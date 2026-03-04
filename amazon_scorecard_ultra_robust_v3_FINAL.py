"""
Amazon Quality Scorecard Engine
================================
Sistema de procesamiento y análisis de métricas de calidad para conductores Amazon.
Soporta PostgreSQL y SQLite con auto-migraciones y validaciones robustas.

Versión: 3.0
Fecha: Febrero 2026
"""

import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule
import warnings
import logging
from typing import Dict, List, Tuple, Optional
import re
import os
import sqlite3
import io
HAS_POSTGRES = False
try:
    import psycopg2
    from psycopg2 import sql
    HAS_POSTGRES = True
except Exception:
    pass

HAS_BCRYPT = False
try:
    import bcrypt
    HAS_BCRYPT = True
except Exception:
    pass

HAS_PDFPLUMBER = False
try:
    import pdfplumber
    HAS_PDFPLUMBER = True
except Exception:
    pass

from datetime import datetime, timedelta

# Configurar logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

warnings.filterwarnings('ignore')

class Config:
    """Configuración centralizada del sistema"""
    
    # Límites de validación
    MAX_DNR = 500  # Aumentado para evitar caps en datos históricos acumulados
    MAX_FALSE_SCAN = 2000  # Aumentado
    MAX_CONDUCTORES = 5000  # Máximo conductores esperados
    
    # Nombres de archivos esperados (patrones)
    PATTERN_CONCESSIONS = r'.*concessions.*\.(csv|xlsx)' # Incluye DSC-concessions
    PATTERN_QUALITY = r'.*quality.*overview.*\.(csv|xlsx)'
    PATTERN_FALSE_SCAN = r'.*false.*scan.*\.html'
    PATTERN_DWC = r'.*(dwc|iadc).*\.(csv|html)'
    PATTERN_FDPS = r'.*fdps.*\.(xlsx|csv)'
    PATTERN_DAILY = r'.*daily.*report.*\.html'
    PATTERN_OFFICIAL_SCORECARD = r'.*scorecard.*3\.0.*\.pdf'
    
    # Columnas requeridas por archivo
    REQUIRED_CONCESSIONS = ['Nombre del agente de entrega', 'ID de agente de entrega', 
                           'Paquetes entregados no recibidos (DNR)']
    REQUIRED_QUALITY = ['ID del transportista', 'DCR']
    
    # Valores por defecto
    DEFAULT_DNR = 0
    DEFAULT_FS = 0
    DEFAULT_DCR = 1.0
    DEFAULT_POD = 1.0
    DEFAULT_CC = 1.0
    DEFAULT_FDPS = 1.0
    DEFAULT_RTS = 0.0
    DEFAULT_CDF = 1.0

def safe_number(val, default=0.0) -> float:
    """Convierte a número de forma segura manejando comas y strings"""
    try:
        if pd.isna(val) or val == '' or val == '-':
            return default
        s = str(val).replace(',', '.').replace('%', '').strip()
        return float(s)
    except (ValueError, TypeError):
        return default

def safe_percentage(val) -> float:
    """Convierte porcentaje string a float de forma segura (0-1.0) con validación de rango"""
    try:
        if pd.isna(val) or val == "" or val == "-":
            return 1.0  # Default a 100% si no hay dato
        
        # Detectar si tiene el símbolo % para ser más precisos
        has_percent = "%" in str(val)
        
        # Normalizar: eliminar %, espacios, y unificar separador decimal
        s = str(val).replace("%", "").strip()
        
        # Amazon a veces usa puntos para miles y comas para decimales (EU)
        if "," in s and "." in s:
            if s.find(",") > s.find("."):
                s = s.replace(".", "").replace(",", ".") # EU: 1.234,56 -> 1234.56
            else:
                s = s.replace(",", "") # US: 1,234.56 -> 1234.56
        elif "," in s:
            s = s.replace(",", ".")
            
        num = float(s)
        
        # Si explícitamente tiene %, siempre dividimos por 100
        if has_percent:
            return max(0.0, min(num / 100.0, 1.0))
        
        # Si el número es gigante (ej: 9494), probablemente es un ID capturado por error
        if num > 100.0:
            return 1.0
            
        if num > 1.0:
            return num / 100.0
            
        # Asegurar rango 0-1.0
        return max(0.0, min(num, 1.0))
    except (ValueError, TypeError):
        return 1.0

def clean_id(val) -> str:
    """Limpia y normaliza IDs de forma segura"""
    if pd.isna(val) or val == '':
        return "UNKNOWN"
    return str(val).strip().upper()

def truncate_sheet_name(name: str, max_length: int = 31) -> str:
    """Trunca nombre de hoja Excel a máximo permitido"""
    # Eliminar caracteres no permitidos en nombres de hoja
    invalid_chars = ['/', '\\', '?', '*', '[', ']', ':']
    clean_name = name
    for char in invalid_chars:
        clean_name = clean_name.replace(char, '_')
    
    # Truncar si es necesario
    if len(clean_name) > max_length:
        clean_name = clean_name[:max_length]
    
    return clean_name

def validate_dataframe(df: pd.DataFrame, required_cols: List[str], 
                      name: str) -> Tuple[bool, str]:
    """Valida que un DataFrame tenga las columnas requeridas"""
    if df is None or df.empty:
        return False, f"{name} está vacío"
    
    missing_cols = [col for col in required_cols if col not in df.columns]
    
    if missing_cols:
        return False, f"{name} le faltan columnas: {', '.join(missing_cols)}"
    
    return True, "OK"

import io

def read_csv_safe(filepath_or_buffer, encoding: str = 'utf-8') -> Optional[pd.DataFrame]:
    """Lee CSV con manejo robusto de encoding (soporta paths o buffers)"""
    encodings = [encoding, 'utf-8', 'latin-1', 'iso-8859-1', 'cp1252']
    
    for enc in encodings:
        try:
            # Si es un buffer, resetear posición
            if hasattr(filepath_or_buffer, 'seek'):
                filepath_or_buffer.seek(0)
                
            df = pd.read_csv(filepath_or_buffer, encoding=enc)
            return df
        except UnicodeDecodeError:
            continue
        except Exception as e:
            logger.error(f"Error leyendo CSV: {str(e)}")
            return None
    return None

def read_any_safe(filepath_or_buffer, filename: str = "") -> Optional[pd.DataFrame]:
    """Detecta extensión y usa el lector adecuado"""
    name = filename.lower() or (filepath_or_buffer.name.lower() if hasattr(filepath_or_buffer, 'name') else "")
    
    if name.endswith('.csv'):
        return read_csv_safe(filepath_or_buffer)
    elif name.endswith('.xlsx') or name.endswith('.xls'):
        return read_excel_safe(filepath_or_buffer)
    elif name.endswith('.html'):
        return read_html_safe(filepath_or_buffer)
    
    # Si no tiene extensión clara, probar CSV por defecto
    return read_csv_safe(filepath_or_buffer)

def read_html_safe(filepath_or_buffer) -> Optional[pd.DataFrame]:
    """Lee HTML y extrae tabla de False Scan o DWC (soporta paths o buffers)"""
    try:
        if hasattr(filepath_or_buffer, 'read'):
            content = filepath_or_buffer.read()
            if isinstance(content, bytes):
                content = content.decode('utf-8', errors='ignore')
        else:
            with open(filepath_or_buffer, 'r', encoding='utf-8', errors='ignore') as f:
                content = f.read()
        
        dfs = pd.read_html(io.StringIO(content))
        
        # Buscar tabla con datos
        for df in dfs:
            if df.empty: continue
            
            # Aplanar multi-index si existe (Amazon usa mucho esto)
            if isinstance(df.columns, pd.MultiIndex):
                # Intentamos unir los niveles o quedarnos con el más descriptivo
                new_cols = []
                for col_tuple in df.columns.values:
                    # Filtrar 'Unnamed' y unir
                    parts = [str(p) for p in col_tuple if 'Unnamed' not in str(p)]
                    if not parts: # Si todo era Unnamed, coger el último
                        new_cols.append(str(col_tuple[-1]))
                    else:
                        new_cols.append(" ".join(parts))
                df.columns = new_cols
            
            df.columns = [str(c).strip() for c in df.columns]
            cols_lower = [str(c).lower() for c in df.columns]
            
            # Caso 1: False Scan
            if any('transporter id' in c or 'driver id' in c for c in cols_lower) and \
               any('false scan' in c for c in cols_lower):
                return df
            
            # Caso 2: DWC/IADC (Nuevo formato o antiguo)
            if any('transporter id' in c or 'driver id' in c for c in cols_lower) and \
               (any('dwc' in c for c in cols_lower) or any('iadc' in c for c in cols_lower)):
                return df
            
            # Caso 3: Daily Report
            if any('transporter id' in c or 'driver id' in c for c in cols_lower) and \
               any('rts' in c or 'dnr' in c for c in cols_lower):
                return df
                
        return None
    except Exception as e:
        logger.error(f"Error leyendo HTML: {str(e)}")
        return None

def read_excel_safe(filepath_or_buffer) -> Optional[pd.DataFrame]:
    """Lee Excel buscando header dinámicamente en todas las hojas (soporta paths o buffers)"""
    try:
        if hasattr(filepath_or_buffer, 'seek'):
            filepath_or_buffer.seek(0)
            
        xl = pd.ExcelFile(filepath_or_buffer, engine='openpyxl')
        sheets = xl.sheet_names
        
        # Priorizar hojas que suelen tener los datos resumidos por transportista
        priority_sheets = ['DNR by Transporter ID', 'DSC by Transporter ID', 'DNR Concessions', 'Sheet1', 'Feuille1', 'Hoja1']
        sorted_sheets = [s for s in priority_sheets if s in sheets] + [s for s in sheets if s not in priority_sheets]
        
        for sheet in sorted_sheets:
            for skip in range(25):
                try:
                    df = pd.read_excel(xl, sheet_name=sheet, skiprows=skip)
                    if df is None or df.empty: continue
                    
                    df.columns = [str(c).strip() for c in df.columns]
                    cols = [c.lower() for c in df.columns]
                    unnamed_count = sum(1 for c in cols if 'unnamed' in c)
                    
                    # Si hay demasiadas columnas Unnamed, probablemente no es el header real
                    if unnamed_count > len(cols) * 0.5 and len(cols) > 2:
                        continue

                    # Para aceptar una hoja, necesitamos al menos una columna de ID y algo de métricas
                    has_id = any(('id' in c or 'agente' in c or 'driver' in c or 'transporter' in c) and 'by' not in c and 'report' not in c for c in cols)
                    has_metrics = any(k in col for k in ['dnr', 'concessions', 'delivered', 'rts', 'quality', 'score'] for col in cols)
                    
                    if has_id and has_metrics:
                        logger.info(f"✓ Header detectado en hoja '{sheet}', fila {skip}")
                        return df
                except:
                    continue
        return None
    except Exception as e:
        logger.error(f"Error leyendo Excel: {str(e)}")
        return None

def find_file(pattern: str, search_path: str = ".") -> Optional[str]:
    """Busca un archivo recursivamente usando un patrón regex"""
    for root, dirs, files in os.walk(search_path):
        for file in files:
            if re.match(pattern, file, re.IGNORECASE):
                return os.path.join(root, file)
    return None

def process_fdps(df: pd.DataFrame) -> pd.DataFrame:
    """Procesa archivo FDPS con validaciones"""
    if df is None or df.empty:
        return pd.DataFrame(columns=['ID', 'FDPS'])
    
    logger.info("Procesando FDPS...")
    
    # Buscar columna de ID y FDPS
    id_col = None
    fdps_col = None
    
    for col in df.columns:
        col_str = str(col).lower()
        if 'id' in col_str and ('transporter' in col_str or 'driver' in col_str or 'agente' in col_str):
            id_col = col
        if 'fdps' in col_str and 'share' not in col_str:
            fdps_col = col

    if id_col is not None and fdps_col is not None:
        df = df.rename(columns={id_col: 'ID', fdps_col: 'FDPS'})
        df['ID'] = df['ID'].apply(clean_id)
        df['FDPS'] = df['FDPS'].apply(safe_percentage)
        logger.info(f"✓ FDPS procesado: {len(df)} conductores")
        return df[['ID', 'FDPS']]
    
    logger.warning("No se encontraron columnas de ID o FDPS en el reporte FDPS")
    return pd.DataFrame(columns=['ID', 'FDPS'])

def process_concessions(df: pd.DataFrame) -> pd.DataFrame:
    """Procesa archivo Concessions con validaciones robustas de columnas"""
    if df is None or df.empty:
        return pd.DataFrame(columns=['ID', 'Nombre', 'DNR', 'RTS', 'Entregados'])
    
    logger.info("Procesando Concessions...")
    
    # Mapeo mucho más estricto para evitar confundir IDs con métricas
    mapping = {}
    
    # 1. Identificar las mejores columnas para cada métrica
    best_cols = {}
    
    for col in df.columns:
        c = str(col).lower()
        
        # ID
        if "id" in c and any(k in c for k in ["agente", "agent", "transporter", "transportista", "driver"]):
            best_cols["ID"] = col
        # Nombre
        elif any(k in c for k in ["nombre", "name"]) and "agente" in c:
            best_cols["Nombre"] = col
        # DNR (MEJORADO: Detectar columnas de semana específica)
        elif (any(k in c for k in ["dnr", "no recibidos"]) or (c == "concessions")) and \
             not re.search(r'\bid\b', c) and "tracking" not in c:
            
            # NUEVO: Detectar columnas con patrón YYYY-WW_DNR (ej: 2026-05_DNR)
            week_pattern = re.match(r'(\d{4})-(\d{2})_dnr', c)
            
            is_better = "DNR" not in best_cols
            
            if not is_better:
                current_col = str(best_cols["DNR"]).lower()
                current_has_dpmo = "dpmo" in current_col or "%" in current_col
                current_is_total = "total" in current_col
                current_week_pattern = re.match(r'(\d{4})-(\d{2})_dnr', current_col)
                
                new_has_dpmo = "dpmo" in c or "%" in c
                new_is_total = "total" in c
                
                # PRIORIDAD 1: Evitar DPMO y %
                if current_has_dpmo and not new_has_dpmo:
                    is_better = True
                # PRIORIDAD 2: Preferir columna de semana específica sobre Total
                elif not current_has_dpmo and not new_has_dpmo:
                    if week_pattern and current_is_total:
                        # Columna de semana específica es mejor que Total
                        is_better = True
                        logger.info(f"  🎯 Detectada columna de semana específica: {col} (en lugar de Total)")
                    elif week_pattern and current_week_pattern:
                        # Si ambas son semanas, preferir la más reciente (última)
                        current_week = int(current_week_pattern.group(2))
                        new_week = int(week_pattern.group(2))
                        if new_week > current_week:
                            is_better = True
                            logger.info(f"  🎯 Usando semana más reciente: W{new_week:02d} (en lugar de W{current_week:02d})")
                    elif not week_pattern and not current_week_pattern:
                        # Si ninguna es semana específica, preferir "Total"
                        if new_is_total and not current_is_total:
                            is_better = True
            
            if is_better:
                best_cols["DNR"] = col
        # RTS (Preferir % si existe)
        elif "rts" in c or "devueltos" in c:
            if "RTS" not in best_cols or "%" in c:
                best_cols["RTS"] = col
        # Entregados
        elif ("entregados" in c or "delivered" in c) and "no" not in c:
            is_better = "Entregados" not in best_cols
            if not is_better:
                if "total" in c and "total" not in str(best_cols["Entregados"]).lower():
                    is_better = True
            if is_better:
                best_cols["Entregados"] = col
            
    # Crear el mapping inverso para rename
    mapping = {v: k for k, v in best_cols.items()}
    df = df.rename(columns=mapping)
    
    # Asegurar columnas mínimas
    for col in ['ID', 'Nombre', 'DNR', 'RTS', 'Entregados']:
        if col not in df.columns:
            if col == 'ID':
                logger.error("No se pudo identificar la columna de ID en Concessions")
                return pd.DataFrame(columns=['ID', 'Nombre', 'DNR', 'RTS', 'Entregados'])
            elif col == 'Nombre':
                df['Nombre'] = df['ID']
            elif col == 'DNR': df['DNR'] = 0
            elif col == 'RTS': df['RTS'] = 0.0
            elif col == 'Entregados': df['Entregados'] = 0
    
    df['ID'] = df['ID'].apply(clean_id)
    df = df[df['ID'] != 'UNKNOWN']
    
    # Validar tipos
    df['DNR'] = df['DNR'].apply(lambda x: safe_number(x, 0))
    df['Entregados'] = df['Entregados'].apply(lambda x: safe_number(x, 0))
    df['RTS'] = df['RTS'].apply(safe_percentage)
    
    # --- LÓGICA DE AGRUPACIÓN INTELIGENTE CON ANTI-DUPLICACIÓN (CORREGIDA) ---
    logger.info(f"  -> Procesando {len(df)} registros de Concessions...")
    
    # ⭐ PARCHE CRÍTICO: Eliminar duplicados ANTES de agrupar
    # Esto previene que archivos duplicados multipliquen los DNRs
    original_count = len(df)
    df = df.drop_duplicates(subset=['ID'], keep='first')
    duplicates_removed = original_count - len(df)
    
    if duplicates_removed > 0:
        logger.warning(f"  ⚠️ Se eliminaron {duplicates_removed} registros duplicados del mismo conductor")
    
    logger.info(f"  -> {len(df)} conductores únicos encontrados")
    
    # Ahora agrupamos por ID (esto es útil solo si hay datos diarios)
    # Si los datos ya vienen consolidados semanalmente, la agrupación no cambia nada
    df_agg = df.groupby('ID').agg({
        'Nombre': 'first',
        'DNR': 'sum',
        'RTS': 'mean',
        'Entregados': 'sum'
    }).reset_index()
    
    # Cap de seguridad: limitar DNR al máximo configurado
    # Nota: Con el drop_duplicates previo, ya no necesitamos lógica compleja
    df_agg['DNR'] = df_agg['DNR'].apply(lambda x: min(x, Config.MAX_DNR))
    
    logger.info(f"✓ Concessions procesado: {len(df_agg)} conductores")
    return df_agg[['ID', 'Nombre', 'DNR', 'RTS', 'Entregados']]

def process_quality(df: pd.DataFrame) -> pd.DataFrame:
    """Procesa archivo Quality Overview con validaciones robustas"""
    if df is None or df.empty:
        return pd.DataFrame(columns=['ID', 'DCR', 'POD', 'CC', 'CDF'])
    
    logger.info("Procesando Quality Overview...")
    
    # Mapeo flexible
    mapping = {}
    found_targets = set()
    for col in df.columns:
        c = str(col).lower()
        target = None
        if 'id' in c and ('transportista' in c or 'transporter' in c or 'driver' in c):
            target = 'ID'
        elif 'dcr' in c: target = 'DCR'
        elif 'pod' in c: target = 'POD'
        elif 'cc' in c: target = 'CC'
        elif 'cdf' in c: target = 'CDF'
            
        if target and target not in found_targets:
            mapping[col] = target
            found_targets.add(target)
            
    df = df.rename(columns=mapping)
    
    # Limpiar IDs
    if 'ID' in df.columns:
        df['ID'] = df['ID'].apply(clean_id)
    else:
        return pd.DataFrame(columns=['ID', 'DCR', 'POD', 'CC', 'CDF'])
    
    # Validar métricas
    for col in ['DCR', 'POD', 'CC', 'CDF']:
        if col in df.columns:
            df[col] = df[col].apply(safe_percentage)
        else:
            df[col] = getattr(Config, f'DEFAULT_{col}')
    
    logger.info(f"✓ Quality Overview procesado: {len(df)} conductores")
    return df[['ID', 'DCR', 'POD', 'CC', 'CDF']]

def process_false_scan(df: pd.DataFrame) -> pd.DataFrame:
    """Procesa archivo False Scan HTML con validaciones"""
    if df is None or df.empty:
        return pd.DataFrame(columns=['ID', 'FS_Count'])
    
    logger.info("Procesando False Scan...")
    
    # Aplanar multi-index si existe
    if isinstance(df.columns, pd.MultiIndex):
        df.columns = df.columns.get_level_values(-1)
    
    # Buscar columna de ID
    id_col = None
    for col in df.columns:
        if 'transporter' in str(col).lower() or 'driver' in str(col).lower():
            if 'id' in str(col).lower():
                id_col = col
                break
    
    if id_col is None:
        logger.error("No se encontró columna de ID en False Scan")
        return pd.DataFrame(columns=['ID', 'FS_Count'])
    
    # Buscar columna de False Scan count
    fs_col = None
    for col in df.columns:
        if 'false scan' in str(col).lower() and 'share' not in str(col).lower():
            fs_col = col
            break
    
    if fs_col is None:
        logger.error("No se encontró columna de False Scan count")
        return pd.DataFrame(columns=['ID', 'FS_Count'])
    
    # Renombrar
    df = df.rename(columns={
        id_col: 'ID',
        fs_col: 'FS_Count'
    })
    
    # Limpiar
    df['ID'] = df['ID'].apply(clean_id)
    df['FS_Count'] = df['FS_Count'].fillna(0).astype(int)
    
    # Detectar valores extremos
    extreme_fs = df[df['FS_Count'] > Config.MAX_FALSE_SCAN]
    if len(extreme_fs) > 0:
        logger.warning(f"⚠️ {len(extreme_fs)} conductores con FS > {Config.MAX_FALSE_SCAN} detectados")
    
    logger.info(f"✓ False Scan procesado: {len(df)} conductores")
    return df[['ID', 'FS_Count']]

def process_dwc(df: pd.DataFrame) -> pd.DataFrame:
    """Procesa archivo DWC/IADC con validaciones"""
    logger.info("Procesando DWC/IADC...")
    
    # 1. Detectar ID del transportista
    id_col = None
    for col in df.columns:
        if 'transporter id' in str(col).lower() or 'driver id' in str(col).lower():
            id_col = col
            break
            
    if id_col is None:
        logger.error("No se encontró columna de ID en DWC/IADC")
        return pd.DataFrame(columns=['ID', 'DNR_RISK_EVENTS', 'CC_DWC', 'IADC'])

    df = df.rename(columns={id_col: 'ID'})
    df['ID'] = df['ID'].apply(clean_id)
    
    # 2. Buscar eventos de riesgo DNR (Formato antiguo)
    dnr_risk_events = pd.Series(0.0, index=df.index)
    if 'Type' in df.columns and 'Total' in df.columns:
        # Filtrar solo filas con DNR Risk
        dnr_mask = df['Type'].str.contains('DNR Risk', case=False, na=False)
        # Agrupar por ID (si el archivo viene en filas por evento)
        risk_group = df[dnr_mask].groupby('ID')['Total'].sum()
    else:
        # Formato nuevo: El riesgo suele estar en columnas específicas
        for col in df.columns:
            c_lower = str(col).lower()
            if 'dnr risk' in c_lower and 'total' in c_lower:
                series = df[col]
                # Si col es duplicado, series será un DataFrame
                if isinstance(series, pd.DataFrame):
                    series = series.iloc[:, 0] # Tomar solo la primera
                
                vals = series.apply(lambda x: safe_number(x, 0)).fillna(0)
                dnr_risk_events = dnr_risk_events + vals
    
    # 3. Buscar DWC% e IADC% (Formato nuevo)
    dwc_col = None
    iadc_col = None
    for col in df.columns:
        c = str(col).lower()
        if 'dwc' in c and '%' in c: dwc_col = col
        if 'iadc' in c and '%' in c: iadc_col = col
        
    res = pd.DataFrame({'ID': df['ID']})
    
    if 'risk_group' in locals():
        res = res.merge(risk_group.reset_index().rename(columns={'Total': 'DNR_RISK_EVENTS'}), on='ID', how='left')
    else:
        res['DNR_RISK_EVENTS'] = dnr_risk_events
        
    if dwc_col:
        series = df[dwc_col]
        if isinstance(series, pd.DataFrame): series = series.iloc[:, 0]
        res['CC_DWC'] = series.apply(safe_percentage)
    if iadc_col:
        series = df[iadc_col]
        if isinstance(series, pd.DataFrame): series = series.iloc[:, 0]
        res['IADC'] = series.apply(safe_percentage)
        
    res = res.fillna(0)
    
    # 4. Agrupar por ID de forma segura (evitando KeyError si faltan columnas)
    agg_dict = {"DNR_RISK_EVENTS": "sum"}
    if "CC_DWC" in res.columns:
        agg_dict["CC_DWC"] = "max"
    if "IADC" in res.columns:
        agg_dict["IADC"] = "max"
        
    res = res.groupby("ID").agg(agg_dict).reset_index()
    
    logger.info(f"✓ DWC/IADC procesado: {len(res)} conductores")
    return res

def process_daily_report(df: pd.DataFrame) -> pd.DataFrame:
    """Procesa reportes diarios para agregar datos a la semana"""
    logger.info("Procesando Daily Report...")
    
    mapping = {}
    found_targets = set()
    for col in df.columns:
        c = str(col).lower()
        target = None
        if 'id' in c: target = 'ID'
        elif 'dnr' in c and 'id' not in c and 'tracking' not in c: target = 'DNR'
        elif 'rts' in c: target = 'RTS'
        elif 'pod' in c and 'fail' in c: target = 'POD_Fails'
        elif 'cc' in c and 'fail' in c: target = 'CC_Fails'
        elif ('delivered' in c or 'entregados' in c) and 'not' not in c: target = 'Entregados'
        
        if target and target not in found_targets:
            mapping[col] = target
            found_targets.add(target)
            
    df = df.rename(columns=mapping)
    
    if 'ID' not in df.columns:
        return pd.DataFrame()
        
    df['ID'] = df['ID'].apply(clean_id)
    
    # Asegurar columnas numéricas
    for col in ['DNR', 'Entregados', 'POD_Fails', 'CC_Fails']:
        if col in df.columns:
            df[col] = df[col].apply(lambda x: safe_number(x, 0))
        else:
            df[col] = 0
            
    if 'RTS' in df.columns:
        df['RTS'] = df['RTS'].apply(safe_percentage)
    else:
        df['RTS'] = 0.0
        
    return df[['ID', 'DNR', 'RTS', 'Entregados', 'POD_Fails', 'CC_Fails']]

def merge_data_smart(df_concessions: pd.DataFrame,
                     df_quality: Optional[pd.DataFrame],
                     df_false_scan: Optional[pd.DataFrame],
                     df_dwc: Optional[pd.DataFrame],
                     df_fdps: Optional[pd.DataFrame] = None,
                     df_daily: Optional[pd.DataFrame] = None) -> pd.DataFrame:
    """
    Merge inteligente de todos los datasets con manejo de missing data
    """
    logger.info("Iniciando merge inteligente de datos...")
    
    # Base: Concessions (tiene nombres + DNR correcto)
    df_base = df_concessions.copy()
    total_base = len(df_base)
    logger.info(f"  Base (Concessions): {total_base} conductores")
    
    # Merge con Quality Overview (métricas)
    if df_quality is not None and not df_quality.empty:
        before = len(df_base)
        df_base = df_base.merge(df_quality, on='ID', how='left')
        matched = df_base['DCR'].notna().sum() if 'DCR' in df_base.columns else 0
        logger.info(f"  + Quality: {matched}/{before} conductores emparejados")
    
    # Merge con Daily Report (Complemento si falta Quality o para mayor detalle)
    if df_daily is not None and not df_daily.empty:
        # Calcular RTS pesado por entregados por día para que la media sea real
        df_daily['rts_count'] = df_daily['RTS'] * df_daily['Entregados']
        
        daily_agg = df_daily.groupby('ID').agg({
            'DNR': 'sum',
            'rts_count': 'sum',
            'Entregados': 'sum',
            'POD_Fails': 'sum',
            'CC_Fails': 'sum'
        }).reset_index()
        
        # Recalcular RTS real (total devueltos / total entregados)
        daily_agg['RTS'] = daily_agg.apply(
            lambda x: x['rts_count'] / x['Entregados'] if x['Entregados'] > 0 else 0.0, axis=1
        )
        
        df_base = df_base.merge(daily_agg, on='ID', how='left', suffixes=('', '_daily'))
        
        matched = df_base['Entregados_daily'].notna().sum()
        logger.info(f"  + Daily Report: {matched}/{len(df_base)} conductores emparejados")
        
        # Complementar métricas si faltan o son 0
        df_base['DNR'] = df_base.apply(
            lambda x: max(x['DNR'], x.get('DNR_daily', 0)) if not pd.isna(x.get('DNR_daily')) else x['DNR'], axis=1
        )
        
        # Corregir Entregados si en concessions era 0 pero en daily hay datos
        if 'Entregados_daily' in df_base.columns:
            df_base['Entregados'] = df_base.apply(
                lambda x: x['Entregados_daily'] if (pd.isna(x['Entregados']) or x['Entregados'] == 0) else x['Entregados'], axis=1
            )
        
        if 'RTS_daily' in df_base.columns:
            df_base['RTS'] = df_base.apply(
                lambda x: x['RTS_daily'] if pd.isna(x['RTS']) or x['RTS'] == 0 else x['RTS'], axis=1
            )
        
        # Calcular POD/CC aproximado si no vienen en Quality pero hay daily
        if 'Entregados_daily' in df_base.columns:
            if 'POD' not in df_base.columns: df_base['POD'] = np.nan
            if 'CC' not in df_base.columns: df_base['CC'] = np.nan
            
            def calc_metric(entregados, fails):
                if pd.isna(entregados) or entregados <= 0: return np.nan
                return max(0.0, min(1.0, 1.0 - (fails / entregados)))

            df_base['POD'] = df_base.apply(
                lambda x: calc_metric(x['Entregados_daily'], x['POD_Fails']) if pd.isna(x.get('POD')) or x.get('POD') == 1.0 else x['POD'], axis=1
            )
            df_base['CC'] = df_base.apply(
                lambda x: calc_metric(x['Entregados_daily'], x['CC_Fails']) if pd.isna(x.get('CC')) or x.get('CC') == 1.0 else x['CC'], axis=1
            )
            
    # Rellenar NaN con valores por defecto para métricas core
    for col in ['DCR', 'POD', 'CC', 'CDF', 'FDPS', 'RTS', 'IADC']:
        if col in df_base.columns:
            # Asegurar que son porcentajes 0-1
            df_base[col] = df_base[col].apply(lambda x: max(0.0, min(float(x), 1.0)) if not pd.isna(x) else np.nan)
            default_val = getattr(Config, f'DEFAULT_{col}', 1.0 if col != 'RTS' else 0.0)
            df_base[col] = df_base[col].fillna(default_val)
        else:
            df_base[col] = getattr(Config, f'DEFAULT_{col}', 1.0 if col != 'RTS' else 0.0)
    
    # Cap final de DNR
    df_base['DNR'] = df_base['DNR'].apply(lambda x: min(x, Config.MAX_DNR))
    
    # Merge con False Scan
    if df_false_scan is not None and not df_false_scan.empty:
        before = len(df_base)
        df_base = df_base.merge(df_false_scan, on='ID', how='left')
        if 'FS_Count' in df_base.columns:
            matched = df_base['FS_Count'].notna().sum()
            logger.info(f"  + False Scan: {matched}/{before} conductores emparejados")
            df_base['FS_Count'] = df_base['FS_Count'].fillna(Config.DEFAULT_FS).astype(int)
        else:
            df_base['FS_Count'] = Config.DEFAULT_FS
    else:
        df_base['FS_Count'] = Config.DEFAULT_FS
    
    # Merge con DWC/IADC (eventos de riesgo + porcentajes)
    if df_dwc is not None and not df_dwc.empty:
        before = len(df_base)
        df_base = df_base.merge(df_dwc, on='ID', how='left')
        
        # Si tenemos CC_DWC de este archivo, podemos usarlo para mejorar el CC
        if 'CC_DWC' in df_base.columns:
            df_base['CC'] = df_base['CC_DWC'].fillna(df_base['CC'])
            
        if 'DNR_RISK_EVENTS' in df_base.columns:
            matched = df_base['DNR_RISK_EVENTS'].notna().sum()
            logger.info(f"  + DWC/IADC: {matched}/{before} conductores emparejados")
            df_base['DNR_RISK_EVENTS'] = df_base['DNR_RISK_EVENTS'].fillna(0).astype(int)
        else:
            df_base['DNR_RISK_EVENTS'] = 0
    else:
        df_base['DNR_RISK_EVENTS'] = 0
        df_base['IADC'] = 0.0
    
    # Merge con FDPS
    if df_fdps is not None and not df_fdps.empty:
        before = len(df_base)
        df_base = df_base.merge(df_fdps, on='ID', how='left')
        if 'FDPS' in df_base.columns:
            df_base['FDPS'] = df_base['FDPS'].fillna(Config.DEFAULT_FDPS)
        else:
            df_base['FDPS'] = Config.DEFAULT_FDPS
    else:
        df_base['FDPS'] = Config.DEFAULT_FDPS
    
    # Validar resultado final
    logger.info(f"✓ Merge completado: {len(df_base)} conductores en dataset final")
    
    return df_base

# ═══════════════════════════════════════════════════════════════
# SISTEMA DE SCORING CON VALIDACIONES
# ═══════════════════════════════════════════════════════════════

def calculate_score_v3_robust(row: pd.Series, targets: Optional[Dict] = None) -> Tuple[str, str, int]:
    """
    Sistema de scoring V3 con validaciones robustas
    Usa targets dinámicos si se proporcionan
    """
    # Targets por defecto si no hay externos
    t = {
        'target_dnr': 0.5, 'target_dcr': 0.995, 'target_pod': 0.99,
        'target_cc': 0.99, 'target_fdps': 0.98, 'target_rts': 0.01, 'target_cdf': 0.95
    }
    if targets:
        t.update(targets)
    
    # Obtener valores con validación
    dnr = safe_number(row.get('DNR', 0), default=0)
    fs_count = safe_number(row.get('FS_Count', 0), default=0)
    dnr_risk = safe_number(row.get('DNR_RISK_EVENTS', 0), default=0)
    dcr = safe_number(row.get('DCR', 1.0), default=1.0)
    pod = safe_number(row.get('POD', 1.0), default=1.0)
    cc = safe_number(row.get('CC', 1.0), default=1.0)
    fdps = safe_number(row.get('FDPS', 1.0), default=1.0)
    rts = safe_number(row.get('RTS', 0.0), default=0.0)
    cdf = safe_number(row.get('CDF', 1.0), default=1.0)
    
    # Validar rangos
    dnr = max(0, min(dnr, Config.MAX_DNR))
    fs_count = max(0, min(fs_count, Config.MAX_FALSE_SCAN))
    dcr = max(0, min(dcr, 1.0))
    pod = max(0, min(pod, 1.0))
    cc = max(0, min(cc, 1.0))
    fdps = max(0, min(fdps, 1.0))
    rts = max(0, min(rts, 1.0))
    cdf = max(0, min(cdf, 1.0))
    
    issues = []
    score = 100
    
    # === DNR (peso x3) ===
    if dnr >= 4:
        issues.append(f"🚨 {int(dnr)} DNR (CRÍTICO)")
        score -= 70
    elif dnr >= 3:
        issues.append(f"🚨 {int(dnr)} DNR (MUY GRAVE)")
        score -= 60
    elif dnr > t['target_dnr'] * 3:
        issues.append(f"⚡ {int(dnr)} DNR")
        score -= 25
    elif dnr > t['target_dnr']:
        issues.append(f"⚡ {int(dnr)} DNR")
        score -= 15
    
    # === FALSE SCANS ===
    fs_int = int(fs_count)
    if fs_int >= 100:
        issues.append(f"❌ {fs_int} FS (CRÍTICO)")
        score -= 40
    elif fs_int >= 20:
        issues.append(f"⚠️ {fs_int} FS")
        score -= 20
    elif fs_int >= 5:
        issues.append(f"ℹ️ {fs_int} FS")
        score -= 5
    
    # === DCR ===
    if dcr < t['target_dcr'] - 0.05:
        issues.append(f"📦 DCR Crítico {dcr:.1%}")
        score -= 40
    elif dcr < t['target_dcr']:
        issues.append(f"📦 DCR Bajo {dcr:.1%}")
        score -= 15
    
    # === POD ===
    if pod < t['target_pod'] - 0.10:
        issues.append(f"📸 POD Crítico {pod:.1%}")
        score -= 25
    elif pod < t['target_pod']:
        issues.append(f"📸 POD Bajo {pod:.1%}")
        score -= 10
    
    # === CC ===
    if cc < t['target_cc']:
        issues.append(f"📞 CC Bajo {cc:.1%}")
        score -= 10
    
    # === FDPS ===
    if fdps < t['target_fdps']:
        issues.append(f"🚚 FDPS Bajo {fdps:.1%}")
        score -= 10

    # === RTS ===
    if rts > t['target_rts'] * 2:
        issues.append(f"🔄 RTS Alto {rts:.1%}")
        score -= 15
    elif rts > t['target_rts']:
        issues.append(f"🔄 RTS {rts:.1%}")
        score -= 8
    
    # === CDF ===
    if cdf < t['target_cdf']:
        issues.append(f"⭐ CDF Bajo {cdf:.1%}")
        score -= 15
    
    # Asegurar score mínimo 0
    score = max(0, score)
    
    # Determinar calificación final
    if score >= 90:
        calificacion = "💎 FANTASTIC"
    elif score >= 80:
        calificacion = "🥇 GREAT"
    elif score >= 60:
        calificacion = "⚠️ FAIR"
    else:
        calificacion = "🛑 POOR"
    
    return calificacion, ", ".join(issues) if issues else "Óptimo", score

# ═══════════════════════════════════════════════════════════════
# GENERACIÓN DE EXCEL
# ═══════════════════════════════════════════════════════════════

def extract_info_from_path(path: str) -> Tuple[str, str]:
    """Extrae semana y centro del nombre del archivo con normalización inteligente"""
    if not path:
        return "N/A", "TDSL"
    
    filename = os.path.basename(path)
    
    # 1. Normalizar Semana
    week = "N/A"
    
    # Patrón estándar: W05, Week 5, Semana 05, S5
    week_match = re.search(r'(?:W|Week|Semana|S)[_\s-]*(\d+)', filename, re.IGNORECASE)
    if week_match:
        num = int(week_match.group(1))
        week = f"W{num:02d}"
    else:
        # Intentar buscar por formato de fecha YYYY-MM-DD
        date_match = re.search(r'(\d{4})[_-](\d{1,2})[_-](\d{1,2})', filename)
        if date_match:
            try:
                from datetime import timedelta
                dt = datetime(int(date_match.group(1)), int(date_match.group(2)), int(date_match.group(3)))
                # Shifting 1 day forward to align Sunday (Amazon start) with Monday (ISO start)
                # This makes Sun-Sat belong to the same ISO week
                dt_shifted = dt + timedelta(days=1)
                isoyear, isoweek, isoday = dt_shifted.isocalendar()
                week = f"W{isoweek:02d}"
            except:
                pass
        
        if week == "N/A":
            # Formato 2026-05 (donde 05 es la semana)
            date_week_match = re.search(r'202\d[_-](\d{1,2})', filename)
            if date_week_match:
                num = int(date_week_match.group(1))
                week = f"W{num:02d}"
    
    # 2. Normalizar Centro (ej: DIC1, VLC1, MAD1, DMA3, ES-TDSL-DIC1)
    # Busca 3-4 letras seguidas de un número
    center_match = re.search(r'([A-Z]{3,4}\d)', filename, re.IGNORECASE)
    if center_match:
        center = center_match.group(1).upper()
    else:
        # Buscar en el path si no está en el filename
        center_match_path = re.search(r'([A-Z]{3,4}\d)', path, re.IGNORECASE)
        center = center_match_path.group(1).upper() if center_match_path else "TDSL"
    
    return week, center

def create_professional_excel(df: pd.DataFrame, output_path: str, 
                              center_name: str = "TDSL", week: str = "N/A") -> bool:
    """
    Genera un Scorecard de Excel con diseño de Dashboard profesional optimizado.
    """
    if df is None or df.empty:
        logger.warning("No hay datos para generar el Excel.")
        return False
        
    try:
        logger.info(f"Generando Dashboard Excel para {len(df)} conductores...")
        
        wb = Workbook()
        wb.remove(wb.active)
        
        # --- PALETA DE COLORES AMAZON ---
        AMZ_DARK = '232F3E'
        AMZ_ORANGE = 'FF9900'
        AMZ_BLUE = '146EB4'
        WHITE = 'FFFFFF'
        
        # Estilos Base
        font_header = Font(name='Segoe UI', size=11, bold=True, color=WHITE)
        font_title = Font(name='Segoe UI', size=16, bold=True, color=WHITE)
        fill_header = PatternFill(start_color=AMZ_DARK, end_color=AMZ_DARK, fill_type='solid')
        fill_orange = PatternFill(start_color=AMZ_ORANGE, end_color=AMZ_ORANGE, fill_type='solid')
        align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
        border_thin = Border(
            left=Side(style='thin', color='DDDDDD'),
            right=Side(style='thin', color='DDDDDD'),
            top=Side(style='thin', color='DDDDDD'),
            bottom=Side(style='thin', color='DDDDDD')
        )

        # 1. 📊 HOJA RESUMEN (DASHBOARD)
        ws_summary = wb.create_sheet('📊 DASHBOARD RESUMEN', 0)
        ws_summary.sheet_view.showGridLines = False
        
        # Título Principal (Corregido merge a columna I para cubrir 4 tarjetas)
        ws_summary.merge_cells('B2:I3')
        title_cell = ws_summary['B2']
        title_cell.value = f'WINIW QUALITY SCORECARD - {center_name} - {week}'
        title_cell.font = font_title
        title_cell.fill = fill_header
        title_cell.alignment = align_center
        
        # --- CARDS DE MÉTRICAS CLAVE ---
        calif_counts = df['CALIFICACION'].value_counts()
        total_drv = len(df)
        
        def create_metric_card(ws, start_col, start_row, label, value, subtext="", color="146EB4"):
            fill_color = PatternFill(start_color=color, end_color=color, fill_type='solid')
            
            # Label box
            cell_label = ws.cell(row=start_row, column=start_col)
            cell_label.value = label
            cell_label.font = Font(name='Segoe UI', size=9, bold=True, color=WHITE)
            cell_label.fill = fill_color
            cell_label.alignment = align_center
            
            # Value box
            cell_val = ws.cell(row=start_row+1, column=start_col)
            cell_val.value = value
            cell_val.font = Font(name='Segoe UI', size=18, bold=True)
            cell_val.alignment = align_center
            
            # Subtext box
            cell_sub = ws.cell(row=start_row+2, column=start_col)
            cell_sub.value = subtext
            cell_sub.font = Font(name='Segoe UI', size=8, italic=True, color='666666')
            cell_sub.alignment = align_center

        row_cards = 5
        # ✅ CÁLCULO CORRECTO DE MÉTRICAS (en Python, no con fórmulas Excel)
        # Esto evita el bug de incluir la fila de totales en los promedios
        dnr_total = int(df['DNR'].sum()) if 'DNR' in df.columns else 0
        dnr_promedio = float(df['DNR'].mean()) if 'DNR' in df.columns else 0.0
        fs_total = int(df['FS_Count'].sum()) if 'FS_Count' in df.columns else 0
        fs_promedio = float(df['FS_Count'].mean()) if 'FS_Count' in df.columns else 0.0
        dcr_avg = float(df['DCR'].mean()) if 'DCR' in df.columns else 1.0
        pod_avg = float(df['POD'].mean()) if 'POD' in df.columns else 1.0
        
        metrics_to_show = [
            ("TOTAL CONDUCTORES", total_drv, f"Activos en {center_name}", AMZ_DARK),
            ("DNR PROMEDIO", f"{dnr_promedio:.2f}", f"Total: {dnr_total}", "E53935"),
            ("FALSE SCANS", f"{fs_promedio:.1f}", f"Total: {fs_total}", "FB8C00"),
            ("DCR PROMEDIO", f"{dcr_avg:.2%}", "Calidad Entrega", "43A047")
        ]
        
        for i, (l, v, s, c) in enumerate(metrics_to_show):
            col_pos = 2 + (i*2)
            create_metric_card(ws_summary, col_pos, row_cards, l, v, s, c)
            ws_summary.merge_cells(start_row=row_cards, start_column=col_pos, end_row=row_cards, end_column=col_pos+1)
            ws_summary.merge_cells(start_row=row_cards+1, start_column=col_pos, end_row=row_cards+2, end_column=col_pos+1)

        # --- DISTRIBUCIÓN DE CALIFICACIONES ---
        ws_summary.cell(row=9, column=2, value="ESTADO DE LA FLOTA").font = Font(bold=True, size=11)
        dist_row = 10
        dist_colors = {'💎 FANTASTIC': '4CAF50', '🥇 GREAT': '8BC34A', '⚠️ FAIR': 'FFC107', '🛑 POOR': 'F44336'}
        
        for i, (cat, color) in enumerate(dist_colors.items()):
            count = int(calif_counts.get(cat, 0))
            pct = count / total_drv if total_drv > 0 else 0
            
            curr_r = dist_row + i
            ws_summary.cell(row=curr_r, column=2, value=cat).font = Font(bold=True)
            ws_summary.cell(row=curr_r, column=3, value=count).alignment = Alignment(horizontal='center')
            ws_summary.cell(row=curr_r, column=4, value=pct).number_format = '0%'
            # Barra visual
            ws_summary.cell(row=curr_r, column=5).fill = PatternFill(start_color=color, end_color=color, fill_type='solid')

        # 2. 📋 HOJA DETALLE
        ws_detail = wb.create_sheet('📋 DETALLE COMPLETO', 1)
        ws_detail.merge_cells('A1:O2')
        detail_header = ws_detail['A1']
        detail_header.value = f"REPORTE DETALLADO CONDUCTORES - {center_name} - SEMANA {week}"
        detail_header.font = font_title
        detail_header.fill = fill_header
        detail_header.alignment = align_center
        
        headers = ['CONDUCTOR', 'ID AGENTE', 'STATUS', 'SCORE', 'ENTREGADOS', 'DNR', 'FALSE SCAN',
                   'RIESGO DNR', 'DCR %', 'POD %', 'CC %', 'FDPS %', 'RTS %', 'CDF %', 'OBSERVACIONES']
        
        for col, h in enumerate(headers, 1):
            cell = ws_detail.cell(row=4, column=col, value=h)
            cell.font = font_header
            cell.fill = fill_orange
            cell.alignment = align_center
            cell.border = border_thin
        
        df_sorted = df.sort_values(['SCORE', 'DNR'], ascending=[False, True]).reset_index(drop=True)
        for idx, (_, row_data) in enumerate(df_sorted.iterrows(), start=5):
            ws_detail.cell(row=idx, column=1, value=str(row_data['Nombre']))
            ws_detail.cell(row=idx, column=2, value=str(row_data['ID']))
            
            status = str(row_data['CALIFICACION'])
            c_status = ws_detail.cell(row=idx, column=3, value=status)
            c_status.font = Font(bold=True)
            if '💎' in status: c_status.font = Font(bold=True, color='2E7D32')
            elif '🛑' in status: c_status.font = Font(bold=True, color='C62828')
            
            ws_detail.cell(row=idx, column=4, value=float(row_data['SCORE'])).number_format = '0'
            ws_detail.cell(row=idx, column=5, value=float(row_data['Entregados'])).number_format = '#,##0'
            ws_detail.cell(row=idx, column=6, value=float(row_data['DNR'])).number_format = '0'
            ws_detail.cell(row=idx, column=7, value=float(row_data['FS_Count'])).number_format = '#,##0'
            ws_detail.cell(row=idx, column=8, value=float(row_data['DNR_RISK_EVENTS'])).number_format = '0'
            
            for col_idx, col_name in enumerate(['DCR', 'POD', 'CC', 'FDPS', 'RTS', 'CDF'], start=9):
                cell = ws_detail.cell(row=idx, column=col_idx, value=float(row_data[col_name]))
                cell.number_format = '0.0%'
            
            cell_obs = ws_detail.cell(row=idx, column=15, value=str(row_data['DETALLES']))
            cell_obs.alignment = Alignment(horizontal='left', wrap_text=True, vertical='center')
            
            for col in range(1, 16): 
                ws_detail.cell(row=idx, column=col).border = border_thin

        # Formato Condicional Optimizado
        last_row = len(df_sorted) + 4
        if last_row >= 5:
            for col in ['D', 'I', 'J', 'K', 'L', 'N']:
                ws_detail.conditional_formatting.add(f'{col}5:{col}{last_row}',
                    ColorScaleRule(start_type='num', start_value=0.7, start_color='F8696B',
                                   mid_type='num', mid_value=0.9, mid_color='FFEB84',
                                   end_type='num', end_value=1.0, end_color='63BE7B'))
            for col in ['F', 'G', 'H', 'M']:
                ws_detail.conditional_formatting.add(f'{col}5:{col}{last_row}',
                    ColorScaleRule(start_type='percentile', start_value=0, start_color='63BE7B',
                                   mid_type='percentile', mid_value=85, mid_color='FFEB84',
                                   end_type='percentile', end_value=98, end_color='F8696B'))

        ws_detail.freeze_panes = 'A5'
        ws_detail.auto_filter.ref = f'A4:O{last_row}'
        
        # Ajuste de anchos
        col_widths = {'A': 28, 'B': 14, 'C': 15, 'D': 8, 'E': 12, 'O': 55}
        for col, width in col_widths.items(): ws_detail.column_dimensions[col].width = width
        for col in 'FGHIJKLMN': ws_detail.column_dimensions[col].width = 11

        # 3. 🏆 HOJA RANKING & ALERTAS
        ws_rank = wb.create_sheet('🏆 RANKING & ALERTAS', 2)
        ws_rank.sheet_view.showGridLines = False
        
        # Estilos específicos para rankings
        fill_green = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
        font_green = Font(color='006100', bold=True)
        fill_red = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
        font_red = Font(color='9C0006', bold=True)
        
        # TOP 10 FANTASTIC
        ws_rank.merge_cells('B2:D2')
        ws_rank['B2'] = "⭐ TOP 10 CONDUCTORES (FANTASTIC)"
        ws_rank['B2'].font = Font(name='Segoe UI', size=12, bold=True, color='2E7D32')
        
        headers_rank = ['Nombre', 'Score', 'DNR']
        for col, h in enumerate(headers_rank, 2):
            cell = ws_rank.cell(row=3, column=col, value=h)
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = align_center
            
        top_10 = df_sorted.sort_values(['SCORE', 'DNR'], ascending=[False, True]).head(10)
        for idx, (_, r) in enumerate(top_10.iterrows(), 4):
            ws_rank.cell(row=idx, column=2, value=r['Nombre']).fill = fill_green
            ws_rank.cell(row=idx, column=3, value=r['SCORE']).font = font_green
            ws_rank.cell(row=idx, column=4, value=r['DNR'])
            
        # BOTTOM 10 / RED FLAGS
        ws_rank.merge_cells('F2:H2')
        ws_rank['F2'] = "🚨 RED FLAGS (POOR/FAIR)"
        ws_rank['F2'].font = Font(name='Segoe UI', size=12, bold=True, color='C62828')
        
        for col, h in enumerate(headers_rank, 6):
            cell = ws_rank.cell(row=3, column=col, value=h)
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = align_center
            
        bottom_10 = df_sorted.sort_values(['SCORE', 'DNR'], ascending=[True, False]).head(15)
        for idx, (_, r) in enumerate(bottom_10.iterrows(), 4):
            ws_rank.cell(row=idx, column=6, value=r['Nombre']).fill = fill_red
            ws_rank.cell(row=idx, column=7, value=r['SCORE']).font = font_red
            ws_rank.cell(row=idx, column=8, value=r['DNR'])

        ws_rank.column_dimensions['B'].width = 35
        ws_rank.column_dimensions['F'].width = 35
        
        # 4. 💬 HOJA FEEDBACK (PULIDO)
        ws_feed = wb.create_sheet('💬 FEEDBACK', 3)
        ws_feed.append(['CONDUCTOR', 'CALIFICACIÓN', 'PUNTOS DE MEJORA'])
        for col, h in enumerate(['CONDUCTOR', 'CALIFICACIÓN', 'PUNTOS DE MEJORA'], 1):
            cell = ws_feed.cell(row=1, column=col)
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = align_center
            
        for _, r in df_sorted.iterrows(): 
            ws_feed.append([r['Nombre'], r['CALIFICACION'], r['DETALLES']])
            
        ws_feed.column_dimensions['A'].width = 30
        ws_feed.column_dimensions['B'].width = 15
        ws_feed.column_dimensions['C'].width = 85
        for row_idx in range(2, len(df_sorted) + 2):
            ws_feed.cell(row=row_idx, column=3).alignment = Alignment(wrap_text=True, vertical='center')

        wb.save(output_path)
        logger.info(f"✅ Dashboard Excel Optimizado guardado en: {output_path}")
        return True
    except Exception as e:
        logger.error(f"❌ Error Dashboard Excel: {str(e)}")
        return False

def delete_scorecard_batch(week: str, center: str, db_config: Optional[Dict] = None) -> bool:
    """Elimina los datos de una semana y centro específicos (Para corregir errores de volcado)"""
    try:
        conn = get_db_connection(db_config)
        cursor = conn.cursor()
        
        q = "DELETE FROM scorecards WHERE semana = %s AND centro = %s" if db_config and db_config.get('type') == 'postgresql' else "DELETE FROM scorecards WHERE semana = ? AND centro = ?"
        cursor.execute(q, (week, center))
        
        rows = cursor.rowcount
        conn.commit()
        conn.close()
        
        logger.info(f"🗑️ Se eliminaron {rows} registros previos de {center} para la semana {week}.")
        return True
    except Exception as e:
        logger.error(f"Error eliminando lote: {e}")
        return False

def find_file_in_dir(pattern: str, directory: str) -> Optional[str]:
    """Busca un archivo en un directorio específico usando un patrón"""
    for file in os.listdir(directory):
        if re.match(pattern, file, re.IGNORECASE):
            return os.path.join(directory, file)
    return None

def process_single_batch(path_concessions, path_quality=None, path_false_scan=None, 
                         path_dwc=None, path_fdps=None, path_daily=None, targets=None) -> Optional[pd.DataFrame]:
    """Procesa un único lote de archivos y devuelve el DataFrame final. 
    Soporta rutas individuales o listas de rutas/buffers."""
    try:
        # Función auxiliar para leer uno o varios archivos y concatenarlos
        def read_multiple(paths):
            if paths is None: return None
            if not isinstance(paths, list): paths = [paths]
            
            valid_dfs = []
            for p in paths:
                df = read_any_safe(p, str(p))
                if df is not None and not df.empty:
                    valid_dfs.append(df)
            
            if not valid_dfs: return None
            return pd.concat(valid_dfs, ignore_index=True) if len(valid_dfs) > 1 else valid_dfs[0]

        df_concessions = read_multiple(path_concessions)
        df_quality = read_multiple(path_quality)
        df_false_scan_html = read_multiple(path_false_scan)
        df_dwc = read_multiple(path_dwc)
        df_fdps = read_multiple(path_fdps)
        df_daily_raw = read_multiple(path_daily)
        
        if df_concessions is None or df_concessions.empty:
            logger.error("No se pudo leer el archivo de Concessions (obligatorio)")
            return None
            
        df_conc_clean = process_concessions(df_concessions)
        if df_conc_clean is not None and not df_conc_clean.empty and 'ID' in df_conc_clean.columns:
            if df_conc_clean['ID'].duplicated().any():
                df_conc_clean = df_conc_clean.drop_duplicates(subset='ID', keep='first')
        
        df_qual_clean = process_quality(df_quality) if df_quality is not None and not df_quality.empty else None
        df_fs_clean = process_false_scan(df_false_scan_html) if df_false_scan_html is not None and not df_false_scan_html.empty else None
        df_dwc_clean = process_dwc(df_dwc) if df_dwc is not None and not df_dwc.empty else None
        df_fdps_clean = process_fdps(df_fdps) if df_fdps is not None and not df_fdps.empty else None
        df_daily_clean = process_daily_report(df_daily_raw) if df_daily_raw is not None and not df_daily_raw.empty else None
        
        df_merged = merge_data_smart(df_conc_clean, df_qual_clean, df_fs_clean, df_dwc_clean, df_fdps_clean, df_daily_clean)
        
        # --- LIMPIEZA FINAL DE SEGURIDAD (ANTI-EXPLOSIÓN) ---
        if 'DNR' in df_merged.columns:
            # Forzamos que el DNR nunca supere el límite, evitando IDs de Amazon en el Excel
            df_merged['DNR'] = df_merged['DNR'].apply(lambda x: min(float(x), Config.MAX_DNR) if not pd.isna(x) else 0.0)
        
        if 'FS_Count' in df_merged.columns:
            df_merged['FS_Count'] = df_merged['FS_Count'].apply(lambda x: min(float(x), Config.MAX_FALSE_SCAN) if not pd.isna(x) else 0.0)

        # Eliminar duplicados finales por si acaso
        if 'ID' in df_merged.columns:
            df_merged = df_merged.drop_duplicates(subset='ID', keep='first')
            
        results = df_merged.apply(lambda x: calculate_score_v3_robust(x, targets=targets), axis=1, result_type='expand')
        df_merged['CALIFICACION'] = results[0]
        df_merged['DETALLES'] = results[1]
        df_merged['SCORE'] = results[2]
        
        return df_merged
    except Exception as e:
        logger.error(f"Error en procesamiento: {str(e)}")
        import traceback
        traceback.print_exc()
        return None

# ═══════════════════════════════════════════════════════════════
# BASE DE DATOS (FASE 2/3 - SQLite & PostgreSQL)
# ═══════════════════════════════════════════════════════════════

def get_db_connection(db_config: Optional[Dict] = None):
    """Crea o conecta a la base de datos (PostgreSQL o SQLite)"""
    if db_config and db_config.get('type') == 'postgresql':
        if not HAS_POSTGRES:
            raise ImportError("Librería 'psycopg2' no encontrada. Instálala con: pip install psycopg2-binary")
        
        return psycopg2.connect(
            host=db_config.get('host', 'localhost'),
            database=db_config.get('database', 'postgres'),
            user=db_config.get('user', 'postgres'),
            password=db_config.get('password', ''),
            port=db_config.get('port', 5432)
        )
    else:
        # Por defecto SQLite
        db_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "amazon_quality.db")
        conn = sqlite3.connect(db_path)
        # Habilitar acceso por nombre de columna en sqlite
        conn.row_factory = sqlite3.Row
        return conn

import hashlib

def hash_password(password: str) -> str:
    """
    Encripta la contraseña usando bcrypt (preferido) o SHA-256 (fallback).
    
    Args:
        password: Contraseña en texto plano
        
    Returns:
        Hash de la contraseña
        
    Notes:
        - bcrypt es más seguro (salt automático, resistente a rainbow tables)
        - SHA-256 se usa como fallback si bcrypt no está instalado
    """
    if HAS_BCRYPT:
        # Usar bcrypt con salt automático (más seguro)
        salt = bcrypt.gensalt()
        hashed = bcrypt.hashpw(password.encode('utf-8'), salt)
        # Agregar prefijo para identificar que es bcrypt
        return 'bcrypt:' + hashed.decode('utf-8')
    else:
        # Fallback a SHA-256 (menos seguro pero funcional)
        logger.warning("Usando SHA-256 para hash de contraseña. Considera instalar bcrypt para mejor seguridad.")
        return 'sha256:' + hashlib.sha256(password.encode()).hexdigest()

def verify_password(password: str, hashed: str) -> bool:
    """
    Verifica una contraseña contra su hash.
    
    Args:
        password: Contraseña en texto plano
        hashed: Hash almacenado en base de datos
        
    Returns:
        True si la contraseña coincide, False en caso contrario
        
    Notes:
        - Detecta automáticamente el método de hash usado (bcrypt o SHA-256)
        - Compatible con hashes antiguos (SHA-256) y nuevos (bcrypt)
    """
    try:
        if hashed.startswith('bcrypt:'):
            # Hash bcrypt
            if HAS_BCRYPT:
                hash_only = hashed.replace('bcrypt:', '')
                return bcrypt.checkpw(password.encode('utf-8'), hash_only.encode('utf-8'))
            else:
                logger.error("Hash es bcrypt pero biblioteca no está instalada")
                return False
        elif hashed.startswith('sha256:'):
            # Hash SHA-256
            hash_only = hashed.replace('sha256:', '')
            test_hash = hashlib.sha256(password.encode()).hexdigest()
            return test_hash == hash_only
        else:
            # Compatibilidad con hashes antiguos sin prefijo (asumimos SHA-256)
            test_hash = hashlib.sha256(password.encode()).hexdigest()
            return test_hash == hashed
    except Exception as e:
        logger.error(f"Error verificando contraseña: {e}")
        return False

def init_database(db_config: Optional[Dict] = None):
    """Inicializa las tablas optimizadas para producción en PostgreSQL o SQLite"""
    try:
        conn = get_db_connection(db_config)
        cursor = conn.cursor()
        is_postgres = db_config and db_config.get('type') == 'postgresql'
        
        # 1. Tabla de Scorecards (Optimizada para Power BI)
        if is_postgres:
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS scorecards (
                    id SERIAL PRIMARY KEY,
                    semana VARCHAR(10),
                    fecha_semana DATE,
                    centro VARCHAR(20),
                    driver_id VARCHAR(50),
                    driver_name VARCHAR(255),
                    calificacion VARCHAR(50),
                    score DOUBLE PRECISION,
                    entregados DOUBLE PRECISION,
                    dnr DOUBLE PRECISION,
                    fs_count DOUBLE PRECISION,
                    dnr_risk_events DOUBLE PRECISION,
                    dcr DOUBLE PRECISION,
                    pod DOUBLE PRECISION,
                    cc DOUBLE PRECISION,
                    fdps DOUBLE PRECISION,
                    rts DOUBLE PRECISION,
                    cdf DOUBLE PRECISION,
                    detalles TEXT,
                    uploaded_by VARCHAR(100),
                    timestamp TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    UNIQUE(semana, centro, driver_id)
                )
            ''')
            # Índices PostgreSQL — cubren todos los patrones de acceso habituales
            pg_indexes = [
                # Filtro principal app: centro + semana
                "CREATE INDEX IF NOT EXISTS idx_centro_semana       ON scorecards (centro, semana)",
                # Power BI: covering index con métricas
                "CREATE INDEX IF NOT EXISTS idx_bi_query            ON scorecards (centro, semana, fecha_semana) INCLUDE (score, dnr, dcr, calificacion)",
                # Histórico ordenado por fecha
                "CREATE INDEX IF NOT EXISTS idx_fecha_desc          ON scorecards (fecha_semana DESC)",
                # Búsqueda por conductor
                "CREATE INDEX IF NOT EXISTS idx_driver_id           ON scorecards (driver_id)",
                "CREATE INDEX IF NOT EXISTS idx_driver_name         ON scorecards (LOWER(driver_name))",
                # Filtro por calificación (dashboard alertas)
                "CREATE INDEX IF NOT EXISTS idx_calificacion        ON scorecards (calificacion)",
                # Análisis temporal por centro
                "CREATE INDEX IF NOT EXISTS idx_centro_fecha        ON scorecards (centro, fecha_semana DESC)",
                # Ranking conductores por centro/semana
                "CREATE INDEX IF NOT EXISTS idx_ranking             ON scorecards (centro, semana, score DESC)",
                # Índice parcial: DNR crítico > 5
                "CREATE INDEX IF NOT EXISTS idx_dnr_alto            ON scorecards (centro, semana, dnr) WHERE dnr > 5",
                # Índice parcial: bajo rendimiento
                "CREATE INDEX IF NOT EXISTS idx_poor_fair           ON scorecards (centro, semana, score) WHERE score < 70",
            ]
            for idx_sql in pg_indexes:
                try:
                    cursor.execute(idx_sql)
                except Exception as ie:
                    logger.warning(f"Índice ya existe o error menor: {ie}")
        else:
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS scorecards (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    semana TEXT, fecha_semana DATE, centro TEXT, driver_id TEXT, driver_name TEXT,
                    calificacion TEXT, score FLOAT, entregados FLOAT, dnr FLOAT,
                    fs_count FLOAT, dnr_risk_events FLOAT, dcr FLOAT, pod FLOAT,
                    cc FLOAT, fdps FLOAT, rts FLOAT, cdf FLOAT, detalles TEXT,
                    uploaded_by TEXT,
                    timestamp TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    UNIQUE(semana, centro, driver_id)
                )
            ''')
            # Índices SQLite — sintaxis simplificada (no soporta INCLUDE ni WHERE en todos los casos)
            sqlite_indexes = [
                "CREATE INDEX IF NOT EXISTS idx_centro_semana  ON scorecards (centro, semana)",
                "CREATE INDEX IF NOT EXISTS idx_fecha_desc     ON scorecards (fecha_semana DESC)",
                "CREATE INDEX IF NOT EXISTS idx_driver_id      ON scorecards (driver_id)",
                "CREATE INDEX IF NOT EXISTS idx_driver_name    ON scorecards (driver_name)",
                "CREATE INDEX IF NOT EXISTS idx_calificacion   ON scorecards (calificacion)",
                "CREATE INDEX IF NOT EXISTS idx_centro_fecha   ON scorecards (centro, fecha_semana)",
                "CREATE INDEX IF NOT EXISTS idx_ranking        ON scorecards (centro, semana, score DESC)",
            ]
            for idx_sql in sqlite_indexes:
                try:
                    cursor.execute(idx_sql)
                except Exception as ie:
                    logger.warning(f"Índice SQLite: {ie}")

        # 2. Tabla de Usuarios
        user_id_type = "SERIAL PRIMARY KEY" if is_postgres else "INTEGER PRIMARY KEY AUTOINCREMENT"
        cursor.execute(f'''
            CREATE TABLE IF NOT EXISTS users (
                id {user_id_type},
                username VARCHAR(100) UNIQUE,
                password TEXT,
                role VARCHAR(20),
                active INTEGER DEFAULT 1,
                must_change_password INTEGER DEFAULT 0
            )
        ''')
        # Índices de usuarios (se crean AQUÍ, después de que la tabla existe)
        users_indexes_pg = [
            "CREATE INDEX IF NOT EXISTS idx_users_username ON users (LOWER(username))",
            "CREATE INDEX IF NOT EXISTS idx_users_role     ON users (role) WHERE active = 1",
        ]
        users_indexes_sqlite = [
            "CREATE INDEX IF NOT EXISTS idx_users_username ON users (username)",
            "CREATE INDEX IF NOT EXISTS idx_users_role     ON users (role)",
        ]
        for idx_sql in (users_indexes_pg if is_postgres else users_indexes_sqlite):
            try:
                cursor.execute(idx_sql)
            except Exception as ie:
                logger.warning(f"Índice users: {ie}")

        # 3. Tabla de Targets (Optimizada)
        cursor.execute(f'''
            CREATE TABLE IF NOT EXISTS center_targets (
                centro VARCHAR(50) PRIMARY KEY,
                target_dnr DOUBLE PRECISION DEFAULT 0.5,
                target_dcr DOUBLE PRECISION DEFAULT 0.995,
                target_pod DOUBLE PRECISION DEFAULT 0.99,
                target_cc DOUBLE PRECISION DEFAULT 0.99,
                target_fdps DOUBLE PRECISION DEFAULT 0.98,
                target_rts DOUBLE PRECISION DEFAULT 0.01,
                target_cdf DOUBLE PRECISION DEFAULT 0.95,
                timestamp TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        ''')

        # Auto-Migración: Añadir columnas si no existen (Para bases de datos ya creadas)
        cols_to_add = [
            ("fecha_semana", "DATE" if is_postgres else "DATE"),
            ("uploaded_by", "VARCHAR(100)" if is_postgres else "TEXT")
        ]
        
        for col_name, col_type in cols_to_add:
            try:
                if is_postgres:
                    cursor.execute(f"ALTER TABLE scorecards ADD COLUMN IF NOT EXISTS {col_name} {col_type}")
                else:
                    # SQLite no soporta ADD COLUMN IF NOT EXISTS de forma directa
                    cursor.execute(f"PRAGMA table_info(scorecards)")
                    existing_cols = [c[1] for c in cursor.fetchall()]
                    if col_name not in existing_cols:
                        cursor.execute(f"ALTER TABLE scorecards ADD COLUMN {col_name} {col_type}")
            except Exception as e:
                logger.warning(f"Aviso migración columna {col_name}: {e}")
        
        # Auto-Migración tabla users: Añadir must_change_password
        try:
            if is_postgres:
                cursor.execute("ALTER TABLE users ADD COLUMN IF NOT EXISTS must_change_password INTEGER DEFAULT 0")
            else:
                cursor.execute("PRAGMA table_info(users)")
                existing_user_cols = [c[1] for c in cursor.fetchall()]
                if "must_change_password" not in existing_user_cols:
                    cursor.execute("ALTER TABLE users ADD COLUMN must_change_password INTEGER DEFAULT 0")
        except Exception as e:
            logger.warning(f"Aviso migración columna must_change_password: {e}")

        # ── TABLAS NUEVAS v3.2: station_scorecards y wh_exceptions ────────────
        # 4. station_scorecards
        if is_postgres:
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS station_scorecards (
                    id SERIAL PRIMARY KEY,
                    semana VARCHAR(10) NOT NULL, fecha_semana DATE, centro VARCHAR(20) NOT NULL,
                    overall_score DOUBLE PRECISION, overall_standing VARCHAR(20),
                    rank_station INTEGER, rank_wow INTEGER,
                    safety_tier VARCHAR(20), fico DOUBLE PRECISION, fico_tier VARCHAR(20),
                    speeding_rate DOUBLE PRECISION, speeding_tier VARCHAR(20),
                    mentor_adoption DOUBLE PRECISION, mentor_tier VARCHAR(20),
                    vsa_compliance DOUBLE PRECISION, vsa_tier VARCHAR(20), boc VARCHAR(100),
                    whc_pct DOUBLE PRECISION, whc_tier VARCHAR(20), cas VARCHAR(50),
                    quality_tier VARCHAR(20), dcr_pct DOUBLE PRECISION, dcr_tier VARCHAR(20),
                    dnr_dpmo DOUBLE PRECISION, dnr_tier VARCHAR(20),
                    lor_dpmo DOUBLE PRECISION, lor_tier VARCHAR(20),
                    dsc_dpmo DOUBLE PRECISION, dsc_tier VARCHAR(20),
                    pod_pct DOUBLE PRECISION, pod_tier VARCHAR(20),
                    cc_pct DOUBLE PRECISION, cc_tier VARCHAR(20),
                    ce_dpmo DOUBLE PRECISION, ce_tier VARCHAR(20),
                    cdf_dpmo DOUBLE PRECISION, cdf_tier VARCHAR(20),
                    capacity_tier VARCHAR(20),
                    capacity_next_day DOUBLE PRECISION, capacity_next_day_tier VARCHAR(20),
                    capacity_same_day DOUBLE PRECISION, capacity_same_day_tier VARCHAR(20),
                    focus_area_1 VARCHAR(200), focus_area_2 VARCHAR(200), focus_area_3 VARCHAR(200),
                    uploaded_by VARCHAR(100), timestamp TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    UNIQUE(semana, centro)
                )
            ''')
            for idx in ["CREATE INDEX IF NOT EXISTS idx_ss_centro_semana ON station_scorecards (centro, semana)",
                        "CREATE INDEX IF NOT EXISTS idx_ss_fecha ON station_scorecards (fecha_semana DESC)",
                        "CREATE INDEX IF NOT EXISTS idx_ss_standing ON station_scorecards (overall_standing)"]:
                try: cursor.execute(idx)
                except Exception: pass
        else:
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS station_scorecards (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    semana TEXT NOT NULL, fecha_semana DATE, centro TEXT NOT NULL,
                    overall_score FLOAT, overall_standing TEXT, rank_station INTEGER, rank_wow INTEGER,
                    safety_tier TEXT, fico FLOAT, fico_tier TEXT,
                    speeding_rate FLOAT, speeding_tier TEXT,
                    mentor_adoption FLOAT, mentor_tier TEXT,
                    vsa_compliance FLOAT, vsa_tier TEXT, boc TEXT,
                    whc_pct FLOAT, whc_tier TEXT, cas TEXT,
                    quality_tier TEXT, dcr_pct FLOAT, dcr_tier TEXT,
                    dnr_dpmo FLOAT, dnr_tier TEXT, lor_dpmo FLOAT, lor_tier TEXT,
                    dsc_dpmo FLOAT, dsc_tier TEXT, pod_pct FLOAT, pod_tier TEXT,
                    cc_pct FLOAT, cc_tier TEXT, ce_dpmo FLOAT, ce_tier TEXT,
                    cdf_dpmo FLOAT, cdf_tier TEXT,
                    capacity_tier TEXT, capacity_next_day FLOAT, capacity_next_day_tier TEXT,
                    capacity_same_day FLOAT, capacity_same_day_tier TEXT,
                    focus_area_1 TEXT, focus_area_2 TEXT, focus_area_3 TEXT,
                    uploaded_by TEXT, timestamp TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    UNIQUE(semana, centro)
                )
            ''')
            for idx in ["CREATE INDEX IF NOT EXISTS idx_ss_centro_semana ON station_scorecards (centro, semana)",
                        "CREATE INDEX IF NOT EXISTS idx_ss_fecha ON station_scorecards (fecha_semana DESC)"]:
                try: cursor.execute(idx)
                except Exception: pass

        # 5. wh_exceptions
        wh_pk = "SERIAL PRIMARY KEY" if is_postgres else "INTEGER PRIMARY KEY AUTOINCREMENT"
        wh_str = "VARCHAR(10)" if is_postgres else "TEXT"
        wh_cen = "VARCHAR(20)" if is_postgres else "TEXT"
        wh_did = "VARCHAR(50)" if is_postgres else "TEXT"
        wh_uby = "VARCHAR(100)" if is_postgres else "TEXT"
        cursor.execute(f'''
            CREATE TABLE IF NOT EXISTS wh_exceptions (
                id {wh_pk},
                semana {wh_str} NOT NULL, fecha_semana DATE, centro {wh_cen} NOT NULL,
                driver_id {wh_did} NOT NULL,
                daily_limit_exceeded INTEGER DEFAULT 0,
                weekly_limit_exceeded INTEGER DEFAULT 0,
                under_offwork_limit INTEGER DEFAULT 0,
                workday_limit_exceeded INTEGER DEFAULT 0,
                uploaded_by {wh_uby},
                timestamp TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                UNIQUE(semana, centro, driver_id)
            )
        ''')
        for idx in ["CREATE INDEX IF NOT EXISTS idx_wh_centro_semana ON wh_exceptions (centro, semana)",
                    "CREATE INDEX IF NOT EXISTS idx_wh_driver ON wh_exceptions (driver_id)"]:
            try: cursor.execute(idx)
            except Exception: pass

        # ── MIGRACIÓN v3.2: columnas _oficial en scorecards (solo UPDATE desde PDF) ──
        pdf_cols_to_add = [
            ("entregados_oficial", "DOUBLE PRECISION" if is_postgres else "FLOAT"),
            ("dcr_oficial",        "DOUBLE PRECISION" if is_postgres else "FLOAT"),
            ("pod_oficial",        "DOUBLE PRECISION" if is_postgres else "FLOAT"),
            ("cc_oficial",         "DOUBLE PRECISION" if is_postgres else "FLOAT"),
            ("dsc_dpmo",           "DOUBLE PRECISION" if is_postgres else "FLOAT"),
            ("lor_dpmo",           "DOUBLE PRECISION" if is_postgres else "FLOAT"),
            ("ce_dpmo",            "DOUBLE PRECISION" if is_postgres else "FLOAT"),
            ("cdf_dpmo_oficial",   "DOUBLE PRECISION" if is_postgres else "FLOAT"),
            ("pdf_loaded",         "INTEGER DEFAULT 0"),
        ]
        for col_name, col_type in pdf_cols_to_add:
            try:
                if is_postgres:
                    cursor.execute(f"ALTER TABLE scorecards ADD COLUMN IF NOT EXISTS {col_name} {col_type}")
                else:
                    cursor.execute("PRAGMA table_info(scorecards)")
                    existing = [c[1] for c in cursor.fetchall()]
                    if col_name not in existing:
                        cursor.execute(f"ALTER TABLE scorecards ADD COLUMN {col_name} {col_type}")
            except Exception as e:
                logger.warning(f"Migración v3.2 col {col_name}: {e}")

        # Asegurar que el usuario 'pablo' existe como SUPERADMIN
        q_check = "SELECT id FROM users WHERE LOWER(username) = %s" if is_postgres else "SELECT id FROM users WHERE LOWER(username) = ?"
        cursor.execute(q_check, ("pablo",))
        if not cursor.fetchone():
            admin_pass = hash_password("Admin_Winiw_2026")
            q_ins = "INSERT INTO users (username, password, role, must_change_password) VALUES (%s, %s, %s, %s)" if is_postgres else "INSERT INTO users (username, password, role, must_change_password) VALUES (?, ?, ?, ?)"
            cursor.execute(q_ins, ("pablo", admin_pass, "superadmin", 1))
            logger.info("Usuario 'pablo' creado como SUPERADMIN (requiere cambio de contraseña).")
        else:
            # Actualizar pablo a superadmin si ya existe pero es admin
            q_update = "UPDATE users SET role = %s WHERE LOWER(username) = %s" if is_postgres else "UPDATE users SET role = ? WHERE LOWER(username) = ?"
            cursor.execute(q_update, ("superadmin", "pablo"))
            logger.info("Usuario 'pablo' actualizado a SUPERADMIN.")
        
        conn.commit()
        conn.close()
        return True
    except Exception as e:
        logger.error(f"Error DB: {e}")
        return False

def reset_production_database(db_config: Optional[Dict] = None):
    """Limpia todos los datos de scorecards para empezar de cero (Mantiene usuarios y targets)"""
    try:
        conn = get_db_connection(db_config)
        cursor = conn.cursor()
        cursor.execute("TRUNCATE TABLE scorecards RESTART IDENTITY" if db_config and db_config.get('type') == 'postgresql' else "DELETE FROM scorecards")
        conn.commit()
        conn.close()
        return True
    except Exception as e:
        logger.error(f"Error reset: {e}")
        return False


def week_to_date(week_str: str, year: int = None) -> str:
    """Convierte un string de semana 'W05' a la fecha del lunes de esa semana"""
    try:
        if year is None:
            year = datetime.now().year  # Año dinámico, nunca hardcodeado
        if not week_str or week_str == "N/A":
            return datetime.now().strftime("%Y-%m-%d")
        
        # Extraer número de semana
        match = re.search(r'(\d+)', week_str)
        if not match:
            return datetime.now().strftime("%Y-%m-%d")
            
        week_num = int(match.group(1))
        # Cálculo ISO: 4 de enero es siempre semana 1
        d = datetime(year, 1, 4)
        # Retroceder al lunes de esa semana y saltar X semanas
        start_date = d - timedelta(days=d.weekday()) + timedelta(weeks=week_num-1)
        return start_date.strftime("%Y-%m-%d")
    except:
        return datetime.now().strftime("%Y-%m-%d")

def save_to_database(df: pd.DataFrame, week: str, center: str, db_config: Optional[Dict] = None, uploaded_by: str = "System", clean_first: bool = True) -> bool:
    """Guarda o actualiza los datos en la base de datos (SQLite o PostgreSQL)"""
    try:
        init_database(db_config)
        
        # LIMPIEZA PREVIA: Si vamos a volcar un lote, es mejor limpiar lo que hubiera antes
        # para ese centro y semana, garantizando que el DB sea fiel al último archivo.
        if clean_first:
            delete_scorecard_batch(week, center, db_config)
            
        conn = get_db_connection(db_config)
        cursor = conn.cursor()
        
        is_postgres = db_config and db_config.get('type') == 'postgresql'
        ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        date_week = week_to_date(week)
        
        cols = [
            "semana", "fecha_semana", "centro", "driver_id", "driver_name", "calificacion", "score",
            "entregados", "dnr", "fs_count", "dnr_risk_events", "dcr", "pod", "cc",
            "fdps", "rts", "cdf", "detalles", "uploaded_by", "timestamp"
        ]
        
        # SQLite usa ?, PostgreSQL usa %s
        placeholder = "%s" if is_postgres else "?"
        placeholders = ", ".join([placeholder] * len(cols))
        
        for _, row in df.iterrows():
            vals = (
                week, date_week, center, str(row['ID']), str(row['Nombre']), 
                str(row['CALIFICACION']), float(row['SCORE']),
                float(row['Entregados']), float(row['DNR']), float(row['FS_Count']),
                float(row['DNR_RISK_EVENTS']), float(row['DCR']), float(row['POD']),
                float(row['CC']), float(row['FDPS']), float(row['RTS']), 
                float(row['CDF']), str(row['DETALLES']), uploaded_by, ts
            )
            
            if is_postgres:
                # PostgreSQL ON CONFLICT
                query = f"""
                    INSERT INTO scorecards ({', '.join(cols)})
                    VALUES ({placeholders})
                    ON CONFLICT (semana, centro, driver_id) 
                    DO UPDATE SET 
                        fecha_semana = EXCLUDED.fecha_semana,
                        calificacion = EXCLUDED.calificacion,
                        score = EXCLUDED.score,
                        entregados = EXCLUDED.entregados,
                        dnr = EXCLUDED.dnr,
                        fs_count = EXCLUDED.fs_count,
                        dnr_risk_events = EXCLUDED.dnr_risk_events,
                        dcr = EXCLUDED.dcr,
                        pod = EXCLUDED.pod,
                        cc = EXCLUDED.cc,
                        fdps = EXCLUDED.fdps,
                        rts = EXCLUDED.rts,
                        cdf = EXCLUDED.cdf,
                        detalles = EXCLUDED.detalles,
                        uploaded_by = EXCLUDED.uploaded_by,
                        timestamp = EXCLUDED.timestamp
                """
                cursor.execute(query, vals)
            else:
                # SQLite INSERT OR REPLACE
                query = f"INSERT OR REPLACE INTO scorecards ({', '.join(cols)}) VALUES ({placeholders})"
                cursor.execute(query, vals)
            
        conn.commit()
        
        # Crear vistas dinámicas por centro (Solo para PostgreSQL por organización)
        if is_postgres:
            try:
                # Obtener centros únicos
                cursor.execute("SELECT DISTINCT centro FROM scorecards")
                centros = [r[0] for r in cursor.fetchall()]
                
                for c in centros:
                    # Limpiar nombre para que sea válido en SQL
                    clean_name = "".join([char if char.isalnum() else "_" for char in c.lower()])
                    view_name = f"v_scorecard_{clean_name}"
                    cursor.execute(f"CREATE OR REPLACE VIEW {view_name} AS SELECT * FROM scorecards WHERE centro = '{c}'")
                conn.commit()
            except Exception as ve:
                logger.warning(f"No se pudieron actualizar las vistas: {str(ve)}")

        conn.close()
        
        # Limpieza final de duplicados físicos y normalización de semanas
        clean_database_duplicates(db_config)
        
        logger.info(f"✅ {len(df)} registros sincronizados con DB ({'PostgreSQL' if is_postgres else 'SQLite'})")
        return True
    except Exception as e:
        logger.error(f"Error guardando en DB: {str(e)}")
        if 'conn' in locals(): conn.close()
        return False

def get_center_targets(center: str, db_config: Optional[Dict] = None) -> Dict:
    """Obtiene los targets guardados para un centro o los defaults"""
    try:
        init_database(db_config)
        conn = get_db_connection(db_config)
        cursor = conn.cursor()
        
        q = "SELECT * FROM center_targets WHERE centro = %s" if db_config and db_config.get('type') == 'postgresql' else "SELECT * FROM center_targets WHERE centro = ?"
        cursor.execute(q, (center,))
        row = cursor.fetchone()
        conn.close()
        
        if row:
            if db_config and db_config.get('type') == 'postgresql':
                return {
                    'centro': row[0], 'target_dnr': row[1], 'target_dcr': row[2],
                    'target_pod': row[3], 'target_cc': row[4], 'target_fdps': row[5],
                    'target_rts': row[6], 'target_cdf': row[7]
                }
            else: # SQLite Row
                return dict(row)
    except Exception as e:
        logger.error(f"Error obteniendo targets: {e}")
    
    # Defaults si no existe o hay error
    return {
        'centro': center, 'target_dnr': 0.5, 'target_dcr': 0.995,
        'target_pod': 0.99, 'target_cc': 0.99, 'target_fdps': 0.98,
        'target_rts': 0.01, 'target_cdf': 0.95
    }

def save_center_targets(targets: Dict, db_config: Optional[Dict] = None):
    """Guarda o actualiza los targets para un centro"""
    try:
        conn = get_db_connection(db_config)
        cursor = conn.cursor()
        is_postgres = db_config and db_config.get('type') == 'postgresql'
        
        cols = ['centro', 'target_dnr', 'target_dcr', 'target_pod', 'target_cc', 'target_fdps', 'target_rts', 'target_cdf']
        vals = [targets[c] for c in cols]
        
        if is_postgres:
            q = f"""
                INSERT INTO center_targets ({', '.join(cols)}) VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
                ON CONFLICT (centro) DO UPDATE SET
                target_dnr=EXCLUDED.target_dnr, target_dcr=EXCLUDED.target_dcr, 
                target_pod=EXCLUDED.target_pod, target_cc=EXCLUDED.target_cc,
                target_fdps=EXCLUDED.target_fdps, target_rts=EXCLUDED.target_rts,
                target_cdf=EXCLUDED.target_cdf, timestamp=CURRENT_TIMESTAMP
            """
        else:
            q = f"INSERT OR REPLACE INTO center_targets ({', '.join(cols)}) VALUES (?, ?, ?, ?, ?, ?, ?, ?)"
        
        cursor.execute(q, vals)
        conn.commit()
        conn.close()
        return True
    except Exception as e:
        logger.error(f"Error guardando targets: {e}")
        return False

def clean_database_duplicates(db_config: Optional[Dict] = None) -> Tuple[bool, int]:
    """
    1. Normaliza formatos de semana (ej: W5 -> W05)
    2. Elimina duplicados físicos que puedan haber quedado por versiones antiguas sin restricciones
    """
    try:
        conn = get_db_connection(db_config)
        cursor = conn.cursor()
        
        is_postgres = db_config and db_config.get('type') == 'postgresql'
        
        # 1. Normalizar Semanas (W5 -> W05)
        cursor.execute("SELECT id, semana FROM scorecards WHERE semana NOT LIKE 'W%' OR LENGTH(semana) < 3")
        rows = cursor.fetchall()
        
        updated = 0
        for r_id, sem in rows:
            match = re.search(r'(\d+)', str(sem))
            if match:
                new_sem = f"W{int(match.group(1)):02d}"
                try:
                    # Intentar actualizar
                    q_upd = "UPDATE scorecards SET semana = %s WHERE id = %s" if is_postgres else "UPDATE scorecards SET semana = ? WHERE id = ?"
                    cursor.execute(q_upd, (new_sem, r_id))
                    updated += 1
                except:
                    # Si falla por UNIQUE, borrar el antiguo que está mal nombrado
                    q_del = "DELETE FROM scorecards WHERE id = %s" if is_postgres else "DELETE FROM scorecards WHERE id = ?"
                    cursor.execute(q_del, (r_id,))
                    updated += 1
        
        # 2. Eliminar duplicados físicos (por si acaso falló el UNIQUE en algún momento)
        # Mantenemos siempre el registro más reciente (el de ID más alto o timestamp mayor)
        if is_postgres:
            cursor.execute("""
                DELETE FROM scorecards 
                WHERE id IN (
                    SELECT id FROM (
                        SELECT id, ROW_NUMBER() OVER (PARTITION BY semana, centro, driver_id ORDER BY timestamp DESC, id DESC) as row_num
                        FROM scorecards
                    ) t WHERE t.row_num > 1
                )
            """)
        else:
            cursor.execute("""
                DELETE FROM scorecards 
                WHERE id NOT IN (
                    SELECT MAX(id) FROM scorecards GROUP BY semana, centro, driver_id
                )
            """)
        
        conn.commit()
        conn.close()
        return True, updated
    except Exception as e:
        logger.error(f"Error limpiando DB: {str(e)}")
        if 'conn' in locals(): conn.close()
        return False, 0

def main():
    """Función principal para procesamiento por lotes (múltiples centros/semanas)"""
    
    print("\n" + "="*80)
    print("AMAZON QUALITY SCORECARD - PROCESADOR MASIVO V3.0")
    print("="*80 + "\n")
    
    base_dir = os.path.dirname(os.path.abspath(__file__))
    
    # 1. LOCALIZAR TODOS LOS ARCHIVOS DE CONCESSIONS (BASE)
    concessions_paths = []
    for root, dirs, files in os.walk(base_dir):
        for file in files:
            if re.match(Config.PATTERN_CONCESSIONS, file, re.IGNORECASE):
                concessions_paths.append(os.path.join(root, file))
    
    if not concessions_paths:
        logger.error(f"No se encontró ningún archivo de Concessions en {base_dir}")
        return False
    
    logger.info(f"Se han detectado {len(concessions_paths)} lotes potenciales para procesar.\n")
    
    processed_count = 0
    processed_batches = set()
    
    for path_concessions in concessions_paths:
        current_dir = os.path.dirname(path_concessions)
        week, center = extract_info_from_path(path_concessions)
        
        batch_key = (week, center)
        if batch_key in processed_batches:
            logger.info(f"Saltando {center}/{week}: Ya procesado.")
            continue
            
        print(f"\n>>> PROCESANDO: Centro {center} | Semana {week}")
        print(f"    Carpeta: {current_dir}")
        
        try:
            # 2. BUSCAR ARCHIVOS COMPLEMENTARIOS EN LA MISMA CARPETA
            path_quality = find_file_in_dir(Config.PATTERN_QUALITY, current_dir)
            path_false_scan = find_file_in_dir(Config.PATTERN_FALSE_SCAN, current_dir)
            path_dwc = find_file_in_dir(Config.PATTERN_DWC, current_dir)
            path_fdps = find_file_in_dir(Config.PATTERN_FDPS, current_dir)
            
            # 3. LEER Y PROCESAR
            df_concessions = read_any_safe(path_concessions, str(path_concessions))
            df_quality = read_any_safe(path_quality, str(path_quality)) if path_quality is not None else None
            df_false_scan_html = read_any_safe(path_false_scan, str(path_false_scan)) if path_false_scan is not None else None
            df_dwc = read_any_safe(path_dwc, str(path_dwc)) if path_dwc is not None else None
            df_fdps = read_any_safe(path_fdps, str(path_fdps)) if path_fdps is not None else None
            
            if df_concessions is None:
                logger.error(f"Saltando lote: No se pudo leer {path_concessions}")
                continue
                
            df_conc_clean = process_concessions(df_concessions)
            
            # Duplicados
            if not df_conc_clean.empty and 'ID' in df_conc_clean.columns:
                if df_conc_clean['ID'].duplicated().any():
                    df_conc_clean = df_conc_clean.drop_duplicates(subset='ID', keep='first')
            
            df_qual_clean = process_quality(df_quality) if df_quality is not None else None
            df_fs_clean = process_false_scan(df_false_scan_html) if df_false_scan_html is not None else None
            df_dwc_clean = process_dwc(df_dwc) if df_dwc is not None else None
            df_fdps_clean = process_fdps(df_fdps) if df_fdps is not None else None
            
            # 4. MERGE Y SCORING
            df_merged = merge_data_smart(df_conc_clean, df_qual_clean, df_fs_clean, df_dwc_clean, df_fdps_clean)
            
            results = df_merged.apply(calculate_score_v3_robust, axis=1, result_type='expand')
            df_merged['CALIFICACION'] = results[0]
            df_merged['DETALLES'] = results[1]
            df_merged['SCORE'] = results[2]
            
            # 5. GENERAR EXCEL Y GUARDAR EN DB
            output_file = os.path.join(current_dir, f'Amazon_Quality_Scorecard_{center}_{week}.xlsx')
            success_excel = create_professional_excel(df_merged, output_file, center_name=center, week=week)
            success_db = save_to_database(df_merged, week, center)
            
            if success_excel:
                processed_count += 1
                processed_batches.add(batch_key)
                print(f"OK: EXCEL GENERADO: {os.path.basename(output_file)}")
            if success_db:
                print(f"OK: DATOS SINCRONIZADOS CON BASE DE DATOS")
            
        except Exception as e:
            logger.error(f"Error procesando lote {center}/{week}: {str(e)}")
            continue

    print("\n" + "="*80)
    print(f"PROCESAMIENTO FINALIZADO: {processed_count} reportes generados exitosamente.")
    print("="*80 + "\n")
    return True


# ═══════════════════════════════════════════════════════════════════════════════
# DSP WEEKLY SCORECARD PDF — PARSING Y GUARDADO (v3.2)
# ═══════════════════════════════════════════════════════════════════════════════

def _parse_station_kpis(text: str, errors: list) -> dict:
    """
    Extrae los KPIs de estación de la página 2 del PDF.
    Probado contra el texto real extraído por pdfplumber del PDF oficial de Amazon.
    
    Formato real del PDF (verificado):
      'Safe Driving Metric (FICO) 831|Fantastic Vehicle Audit (VSA) Compliance 100%|Fantastic'
      'Working Hours Compliance (WHC) 86.36%|Poor'
      'Overall Score: 80.8 | Great'
      'Rank at DIC1: 2 ( 0 WoW)'
    """
    kpis = {}

    def find_val_tier(pattern):
        """Busca patrón, devuelve (float_value, tier_string) o (None, None)."""
        m = re.search(pattern, text, re.IGNORECASE)
        if not m:
            return None, None
        val_str = m.group(1).replace('%', '').strip()
        tier    = m.group(2).strip() if len(m.groups()) >= 2 else None
        try:
            return float(val_str), tier
        except (ValueError, TypeError):
            return None, tier

    # ── Overall Score y Standing ────────────────────────────────────────────
    m = re.search(r'Overall Score:\s*([\d.]+)\s*\|\s*(Fantastic|Great|Fair|Poor)', text)
    if m:
        kpis['overall_score']    = float(m.group(1))
        kpis['overall_standing'] = m.group(2)
    else:
        errors.append("overall_score")

    # ── Rank ────────────────────────────────────────────────────────────────
    # Formato real: "Rank at DIC1: 2 ( 0 WoW)"
    m = re.search(r'Rank at \w+:\s*(\d+)\s*\(\s*([+-]?\d+)\s*WoW\)', text)
    if m:
        kpis['rank_station'] = int(m.group(1))
        kpis['rank_wow']     = int(m.group(2))
    else:
        errors.append("rank_station")

    # ── Tiers de categorías ──────────────────────────────────────────────────
    m = re.search(r'Compliance and Safety\s+(Fantastic|Great|Fair|Poor)', text)
    kpis['safety_tier'] = m.group(1) if m else None

    m = re.search(r'Delivery Quality[^:]*:\s*(Fantastic|Great|Fair|Poor)', text)
    kpis['quality_tier'] = m.group(1) if m else None

    m = re.search(r'^Capacity:\s*(Fantastic|Great|Fair|Poor)', text, re.MULTILINE)
    kpis['capacity_tier'] = m.group(1) if m else None

    # ── Safety: FICO ─────────────────────────────────────────────────────────
    # Formato: "Safe Driving Metric (FICO) 831|Fantastic"
    kpis['fico'], kpis['fico_tier'] = find_val_tier(
        r'(?:Safe Driving Metric|FICO)[^0-9\n]+([\d.]+)\|(Fantastic|Great|Fair|Poor)')

    # ── Speeding ─────────────────────────────────────────────────────────────
    # Formato: "Speeding Event Rate (Per 100 Trips) 0|Fantastic"
    kpis['speeding_rate'], kpis['speeding_tier'] = find_val_tier(
        r'Speeding Event Rate[^\n]+?(\b[\d.]+)\|(Fantastic|Great|Fair|Poor)')

    # ── Mentor Adoption ──────────────────────────────────────────────────────
    # Formato: "Mentor Adoption Rate 100%|Fantastic"
    kpis['mentor_adoption'], kpis['mentor_tier'] = find_val_tier(
        r'Mentor Adoption Rate\s+([\d.]+)%?\|(Fantastic|Great|Fair|Poor)')

    # ── VSA ──────────────────────────────────────────────────────────────────
    # Formato: "Vehicle Audit (VSA) Compliance 100%|Fantastic"
    kpis['vsa_compliance'], kpis['vsa_tier'] = find_val_tier(
        r'Vehicle Audit.*?Compliance\s+([\d.]+)%?\|(Fantastic|Great|Fair|Poor)')

    # ── BOC ──────────────────────────────────────────────────────────────────
    # Formato: "Breach of Contract (BOC) None"
    m = re.search(r'Breach of Contract[^\n]*?(None|Yes)\b', text, re.IGNORECASE)
    kpis['boc'] = m.group(1) if m else None

    # ── WHC ──────────────────────────────────────────────────────────────────
    # Formato: "Working Hours Compliance (WHC) 86.36%|Poor"
    kpis['whc_pct'], kpis['whc_tier'] = find_val_tier(
        r'Working Hours Compliance[^\n]+?([\d.]+)%?\|(Fantastic|Great|Fair|Poor)')
    if kpis.get('whc_pct') is None:
        errors.append("whc_pct")

    # ── CAS ──────────────────────────────────────────────────────────────────
    # Formato: "Comprehensive Audit Score (CAS) In Compliance"
    m = re.search(r'Comprehensive Audit Score[^\n]*?(In Compliance|Non-Compliant)', text, re.IGNORECASE)
    kpis['cas'] = m.group(1) if m else None

    # ── DCR ──────────────────────────────────────────────────────────────────
    # Formato: "Delivery Completion Rate(DCR) 98.61%|Great"
    kpis['dcr_pct'], kpis['dcr_tier'] = find_val_tier(
        r'Delivery Completion Rate[^0-9\n]+([\d.]+)%?\|(Fantastic|Great|Fair|Poor)')
    if kpis.get('dcr_pct') is None:
        errors.append("dcr_pct")

    # ── DNR DPMO ─────────────────────────────────────────────────────────────
    # Formato: "Delivered Not Received(DNR DPMO) 1360|Great"
    kpis['dnr_dpmo'], kpis['dnr_tier'] = find_val_tier(
        r'Delivered Not Received[^0-9\n]+([\d.]+)\|(Fantastic|Great|Fair|Poor)')

    # ── LoR DPMO ─────────────────────────────────────────────────────────────
    # Formato: "Lost on Road (LoR) DPMO 111|Poor"
    kpis['lor_dpmo'], kpis['lor_tier'] = find_val_tier(
        r'Lost on Road[^0-9\n]+([\d.]+)\|(Fantastic|Great|Fair|Poor)')
    if kpis.get('lor_dpmo') is None:
        errors.append("lor_dpmo")

    # ── DSC DPMO ─────────────────────────────────────────────────────────────
    # Formato: "Delivery Success Conditions (DSC DPMO) 1134|Great"
    kpis['dsc_dpmo'], kpis['dsc_tier'] = find_val_tier(
        r'Delivery Success Conditions[^0-9\n]+([\d.]+)\|(Fantastic|Great|Fair|Poor)')

    # ── POD ──────────────────────────────────────────────────────────────────
    # Formato: "Photo-On-Delivery 97.51%|Fantastic"
    kpis['pod_pct'], kpis['pod_tier'] = find_val_tier(
        r'Photo.On.Delivery\s+([\d.]+)%?\|(Fantastic|Great|Fair|Poor)')

    # ── Contact Compliance ───────────────────────────────────────────────────
    # Formato: "Contact Compliance 97.72%|Great"
    kpis['cc_pct'], kpis['cc_tier'] = find_val_tier(
        r'Contact Compliance\s+([\d.]+)%?\|(Fantastic|Great|Fair|Poor)')

    # ── Customer Escalation DPMO ─────────────────────────────────────────────
    # Formato: "Customer escalation DPMO 0|Fantastic"
    kpis['ce_dpmo'], kpis['ce_tier'] = find_val_tier(
        r'Customer escalation DPMO\s+([\d.]+)\|(Fantastic|Great|Fair|Poor)')

    # ── CDF DPMO ─────────────────────────────────────────────────────────────
    # Formato: "Customer Delivery Feedback 1615|Fantastic"
    kpis['cdf_dpmo'], kpis['cdf_tier'] = find_val_tier(
        r'Customer Delivery Feedback\s+([\d.]+)\|(Fantastic|Great|Fair|Poor)')

    # ── Capacity ─────────────────────────────────────────────────────────────
    # Formato: "Next Day Capacity Reliability 145.91%|Fantastic"
    kpis['capacity_next_day'], kpis['capacity_next_day_tier'] = find_val_tier(
        r'Next Day Capacity Reliability\s+([\d.]+)%?\|(Fantastic|Great|Fair|Poor)')

    # Formato: "Same Day/Sub-Same Day Capacity Reliability 100%|Fantastic"
    kpis['capacity_same_day'], kpis['capacity_same_day_tier'] = find_val_tier(
        r'Same Day[^0-9\n]+([\d.]+)%?\|(Fantastic|Great|Fair|Poor)')

    # ── Focus Areas ──────────────────────────────────────────────────────────
    # Formato: "1. Lost on Road (LoR) DPMO\n2. Working Hours Compliance (WHC)\n3. ..."
    idx_focus = text.find('Recommended Focus Areas')
    if idx_focus >= 0:
        focus_block = text[idx_focus:]
        areas = re.findall(r'\d+\.\s+(.+?)(?=\n\d+\.|\nCurrent|$)', focus_block, re.DOTALL)
        kpis['focus_area_1'] = areas[0].strip() if len(areas) > 0 else None
        kpis['focus_area_2'] = areas[1].strip() if len(areas) > 1 else None
        kpis['focus_area_3'] = areas[2].strip() if len(areas) > 2 else None
    else:
        kpis['focus_area_1'] = kpis['focus_area_2'] = kpis['focus_area_3'] = None

    return kpis


def _build_drivers_df(all_rows: list, errors: list) -> pd.DataFrame:
    """
    Construye DataFrame de drivers desde filas extraídas de las páginas 3-5.
    
    Estructura real del PDF (verificada con pdfplumber):
      Página 3: fila 0 es cabecera ['Transporter ID','Delivered','DCR','DSC DPMO',
                'LoR DPMO','POD','CC','CE CDF DPMO', None] — CE y CDF en col[7] y col[8]
      Páginas 4-5: sin cabecera, empiezan directo en datos
    """
    records = []
    header_skipped = False

    for row in all_rows:
        if not row or not row[0]:
            continue

        cell0 = str(row[0]).strip()

        # Saltar fila de cabecera (página 3)
        if not header_skipped and cell0 == 'Transporter ID':
            header_skipped = True
            continue

        # Saltar filas de título de sección (ej: "DSP WEEKLY SUMMARY")
        if not re.match(r'^[A-Z0-9]{10,20}$', cell0):
            continue

        def to_float(val):
            """Convierte valor a float; None si es '-', vacío o inválido."""
            if val in (None, '-', '', 'None'):
                return None
            try:
                return float(str(val).replace('%', '').strip())
            except (ValueError, TypeError):
                return None

        # DCR, POD, CC vienen como "98.96%" → convertir a ratio 0-1
        # usando la función ya existente safe_percentage()
        raw_dcr = row[2] if len(row) > 2 else None
        raw_pod = row[5] if len(row) > 5 else None
        raw_cc  = row[6] if len(row) > 6 else None

        records.append({
            'driver_id':          cell0,
            'entregados_oficial': to_float(row[1] if len(row) > 1 else None),
            'dcr_oficial':        safe_percentage(raw_dcr),     # → 0-1
            'dsc_dpmo':           to_float(row[3] if len(row) > 3 else None),
            'lor_dpmo':           to_float(row[4] if len(row) > 4 else None),
            'pod_oficial':        safe_percentage(raw_pod),     # → 0-1
            'cc_oficial':         safe_percentage(raw_cc),      # → 0-1
            'ce_dpmo':            to_float(row[7] if len(row) > 7 else None),
            'cdf_dpmo_oficial':   to_float(row[8] if len(row) > 8 else None),
        })

    df = pd.DataFrame(records)
    if df.empty:
        errors.append("tabla de drivers vacía")
    else:
        logger.info(f"✓ Drivers extraídos del PDF: {len(df)}")
    return df


def _build_wh_df(rows: list, errors: list) -> pd.DataFrame:
    """
    Construye DataFrame de excepciones WHC desde página 6.
    
    Estructura real (verificada):
      row[0]: ['#','Transporter ID','Daily Limit Exceeded','Weekly Limit Exceeded',
               'Under Offwork Limit','Work Day Limit Exceeded','WH Exception']
      row[1..]: ['1','AOX3PX1MTVS0E','No','No','Yes','No','Yes']
    """
    records = []
    for row in rows:
        if not row or not row[0]:
            continue
        # Saltar cabecera
        if str(row[0]).strip() == '#':
            continue
        # Validar que sea fila de datos (primera col es número)
        try:
            int(str(row[0]).strip())
        except (ValueError, TypeError):
            continue

        if len(row) < 6:
            continue

        def to_bool(val):
            return 1 if str(val or '').strip().lower() == 'yes' else 0

        records.append({
            'driver_id':              str(row[1]).strip(),
            'daily_limit_exceeded':   to_bool(row[2]),
            'weekly_limit_exceeded':  to_bool(row[3]),
            'under_offwork_limit':    to_bool(row[4]),
            'workday_limit_exceeded': to_bool(row[5]),
        })

    df = pd.DataFrame(records)
    if df.empty:
        logger.info("PDF: sin excepciones WHC esta semana")
    else:
        logger.info(f"✓ WHC excepciones extraídas: {len(df)}")
    return df


def parse_dsp_scorecard_pdf(pdf_bytes: bytes) -> dict:
    """
    Función principal: extrae los 3 conjuntos de datos del DSP Weekly Scorecard PDF.

    Returns dict con claves:
      'ok'      → bool
      'meta'    → {'centro': str, 'semana': str, 'year': int}
      'station' → dict con todos los KPIs de estación
      'drivers' → pd.DataFrame con métricas oficiales por conductor
      'wh'      → pd.DataFrame con infracciones Working Hours
      'errors'  → list de campos no encontrados (no fatales)
    """
    if not HAS_PDFPLUMBER:
        return {'ok': False, 'errors': ['pdfplumber no instalado. Ejecuta: pip install pdfplumber'],
                'meta': {}, 'station': {}, 'drivers': pd.DataFrame(), 'wh': pd.DataFrame()}

    result = {
        'ok': False, 'meta': {}, 'station': {},
        'drivers': pd.DataFrame(), 'wh': pd.DataFrame(), 'errors': []
    }

    try:
        with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:

            # ── PÁGINA 1: Meta — centro, semana, año ────────────────────────
            p1 = pdf.pages[0].extract_text() or ""

            # "TDSL at DIC1" → DIC1
            m = re.search(r'TDSL\s+at\s+([A-Z]{2,5}\d)', p1)
            centro = m.group(1).upper() if m else None

            # "Week 7" → W07
            m = re.search(r'Week\s+(\d+)', p1, re.IGNORECASE)
            semana = f"W{int(m.group(1)):02d}" if m else None

            # Año: "2026"
            m = re.search(r'\b(20\d{2})\b', p1)
            year = int(m.group(1)) if m else datetime.now().year

            if not centro:
                result['errors'].append("No se detectó el centro en el PDF")
                return result
            if not semana:
                result['errors'].append("No se detectó la semana en el PDF")
                return result

            result['meta'] = {'centro': centro, 'semana': semana, 'year': year}
            logger.info(f"PDF detectado: {centro} | {semana} | {year}")

            # ── PÁGINA 2: KPIs de estación ──────────────────────────────────
            if len(pdf.pages) > 1:
                p2_text = pdf.pages[1].extract_text() or ""
                result['station'] = _parse_station_kpis(p2_text, result['errors'])
            else:
                result['errors'].append("PDF sin página 2 (KPIs de estación)")

            # ── PÁGINAS 3-5: Tabla de drivers ───────────────────────────────
            # Página 3 tiene cabecera; páginas 4 y 5 solo datos
            all_driver_rows = []
            for page_idx in [2, 3, 4]:
                if page_idx < len(pdf.pages):
                    tbl = pdf.pages[page_idx].extract_table()
                    if tbl:
                        all_driver_rows.extend(tbl)

            if all_driver_rows:
                result['drivers'] = _build_drivers_df(all_driver_rows, result['errors'])
            else:
                result['errors'].append("No se encontraron datos de drivers en páginas 3-5")

            # ── PÁGINA 6: Working Hours Exceptions ──────────────────────────
            if len(pdf.pages) > 5:
                tbl = pdf.pages[5].extract_table()
                if tbl:
                    result['wh'] = _build_wh_df(tbl, result['errors'])

            result['ok'] = True

    except Exception as e:
        result['errors'].append(f"Error al procesar PDF: {str(e)}")
        logger.error(f"parse_dsp_scorecard_pdf error: {e}")

    return result


def save_station_scorecard(station_data: dict, week: str, center: str,
                           db_config=None, uploaded_by: str = "System") -> bool:
    """
    Guarda o actualiza los KPIs de estación en station_scorecards.
    UPSERT por (semana, centro) — reemplaza si ya existe.
    """
    try:
        init_database(db_config)
        conn   = get_db_connection(db_config)
        cursor = conn.cursor()
        is_pg  = db_config and db_config.get('type') == 'postgresql'
        ph     = '%s' if is_pg else '?'
        fecha  = week_to_date(week)

        fields = [
            'semana', 'fecha_semana', 'centro',
            'overall_score', 'overall_standing', 'rank_station', 'rank_wow',
            'safety_tier', 'fico', 'fico_tier', 'speeding_rate', 'speeding_tier',
            'mentor_adoption', 'mentor_tier', 'vsa_compliance', 'vsa_tier',
            'boc', 'whc_pct', 'whc_tier', 'cas',
            'quality_tier', 'dcr_pct', 'dcr_tier', 'dnr_dpmo', 'dnr_tier',
            'lor_dpmo', 'lor_tier', 'dsc_dpmo', 'dsc_tier',
            'pod_pct', 'pod_tier', 'cc_pct', 'cc_tier',
            'ce_dpmo', 'ce_tier', 'cdf_dpmo', 'cdf_tier',
            'capacity_tier', 'capacity_next_day', 'capacity_next_day_tier',
            'capacity_same_day', 'capacity_same_day_tier',
            'focus_area_1', 'focus_area_2', 'focus_area_3',
            'uploaded_by',
        ]

        vals = [
            week, fecha, center,
            station_data.get('overall_score'), station_data.get('overall_standing'),
            station_data.get('rank_station'), station_data.get('rank_wow'),
            station_data.get('safety_tier'),
            station_data.get('fico'), station_data.get('fico_tier'),
            station_data.get('speeding_rate'), station_data.get('speeding_tier'),
            station_data.get('mentor_adoption'), station_data.get('mentor_tier'),
            station_data.get('vsa_compliance'), station_data.get('vsa_tier'),
            station_data.get('boc'), station_data.get('whc_pct'), station_data.get('whc_tier'),
            station_data.get('cas'),
            station_data.get('quality_tier'),
            station_data.get('dcr_pct'), station_data.get('dcr_tier'),
            station_data.get('dnr_dpmo'), station_data.get('dnr_tier'),
            station_data.get('lor_dpmo'), station_data.get('lor_tier'),
            station_data.get('dsc_dpmo'), station_data.get('dsc_tier'),
            station_data.get('pod_pct'), station_data.get('pod_tier'),
            station_data.get('cc_pct'), station_data.get('cc_tier'),
            station_data.get('ce_dpmo'), station_data.get('ce_tier'),
            station_data.get('cdf_dpmo'), station_data.get('cdf_tier'),
            station_data.get('capacity_tier'),
            station_data.get('capacity_next_day'), station_data.get('capacity_next_day_tier'),
            station_data.get('capacity_same_day'), station_data.get('capacity_same_day_tier'),
            station_data.get('focus_area_1'), station_data.get('focus_area_2'),
            station_data.get('focus_area_3'),
            uploaded_by,
        ]

        col_list     = ', '.join(fields)
        placeholders = ', '.join([ph] * len(fields))

        if is_pg:
            update_set = ', '.join(
                f"{f} = EXCLUDED.{f}" for f in fields if f not in ('semana', 'centro')
            )
            query = f"""
                INSERT INTO station_scorecards ({col_list}) VALUES ({placeholders})
                ON CONFLICT (semana, centro) DO UPDATE SET {update_set}
            """
        else:
            query = f"INSERT OR REPLACE INTO station_scorecards ({col_list}) VALUES ({placeholders})"

        cursor.execute(query, vals)
        conn.commit()
        conn.close()
        logger.info(f"✓ station_scorecard guardado: {center} {week} | Score: {station_data.get('overall_score')} {station_data.get('overall_standing')}")
        return True

    except Exception as e:
        logger.error(f"save_station_scorecard error: {e}")
        return False


def update_drivers_from_pdf(drivers_df: pd.DataFrame, week: str, center: str,
                             db_config=None) -> Tuple[int, int]:
    """
    Actualiza SOLO las columnas _oficial en scorecards con los valores del PDF.
    
    - NO elimina filas existentes
    - NO crea filas nuevas
    - NO toca columnas base (dnr, dcr, pod, etc.)
    - Las columnas _oficial solo se rellenan aquí
    
    Returns: (n_actualizados, n_no_encontrados)
    """
    if drivers_df is None or drivers_df.empty:
        return 0, 0

    try:
        conn    = get_db_connection(db_config)
        cursor  = conn.cursor()
        is_pg   = db_config and db_config.get('type') == 'postgresql'
        ph      = '%s' if is_pg else '?'
        updated = 0
        not_found = []

        for _, row in drivers_df.iterrows():
            driver_id = row['driver_id']

            # Verificar existencia
            q_check = f"SELECT id FROM scorecards WHERE semana={ph} AND centro={ph} AND driver_id={ph}"
            cursor.execute(q_check, (week, center, driver_id))
            if not cursor.fetchone():
                not_found.append(driver_id)
                continue

            # UPDATE solo columnas _oficial — nunca toca columnas base
            q_update = f"""
                UPDATE scorecards SET
                    entregados_oficial = {ph},
                    dcr_oficial        = {ph},
                    pod_oficial        = {ph},
                    cc_oficial         = {ph},
                    dsc_dpmo           = {ph},
                    lor_dpmo           = {ph},
                    ce_dpmo            = {ph},
                    cdf_dpmo_oficial   = {ph},
                    pdf_loaded         = 1
                WHERE semana={ph} AND centro={ph} AND driver_id={ph}
            """
            cursor.execute(q_update, (
                row.get('entregados_oficial'),
                row.get('dcr_oficial'),
                row.get('pod_oficial'),
                row.get('cc_oficial'),
                row.get('dsc_dpmo'),
                row.get('lor_dpmo'),
                row.get('ce_dpmo'),
                row.get('cdf_dpmo_oficial'),
                week, center, driver_id
            ))
            updated += 1

        conn.commit()
        conn.close()

        if not_found:
            logger.warning(f"Drivers del PDF sin match en scorecards ({len(not_found)}): {not_found[:5]}{'...' if len(not_found) > 5 else ''}")
        logger.info(f"✓ update_drivers_from_pdf: {updated} actualizados, {len(not_found)} sin match")
        return updated, len(not_found)

    except Exception as e:
        logger.error(f"update_drivers_from_pdf error: {e}")
        return 0, 0


def save_wh_exceptions(wh_df: pd.DataFrame, week: str, center: str,
                       db_config=None, uploaded_by: str = "System") -> bool:
    """
    Guarda las excepciones de Working Hours en wh_exceptions.
    Primero borra las del mismo centro+semana para evitar duplicados,
    luego inserta las nuevas.
    """
    if wh_df is None or wh_df.empty:
        logger.info(f"Sin excepciones WHC para {center} {week}")
        return True

    try:
        conn   = get_db_connection(db_config)
        cursor = conn.cursor()
        is_pg  = db_config and db_config.get('type') == 'postgresql'
        ph     = '%s' if is_pg else '?'
        fecha  = week_to_date(week)

        # Limpiar registros anteriores del mismo lote
        cursor.execute(
            f"DELETE FROM wh_exceptions WHERE semana={ph} AND centro={ph}",
            (week, center)
        )

        for _, row in wh_df.iterrows():
            cursor.execute(
                f"""INSERT INTO wh_exceptions
                    (semana, fecha_semana, centro, driver_id,
                     daily_limit_exceeded, weekly_limit_exceeded,
                     under_offwork_limit, workday_limit_exceeded, uploaded_by)
                    VALUES ({ph},{ph},{ph},{ph},{ph},{ph},{ph},{ph},{ph})
                """,
                (
                    week, fecha, center, row['driver_id'],
                    int(row.get('daily_limit_exceeded', 0)),
                    int(row.get('weekly_limit_exceeded', 0)),
                    int(row.get('under_offwork_limit', 0)),
                    int(row.get('workday_limit_exceeded', 0)),
                    uploaded_by
                )
            )

        conn.commit()
        conn.close()
        logger.info(f"✓ WHC exceptions guardadas: {len(wh_df)} para {center} {week}")
        return True

    except Exception as e:
        logger.error(f"save_wh_exceptions error: {e}")
        return False


def get_station_scorecards(db_config=None) -> pd.DataFrame:
    """Devuelve todos los station_scorecards ordenados por centro y semana desc."""
    try:
        conn   = get_db_connection(db_config)
        is_pg  = db_config and db_config.get('type') == 'postgresql'
        query  = """
            SELECT semana, centro, overall_score, overall_standing,
                   rank_station, rank_wow,
                   fico, fico_tier, whc_pct, whc_tier,
                   dcr_pct, dcr_tier, dnr_dpmo, dnr_tier,
                   lor_dpmo, lor_tier, dsc_dpmo, dsc_tier,
                   pod_pct, pod_tier, cc_pct, cc_tier,
                   ce_dpmo, ce_tier, cdf_dpmo, cdf_tier,
                   speeding_rate, speeding_tier,
                   mentor_adoption, mentor_tier,
                   vsa_compliance, vsa_tier,
                   whc_pct, whc_tier, boc, cas,
                   capacity_next_day, capacity_next_day_tier,
                   capacity_same_day, capacity_same_day_tier,
                   safety_tier, quality_tier, capacity_tier,
                   focus_area_1, focus_area_2, focus_area_3,
                   uploaded_by, timestamp
            FROM station_scorecards
            ORDER BY centro ASC, semana DESC
        """
        df = pd.read_sql_query(query, conn)
        conn.close()
        return df
    except Exception as e:
        logger.error(f"get_station_scorecards error: {e}")
        return pd.DataFrame()


    success = main()
    sys.exit(0 if success else 1)
