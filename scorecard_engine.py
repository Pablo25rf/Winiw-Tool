# Copyright (c) 2026 Winiw — pablo25rf. Todos los derechos reservados.
# Software propietario. Prohibido su uso, copia o distribución sin autorización escrita.
# Ver LICENSE en la raíz del proyecto.
"""
Quality Scorecard Engine v3.9
=====================================
Sistema de procesamiento y análisis de métricas de calidad para conductores de Logística.
Soporta PostgreSQL y SQLite con auto-migraciones y validaciones robustas.

Versión: 3.9
Fecha: Marzo 2026
"""

import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule
import warnings
import logging
import functools
from typing import Dict, List, Tuple, Optional
from contextlib import contextmanager
import re
import os
import sqlite3
import html as _html_escape
import io
import hashlib
HAS_POSTGRES = False
try:
    import psycopg2
    from psycopg2 import sql
    from psycopg2 import pool as pg_pool
    HAS_POSTGRES = True
except ImportError:
    pass  # psycopg2 no instalado — modo SQLite activo

# ── Pool de conexiones PostgreSQL (reutiliza conexiones entre queries) ────────
import threading as _threading
_PG_POOL: "pg_pool.ThreadedConnectionPool | None" = None
_PG_POOL_LOCK = _threading.Lock()

def _get_pg_pool(db_config: dict):
    """
    Devuelve (o crea) un ThreadedConnectionPool para Supabase.
    minconn=1, maxconn=5 — apropiado para 10-30 usuarios con Streamlit Cloud.
    Lock para evitar race condition si dos threads llegan simultáneamente.
    """
    global _PG_POOL
    with _PG_POOL_LOCK:
        if _PG_POOL is None or _PG_POOL.closed:
            _PG_POOL = pg_pool.ThreadedConnectionPool(
                minconn=1, maxconn=5,
                host=db_config.get('host', 'localhost'),
                database=db_config.get('database', 'postgres'),
                user=db_config.get('user', 'postgres'),
                password=db_config.get('password', ''),
                port=db_config.get('port', 5432),
                connect_timeout=10,
                sslmode='require',          # Supabase exige SSL
                keepalives=1,               # TCP keepalives para evitar cortes por inactividad
                keepalives_idle=30,         # Primer keepalive a los 30s sin actividad
                keepalives_interval=10,     # Reintento cada 10s
                keepalives_count=5,         # Máx 5 reintentos antes de cerrar
                options="-c statement_timeout=30000",  # Timeout de queries a 30s
            )
    return _PG_POOL
HAS_BCRYPT = False
try:
    import bcrypt
    HAS_BCRYPT = True
except Exception:
    pass  # bcrypt no instalado — se usará SHA-256 como fallback

HAS_PDFPLUMBER = False
try:
    import pdfplumber
    HAS_PDFPLUMBER = True
except Exception:
    pass  # pdfplumber no instalado — lectura de PDF desactivada

from datetime import datetime, timedelta

logger = logging.getLogger(__name__)
if not logger.handlers:
    logger.setLevel(logging.INFO)

warnings.filterwarnings('ignore', category=pd.errors.DtypeWarning)
warnings.filterwarnings('ignore', message='.*openpyxl.*')

class Config:
    """Configuración centralizada del sistema"""
    
    # Límites de validación
    MAX_DNR = 500  # Aumentado para evitar caps en datos históricos acumulados
    MAX_FALSE_SCAN = 2000  # Aumentado
    MAX_CONDUCTORES = 5000  # Máximo conductores esperados
    # Targets por defecto — I-07: fuente única de verdad
    DEFAULT_TARGETS = {
        'target_dnr': 0.5, 'target_dcr': 0.995, 'target_pod': 0.99,
        'target_cc': 0.99, 'target_fdps': 0.98, 'target_rts': 0.01, 'target_cdf': 0.95
    }
    
    # Nombres de archivos esperados (patrones)
    # DSC-Concessions debe verificarse ANTES que Concessions (tiene prioridad)
    PATTERN_DSC_CONCESSIONS = r'.*dsc.*concessions.*\.(csv|xlsx)'
    PATTERN_CONCESSIONS     = r'(?!.*dsc).*concessions.*\.(csv|xlsx)'  # excluye DSC
    # Detecta: quality_overview, Amazon_Quality_Scorecard, quality-report, etc.
    # Excluye: POD-Quality (que es un PDF de otro tipo)
    PATTERN_QUALITY = r'.*quality.*(overview|scorecard|report).*\.(csv|xlsx)'
    PATTERN_FALSE_SCAN = r'.*false.*scan.*\.html'
    PATTERN_DWC = r'.*(dwc|iadc).*\.(csv|html)'
    PATTERN_FDPS = r'.*fdps.*\.(xlsx|csv)'
    PATTERN_DAILY = r'.*daily.*report.*\.html'
    PATTERN_OFFICIAL_SCORECARD = r'.*scorecard.*3\.0.*\.pdf'
    
    # Columnas requeridas por archivo
    REQUIRED_CONCESSIONS = ['Nombre del agente de entrega', 'ID de agente de entrega', 
                           'Paquetes entregados no recibidos (DNR)']
    REQUIRED_QUALITY = ['ID del transportista', 'DCR']

    # Valores por defecto
    DEFAULT_DNR = 0
    DEFAULT_FS = 0
    DEFAULT_DCR = 1.0
    DEFAULT_POD = 1.0
    DEFAULT_CC = 1.0
    DEFAULT_FDPS = 1.0
    DEFAULT_RTS = 0.0
    DEFAULT_CDF = 1.0

def safe_number(val, default=0.0) -> float:
    """Convierte a número de forma segura manejando comas y strings"""
    try:
        if pd.isna(val) or val == '' or val == '-':
            return default
        s = str(val).replace(',', '.').replace('%', '').strip()
        return float(s)
    except (ValueError, TypeError):
        return default

def safe_percentage(val) -> float:
    """Convierte porcentaje string a float de forma segura (0-1.0) con validación de rango"""
    try:
        if pd.isna(val) or val == "" or val == "-":
            return 1.0  # Default a 100% si no hay dato
        
        # Detectar si tiene el símbolo % para ser más precisos
        has_percent = "%" in str(val)
        
        # Normalizar: eliminar %, espacios, y unificar separador decimal
        s = str(val).replace("%", "").strip()
        
        # Amazon a veces usa puntos para miles y comas para decimales (EU)
        if "," in s and "." in s:
            if s.find(",") > s.find("."):
                s = s.replace(".", "").replace(",", ".") # EU: 1.234,56 -> 1234.56
            else:
                s = s.replace(",", "") # US: 1,234.56 -> 1234.56
        elif "," in s:
            s = s.replace(",", ".")
            
        num = float(s)
        
        # Si explícitamente tiene %, siempre dividimos por 100
        if has_percent:
            return max(0.0, min(num / 100.0, 1.0))
        
        # Si el número es gigante (ej: 9494), probablemente es un ID capturado por error
        if num > 100.0:
            return 1.0
            
        if num > 1.0:
            return num / 100.0
            
        # Asegurar rango 0-1.0
        return max(0.0, min(num, 1.0))
    except (ValueError, TypeError):
        return 1.0

def clean_id(val) -> str:
    """Limpia y normaliza IDs de forma segura"""
    if pd.isna(val) or val == '':
        return "UNKNOWN"
    return str(val).strip().upper()

_INVALID_SHEET_RE = re.compile(r'[/\\?*\[\]:]')

def truncate_sheet_name(name: str, max_length: int = 31) -> str:
    """Trunca nombre de hoja Excel a máximo permitido eliminando caracteres inválidos."""
    return _INVALID_SHEET_RE.sub('_', name)[:max_length]

def validate_dataframe(df: pd.DataFrame, required_cols: List[str], 
                      name: str) -> Tuple[bool, str]:
    """Valida que un DataFrame tenga las columnas requeridas"""
    if df is None or df.empty:
        return False, f"{name} está vacío"
    
    missing_cols = [col for col in required_cols if col not in df.columns]
    
    if missing_cols:
        return False, f"{name} le faltan columnas: {', '.join(missing_cols)}"
    
    return True, "OK"

def read_csv_safe(filepath_or_buffer, encoding: str = 'utf-8') -> Optional[pd.DataFrame]:
    """Lee CSV con manejo robusto de encoding (soporta paths o buffers)"""
    encodings = [encoding, 'utf-8', 'latin-1', 'iso-8859-1', 'cp1252']
    
    for enc in encodings:
        try:
            # Si es un buffer, resetear posición
            if hasattr(filepath_or_buffer, 'seek'):
                filepath_or_buffer.seek(0)
                
            df = pd.read_csv(filepath_or_buffer, encoding=enc)
            return df
        except UnicodeDecodeError:
            continue
        except Exception as e:
            logger.error(f"Error leyendo CSV: {str(e)}")
            return None
    return None

def read_any_safe(filepath_or_buffer, filename: str = "") -> Optional[pd.DataFrame]:
    """Detecta extensión y usa el lector adecuado"""
    name = filename.lower() or (filepath_or_buffer.name.lower() if hasattr(filepath_or_buffer, 'name') else "")
    
    if name.endswith('.csv'):
        return read_csv_safe(filepath_or_buffer)
    elif name.endswith('.xlsx') or name.endswith('.xls'):
        return read_excel_safe(filepath_or_buffer)
    elif name.endswith('.html'):
        return read_html_safe(filepath_or_buffer)
    
    # Si no tiene extensión clara, probar CSV por defecto
    return read_csv_safe(filepath_or_buffer)

def read_html_safe(filepath_or_buffer) -> Optional[pd.DataFrame]:
    """Lee HTML y extrae tabla de False Scan o DWC (soporta paths o buffers)"""
    try:
        if hasattr(filepath_or_buffer, 'read'):
            content = filepath_or_buffer.read()
            if isinstance(content, bytes):
                content = content.decode('utf-8', errors='ignore')
        else:
            with open(filepath_or_buffer, 'r', encoding='utf-8', errors='ignore') as f:
                content = f.read()
        
        dfs = pd.read_html(io.StringIO(content))
        
        # Buscar tabla con datos
        for df in dfs:
            if df.empty: continue
            
            # Aplanar multi-index si existe (Amazon usa mucho esto)
            if isinstance(df.columns, pd.MultiIndex):
                # Intentamos unir los niveles o quedarnos con el más descriptivo
                new_cols = []
                for col_tuple in df.columns.values:
                    # Filtrar 'Unnamed' y unir
                    parts = [str(p) for p in col_tuple if 'Unnamed' not in str(p)]
                    if not parts: # Si todo era Unnamed, coger el último
                        new_cols.append(str(col_tuple[-1]))
                    else:
                        new_cols.append(" ".join(parts))
                df.columns = new_cols
            
            df.columns = [str(c).strip() for c in df.columns]
            cols_lower = [str(c).lower() for c in df.columns]
            
            # Caso 1: False Scan
            if any('transporter id' in c or 'driver id' in c for c in cols_lower) and \
               any('false scan' in c for c in cols_lower):
                return df
            
            # Caso 2: DWC/IADC (Nuevo formato o antiguo)
            if any('transporter id' in c or 'driver id' in c for c in cols_lower) and \
               (any('dwc' in c for c in cols_lower) or any('iadc' in c for c in cols_lower)):
                return df
            
            # Caso 3: Daily Report
            if any('transporter id' in c or 'driver id' in c for c in cols_lower) and \
               any('rts' in c or 'dnr' in c for c in cols_lower):
                return df
                
        return None
    except Exception as e:
        logger.error(f"Error leyendo HTML: {str(e)}")
        return None

def read_excel_safe(filepath_or_buffer) -> Optional[pd.DataFrame]:
    """Lee Excel buscando header dinámicamente en todas las hojas (soporta paths o buffers).
    
    Optimizado: lee cada hoja UNA sola vez con header=None y busca la fila de cabecera
    iterando sobre las filas ya en memoria en vez de re-parsear el fichero 25 veces.
    """
    _ID_KEYS      = {'id', 'agente', 'driver', 'transporter'}
    _METRIC_KEYS  = {'dnr', 'concessions', 'delivered', 'rts', 'quality', 'score'}
    _SKIP_WORDS   = {'by', 'report'}
    MAX_HEADER_SCAN = 25

    def _is_valid_header(row_values) -> bool:
        cols = [str(v).strip().lower() for v in row_values if pd.notna(v)]
        if not cols:
            return False
        unnamed_count = sum(1 for c in cols if 'unnamed' in c)
        if unnamed_count > len(cols) * 0.5 and len(cols) > 2:
            return False
        has_id = any(
            any(k in c for k in _ID_KEYS) and not any(w in c for w in _SKIP_WORDS)
            for c in cols
        )
        has_metrics = any(k in c for k in _METRIC_KEYS for c in cols)
        return has_id and has_metrics

    try:
        if hasattr(filepath_or_buffer, 'seek'):
            filepath_or_buffer.seek(0)

        xl = pd.ExcelFile(filepath_or_buffer, engine='openpyxl')
        sheets = xl.sheet_names

        priority_sheets = ['DNR by Transporter ID', 'DSC by Transporter ID', 'DNR Concessions', 'Sheet1', 'Feuille1', 'Hoja1']
        sorted_sheets = [s for s in priority_sheets if s in sheets] + [s for s in sheets if s not in priority_sheets]

        for sheet in sorted_sheets:
            try:
                raw = pd.read_excel(xl, sheet_name=sheet, header=None, nrows=MAX_HEADER_SCAN + 50)
                if raw is None or raw.empty:
                    continue
                header_row = None
                for i in range(min(MAX_HEADER_SCAN, len(raw))):
                    if _is_valid_header(raw.iloc[i].tolist()):
                        header_row = i
                        break
                if header_row is None:
                    continue
                df = pd.read_excel(xl, sheet_name=sheet, skiprows=header_row)
                df.columns = [str(c).strip() for c in df.columns]
                logger.info(f"✓ Header detectado en hoja '{sheet}', fila {header_row}")
                return df
            except Exception:
                continue
        return None
    except Exception as e:
        logger.error(f"Error leyendo Excel: {str(e)}")
        return None

def find_file(pattern: str, search_path: str = ".") -> Optional[str]:
    """Busca un archivo recursivamente usando un patrón regex"""
    for root, dirs, files in os.walk(search_path):
        for file in files:
            if re.match(pattern, file, re.IGNORECASE):
                return os.path.join(root, file)
    return None

def process_fdps(df: pd.DataFrame) -> pd.DataFrame:
    """Procesa archivo FDPS con validaciones"""
    if df is None or df.empty:
        return pd.DataFrame(columns=['ID', 'FDPS'])
    
    logger.info("Procesando FDPS...")
    
    # Buscar columna de ID y FDPS
    id_col = None
    fdps_col = None
    
    for col in df.columns:
        col_str = str(col).lower()
        if 'id' in col_str and ('transporter' in col_str or 'driver' in col_str or 'agente' in col_str):
            id_col = col
        if 'fdps' in col_str and 'share' not in col_str:
            fdps_col = col

    if id_col is not None and fdps_col is not None:
        df = df.rename(columns={id_col: 'ID', fdps_col: 'FDPS'})
        df['ID'] = df['ID'].apply(clean_id)
        df['FDPS'] = df['FDPS'].apply(safe_percentage)
        logger.info(f"✓ FDPS procesado: {len(df)} conductores")
        return df[['ID', 'FDPS']]
    
    logger.warning("No se encontraron columnas de ID o FDPS en el reporte FDPS")
    return pd.DataFrame(columns=['ID', 'FDPS'])

def process_concessions(df: pd.DataFrame) -> pd.DataFrame:
    """Procesa archivo Concessions con validaciones robustas de columnas"""
    if df is None or df.empty:
        return pd.DataFrame(columns=['ID', 'Nombre', 'DNR', 'RTS', 'Entregados'])
    
    logger.info("Procesando Concessions...")
    
    # Mapeo mucho más estricto para evitar confundir IDs con métricas
    mapping = {}
    
    # 1. Identificar las mejores columnas para cada métrica
    best_cols = {}
    
    for col in df.columns:
        c = str(col).lower()
        
        # ID
        if "id" in c and any(k in c for k in ["agente", "agent", "transporter", "transportista", "driver"]):
            best_cols["ID"] = col
        # Nombre
        elif any(k in c for k in ["nombre", "name"]) and "agente" in c:
            best_cols["Nombre"] = col
        # DNR (MEJORADO: Detectar columnas de semana específica)
        elif (any(k in c for k in ["dnr", "no recibidos"]) or (c == "concessions")) and \
             not re.search(r'\bid\b', c) and "tracking" not in c:
            
            # NUEVO: Detectar columnas con patrón YYYY-WW_DNR (ej: 2026-05_DNR)
            week_pattern = re.match(r'(\d{4})-(\d{2})_dnr', c)
            
            is_better = "DNR" not in best_cols
            
            if not is_better:
                current_col = str(best_cols["DNR"]).lower()
                current_has_dpmo = "dpmo" in current_col or "%" in current_col
                current_is_total = "total" in current_col
                current_week_pattern = re.match(r'(\d{4})-(\d{2})_dnr', current_col)
                
                new_has_dpmo = "dpmo" in c or "%" in c
                new_is_total = "total" in c
                
                # PRIORIDAD 1: Evitar DPMO y %
                if current_has_dpmo and not new_has_dpmo:
                    is_better = True
                # PRIORIDAD 2: Preferir columna de semana específica sobre Total
                elif not current_has_dpmo and not new_has_dpmo:
                    if week_pattern and current_is_total:
                        # Columna de semana específica es mejor que Total
                        is_better = True
                        logger.info(f"  🎯 Detectada columna de semana específica: {col} (en lugar de Total)")
                    elif week_pattern and current_week_pattern:
                        # Si ambas son semanas, preferir la más reciente (última)
                        current_week = int(current_week_pattern.group(2))
                        new_week = int(week_pattern.group(2))
                        if new_week > current_week:
                            is_better = True
                            logger.info(f"  🎯 Usando semana más reciente: W{new_week:02d} (en lugar de W{current_week:02d})")
                    elif not week_pattern and not current_week_pattern:
                        # Si ninguna es semana específica, preferir "Total"
                        if new_is_total and not current_is_total:
                            is_better = True
            
            if is_better:
                best_cols["DNR"] = col
        # RTS (Preferir % si existe)
        elif "rts" in c or "devueltos" in c:
            if "RTS" not in best_cols or "%" in c:
                best_cols["RTS"] = col
        # Entregados
        elif ("entregados" in c or "delivered" in c) and "no" not in c:
            is_better = "Entregados" not in best_cols
            if not is_better:
                if "total" in c and "total" not in str(best_cols["Entregados"]).lower():
                    is_better = True
            if is_better:
                best_cols["Entregados"] = col
            
    # Crear el mapping inverso para rename
    mapping = {v: k for k, v in best_cols.items()}
    df = df.rename(columns=mapping)
    
    # Asegurar columnas mínimas
    for col in ['ID', 'Nombre', 'DNR', 'RTS', 'Entregados']:
        if col not in df.columns:
            if col == 'ID':
                logger.error("No se pudo identificar la columna de ID en Concessions")
                return pd.DataFrame(columns=['ID', 'Nombre', 'DNR', 'RTS', 'Entregados'])
            elif col == 'Nombre':
                df['Nombre'] = df['ID']
            elif col == 'DNR': df['DNR'] = 0
            elif col == 'RTS': df['RTS'] = 0.0
            elif col == 'Entregados': df['Entregados'] = 0
    
    df['ID'] = df['ID'].apply(clean_id)
    df = df[df['ID'] != 'UNKNOWN']
    
    # Validar tipos
    df['DNR'] = df['DNR'].apply(lambda x: safe_number(x, 0))
    df['Entregados'] = df['Entregados'].apply(lambda x: safe_number(x, 0))
    df['RTS'] = df['RTS'].apply(safe_percentage)
    
    # --- LÓGICA DE AGRUPACIÓN INTELIGENTE CON ANTI-DUPLICACIÓN (CORREGIDA) ---
    logger.info(f"  -> Procesando {len(df)} registros de Concessions...")
    
    original_count = len(df)
    df = df.drop_duplicates(subset=['ID'], keep='first')
    duplicates_removed = original_count - len(df)
    
    if duplicates_removed > 0:
        logger.warning(f"  ⚠️ Se eliminaron {duplicates_removed} registros duplicados del mismo conductor")
    
    logger.info(f"  -> {len(df)} conductores únicos encontrados")

    df['DNR'] = pd.to_numeric(df['DNR'], errors='coerce').fillna(0.0).clip(upper=Config.MAX_DNR)

    logger.info(f"✓ Concessions procesado: {len(df)} conductores")
    return df[['ID', 'Nombre', 'DNR', 'RTS', 'Entregados']]

def process_quality(df: pd.DataFrame) -> pd.DataFrame:
    """Procesa archivo Quality Overview con validaciones robustas"""
    if df is None or df.empty:
        return pd.DataFrame(columns=['ID', 'DCR', 'POD', 'CC', 'CDF'])
    
    logger.info("Procesando Quality Overview...")
    
    # Mapeo flexible
    mapping = {}
    found_targets = set()
    for col in df.columns:
        c = str(col).lower()
        target = None
        if 'id' in c and ('transportista' in c or 'transporter' in c or 'driver' in c):
            target = 'ID'
        elif 'dcr' in c: target = 'DCR'
        elif 'pod' in c: target = 'POD'
        elif 'cc' in c: target = 'CC'
        elif 'cdf' in c: target = 'CDF'
            
        if target and target not in found_targets:
            mapping[col] = target
            found_targets.add(target)
            
    df = df.rename(columns=mapping)
    
    # Limpiar IDs
    if 'ID' in df.columns:
        df['ID'] = df['ID'].apply(clean_id)
    else:
        return pd.DataFrame(columns=['ID', 'DCR', 'POD', 'CC', 'CDF'])
    
    # Validar métricas
    for col in ['DCR', 'POD', 'CC', 'CDF']:
        if col in df.columns:
            df[col] = df[col].apply(safe_percentage)
        else:
            df[col] = getattr(Config, f'DEFAULT_{col}')
    
    logger.info(f"✓ Quality Overview procesado: {len(df)} conductores")
    return df[['ID', 'DCR', 'POD', 'CC', 'CDF']]

def process_false_scan(df: pd.DataFrame) -> pd.DataFrame:
    """Procesa archivo False Scan HTML con validaciones"""
    if df is None or df.empty:
        return pd.DataFrame(columns=['ID', 'FS_Count'])
    
    logger.info("Procesando False Scan...")
    
    # Aplanar multi-index si existe
    if isinstance(df.columns, pd.MultiIndex):
        df.columns = df.columns.get_level_values(-1)
    
    # Buscar columna de ID
    id_col = None
    for col in df.columns:
        if 'transporter' in str(col).lower() or 'driver' in str(col).lower():
            if 'id' in str(col).lower():
                id_col = col
                break
    
    if id_col is None:
        logger.error("No se encontró columna de ID en False Scan")
        return pd.DataFrame(columns=['ID', 'FS_Count'])
    
    # Buscar columna de False Scan count
    fs_col = None
    for col in df.columns:
        if 'false scan' in str(col).lower() and 'share' not in str(col).lower():
            fs_col = col
            break
    
    if fs_col is None:
        logger.error("No se encontró columna de False Scan count")
        return pd.DataFrame(columns=['ID', 'FS_Count'])
    
    # Renombrar
    df = df.rename(columns={
        id_col: 'ID',
        fs_col: 'FS_Count'
    })
    
    # Limpiar
    df['ID'] = df['ID'].apply(clean_id)
    df['FS_Count'] = df['FS_Count'].fillna(0).astype(int)
    
    # Detectar valores extremos
    extreme_fs = df[df['FS_Count'] > Config.MAX_FALSE_SCAN]
    if len(extreme_fs) > 0:
        logger.warning(f"⚠️ {len(extreme_fs)} conductores con FS > {Config.MAX_FALSE_SCAN} detectados")
    
    logger.info(f"✓ False Scan procesado: {len(df)} conductores")
    return df[['ID', 'FS_Count']]

def process_dwc(df: pd.DataFrame) -> pd.DataFrame:
    """Procesa archivo DWC/IADC con validaciones"""
    logger.info("Procesando DWC/IADC...")
    
    # 1. Detectar ID del transportista
    id_col = None
    for col in df.columns:
        if 'transporter id' in str(col).lower() or 'driver id' in str(col).lower():
            id_col = col
            break
            
    if id_col is None:
        logger.error("No se encontró columna de ID en DWC/IADC")
        return pd.DataFrame(columns=['ID', 'DNR_RISK_EVENTS', 'CC_DWC', 'IADC'])

    df = df.rename(columns={id_col: 'ID'})
    df['ID'] = df['ID'].apply(clean_id)
    
    # 2. Buscar eventos de riesgo DNR (Formato antiguo)
    dnr_risk_events = pd.Series(0.0, index=df.index)
    risk_group = None  # inicialización explícita — I-02 fix
    if 'Type' in df.columns and 'Total' in df.columns:
        # Filtrar solo filas con DNR Risk
        dnr_mask = df['Type'].str.contains('DNR Risk', case=False, na=False)
        # Agrupar por ID (si el archivo viene en filas por evento)
        risk_group = df[dnr_mask].groupby('ID')['Total'].sum()
    else:
        # Formato nuevo: El riesgo suele estar en columnas específicas
        for col in df.columns:
            c_lower = str(col).lower()
            if 'dnr risk' in c_lower and 'total' in c_lower:
                series = df[col]
                # Si col es duplicado, series será un DataFrame
                if isinstance(series, pd.DataFrame):
                    series = series.iloc[:, 0] # Tomar solo la primera
                
                vals = series.apply(lambda x: safe_number(x, 0)).fillna(0)
                dnr_risk_events = dnr_risk_events + vals
    
    # 3. Buscar DWC% e IADC% (Formato nuevo)
    dwc_col = None
    iadc_col = None
    for col in df.columns:
        c = str(col).lower()
        if 'dwc' in c and '%' in c: dwc_col = col
        if 'iadc' in c and '%' in c: iadc_col = col
        
    res = pd.DataFrame({'ID': df['ID']})
    
    if risk_group is not None and len(risk_group) > 0:
        res = res.merge(risk_group.reset_index().rename(columns={'Total': 'DNR_RISK_EVENTS'}), on='ID', how='left')
    else:
        res['DNR_RISK_EVENTS'] = dnr_risk_events
        
    if dwc_col:
        series = df[dwc_col]
        if isinstance(series, pd.DataFrame): series = series.iloc[:, 0]
        res['CC_DWC'] = series.apply(safe_percentage)
    if iadc_col:
        series = df[iadc_col]
        if isinstance(series, pd.DataFrame): series = series.iloc[:, 0]
        res['IADC'] = series.apply(safe_percentage)
        
    res = res.fillna(0)
    
    # 4. Agrupar por ID de forma segura (evitando KeyError si faltan columnas)
    agg_dict = {"DNR_RISK_EVENTS": "sum"}
    if "CC_DWC" in res.columns:
        agg_dict["CC_DWC"] = "max"
    if "IADC" in res.columns:
        agg_dict["IADC"] = "max"
        
    res = res.groupby("ID").agg(agg_dict).reset_index()
    
    logger.info(f"✓ DWC/IADC procesado: {len(res)} conductores")
    return res

def process_daily_report(df: pd.DataFrame) -> pd.DataFrame:
    """Procesa reportes diarios para agregar datos a la semana"""
    logger.info("Procesando Daily Report...")
    
    mapping = {}
    found_targets = set()
    for col in df.columns:
        c = str(col).lower()
        target = None
        if 'id' in c: target = 'ID'
        elif 'dnr' in c and 'id' not in c and 'tracking' not in c: target = 'DNR'
        elif 'rts' in c: target = 'RTS'
        elif 'pod' in c and 'fail' in c: target = 'POD_Fails'
        elif 'cc' in c and 'fail' in c: target = 'CC_Fails'
        elif ('delivered' in c or 'entregados' in c) and 'not' not in c: target = 'Entregados'
        
        if target and target not in found_targets:
            mapping[col] = target
            found_targets.add(target)
            
    df = df.rename(columns=mapping)
    
    if 'ID' not in df.columns:
        return pd.DataFrame()
        
    df['ID'] = df['ID'].apply(clean_id)
    
    # Asegurar columnas numéricas
    for col in ['DNR', 'Entregados', 'POD_Fails', 'CC_Fails']:
        if col in df.columns:
            df[col] = df[col].apply(lambda x: safe_number(x, 0))
        else:
            df[col] = 0
            
    if 'RTS' in df.columns:
        df['RTS'] = df['RTS'].apply(safe_percentage)
    else:
        df['RTS'] = 0.0
        
    return df[['ID', 'DNR', 'RTS', 'Entregados', 'POD_Fails', 'CC_Fails']]

def merge_data_smart(df_concessions: pd.DataFrame,
                     df_quality: Optional[pd.DataFrame],
                     df_false_scan: Optional[pd.DataFrame],
                     df_dwc: Optional[pd.DataFrame],
                     df_fdps: Optional[pd.DataFrame] = None,
                     df_daily: Optional[pd.DataFrame] = None) -> pd.DataFrame:
    """
    Merge inteligente de todos los datasets con manejo de missing data
    """
    logger.info("Iniciando merge inteligente de datos...")
    
    # Base: Concessions (tiene nombres + DNR correcto)
    df_base = df_concessions.copy()
    total_base = len(df_base)
    logger.info(f"  Base (Concessions): {total_base} conductores")
    
    # Merge con Quality Overview (métricas)
    if df_quality is not None and not df_quality.empty:
        before = len(df_base)
        df_base = df_base.merge(df_quality, on='ID', how='left')
        matched = df_base['DCR'].notna().sum() if 'DCR' in df_base.columns else 0
        logger.info(f"  + Quality: {matched}/{before} conductores emparejados")
    
    # Merge con Daily Report (Complemento si falta Quality o para mayor detalle)
    if df_daily is not None and not df_daily.empty:
        # Calcular RTS pesado por entregados por día para que la media sea real
        df_daily['rts_count'] = df_daily['RTS'] * df_daily['Entregados']
        
        daily_agg = df_daily.groupby('ID').agg({
            'DNR': 'sum',
            'rts_count': 'sum',
            'Entregados': 'sum',
            'POD_Fails': 'sum',
            'CC_Fails': 'sum'
        }).reset_index()
        
        # Recalcular RTS real (total devueltos / total entregados)
        ent_d = daily_agg['Entregados']
        daily_agg['RTS'] = np.where(ent_d > 0, daily_agg['rts_count'] / ent_d, 0.0)

        df_base = df_base.merge(daily_agg, on='ID', how='left', suffixes=('', '_daily'))

        matched = df_base['Entregados_daily'].notna().sum()
        logger.info(f"  + Daily Report: {matched}/{len(df_base)} conductores emparejados")

        if 'DNR_daily' in df_base.columns:
            dnr_d = pd.to_numeric(df_base['DNR_daily'], errors='coerce')
            df_base['DNR'] = np.where(
                dnr_d.notna(),
                np.maximum(pd.to_numeric(df_base['DNR'], errors='coerce').fillna(0), dnr_d.fillna(0)),
                df_base['DNR']
            )

        if 'Entregados_daily' in df_base.columns:
            mask_e = df_base['Entregados'].isna() | (df_base['Entregados'] == 0)
            df_base['Entregados'] = np.where(mask_e, df_base['Entregados_daily'], df_base['Entregados'])

        if 'RTS_daily' in df_base.columns:
            mask_r = df_base['RTS'].isna() | (df_base['RTS'] == 0)
            df_base['RTS'] = np.where(mask_r, df_base['RTS_daily'], df_base['RTS'])

        if 'Entregados_daily' in df_base.columns:
            if 'POD' not in df_base.columns:
                df_base['POD'] = np.nan
            if 'CC' not in df_base.columns:
                df_base['CC'] = np.nan

            ent_col = pd.to_numeric(df_base['Entregados_daily'], errors='coerce')
            for _metric, _fails in [('POD', 'POD_Fails'), ('CC', 'CC_Fails')]:
                if _fails in df_base.columns:
                    _calc = (1.0 - pd.to_numeric(df_base[_fails], errors='coerce') / ent_col).clip(0.0, 1.0)
                    _calc = np.where(ent_col > 0, _calc, np.nan)
                    _mask = df_base[_metric].isna() | (df_base[_metric] == 1.0)
                    df_base[_metric] = np.where(_mask, _calc, df_base[_metric])
            
    for col in ['DCR', 'POD', 'CC', 'CDF', 'FDPS', 'RTS', 'IADC']:
        default_val = getattr(Config, f'DEFAULT_{col}', 1.0 if col != 'RTS' else 0.0)
        if col in df_base.columns:
            df_base[col] = pd.to_numeric(df_base[col], errors='coerce').clip(0.0, 1.0).fillna(default_val)
        else:
            df_base[col] = default_val

    df_base['DNR'] = pd.to_numeric(df_base['DNR'], errors='coerce').fillna(0.0).clip(upper=Config.MAX_DNR)
    
    # Merge con False Scan
    if df_false_scan is not None and not df_false_scan.empty:
        before = len(df_base)
        df_base = df_base.merge(df_false_scan, on='ID', how='left')
        if 'FS_Count' in df_base.columns:
            matched = df_base['FS_Count'].notna().sum()
            logger.info(f"  + False Scan: {matched}/{before} conductores emparejados")
            df_base['FS_Count'] = df_base['FS_Count'].fillna(Config.DEFAULT_FS).astype(int)
        else:
            df_base['FS_Count'] = Config.DEFAULT_FS
    else:
        df_base['FS_Count'] = Config.DEFAULT_FS
    
    # Merge con DWC/IADC (eventos de riesgo + porcentajes)
    if df_dwc is not None and not df_dwc.empty:
        before = len(df_base)
        df_base = df_base.merge(df_dwc, on='ID', how='left')
        
        # Si tenemos CC_DWC de este archivo, podemos usarlo para mejorar el CC
        if 'CC_DWC' in df_base.columns:
            df_base['CC'] = df_base['CC_DWC'].fillna(df_base['CC'])
            
        if 'DNR_RISK_EVENTS' in df_base.columns:
            matched = df_base['DNR_RISK_EVENTS'].notna().sum()
            logger.info(f"  + DWC/IADC: {matched}/{before} conductores emparejados")
            df_base['DNR_RISK_EVENTS'] = df_base['DNR_RISK_EVENTS'].fillna(0).astype(int)
        else:
            df_base['DNR_RISK_EVENTS'] = 0
    else:
        df_base['DNR_RISK_EVENTS'] = 0
        df_base['IADC'] = 0.0
    
    # Merge con FDPS
    if df_fdps is not None and not df_fdps.empty:
        before = len(df_base)
        df_base = df_base.merge(df_fdps, on='ID', how='left')
        if 'FDPS' in df_base.columns:
            df_base['FDPS'] = df_base['FDPS'].fillna(Config.DEFAULT_FDPS)
        else:
            df_base['FDPS'] = Config.DEFAULT_FDPS
    else:
        df_base['FDPS'] = Config.DEFAULT_FDPS
    
    # Validar resultado final
    logger.info(f"✓ Merge completado: {len(df_base)} conductores en dataset final")
    
    return df_base

# ═══════════════════════════════════════════════════════════════
# SISTEMA DE SCORING CON VALIDACIONES
# ═══════════════════════════════════════════════════════════════

def calculate_score_v3_robust(row: pd.Series, targets: Optional[Dict] = None) -> Tuple[str, str, int]:
    """
    Sistema de scoring V3 con validaciones robustas
    Usa targets dinámicos si se proporcionan
    """
    # Targets por defecto desde Config (fuente única de verdad)
    t = dict(Config.DEFAULT_TARGETS)
    if targets:
        t.update(targets)
    
    # Obtener valores con validación
    dnr = safe_number(row.get('DNR', 0), default=0)
    fs_count = safe_number(row.get('FS_Count', 0), default=0)
    dnr_risk = safe_number(row.get('DNR_RISK_EVENTS', 0), default=0)
    dcr = safe_number(row.get('DCR', 1.0), default=1.0)
    pod = safe_number(row.get('POD', 1.0), default=1.0)
    cc = safe_number(row.get('CC', 1.0), default=1.0)
    fdps = safe_number(row.get('FDPS', 1.0), default=1.0)
    rts = safe_number(row.get('RTS', 0.0), default=0.0)
    cdf = safe_number(row.get('CDF', 1.0), default=1.0)
    
    # Validar rangos
    dnr = max(0, min(dnr, Config.MAX_DNR))
    fs_count = max(0, min(fs_count, Config.MAX_FALSE_SCAN))
    dcr = max(0, min(dcr, 1.0))
    pod = max(0, min(pod, 1.0))
    cc = max(0, min(cc, 1.0))
    fdps = max(0, min(fdps, 1.0))
    rts = max(0, min(rts, 1.0))
    cdf = max(0, min(cdf, 1.0))
    
    issues = []
    score = 100
    
    # === DNR (peso x3) ===
    if dnr >= 4:
        issues.append(f"🚨 {int(dnr)} DNR (CRÍTICO)")
        score -= 70
    elif dnr >= 3:
        issues.append(f"🚨 {int(dnr)} DNR (MUY GRAVE)")
        score -= 60
    elif dnr > t['target_dnr'] * 3:
        issues.append(f"⚡ {int(dnr)} DNR")
        score -= 25
    elif dnr > t['target_dnr']:
        issues.append(f"⚡ {int(dnr)} DNR")
        score -= 15
    
    # === FALSE SCANS ===
    fs_int = int(fs_count)
    if fs_int >= 100:
        issues.append(f"❌ {fs_int} FS (CRÍTICO)")
        score -= 40
    elif fs_int >= 20:
        issues.append(f"⚠️ {fs_int} FS")
        score -= 20
    elif fs_int >= 5:
        issues.append(f"ℹ️ {fs_int} FS")
        score -= 5
    
    # === DCR ===
    if dcr < t['target_dcr'] - 0.05:
        issues.append(f"📦 DCR Crítico {dcr:.1%}")
        score -= 40
    elif dcr < t['target_dcr']:
        issues.append(f"📦 DCR Bajo {dcr:.1%}")
        score -= 15
    
    # === POD ===
    if pod < t['target_pod'] - 0.10:
        issues.append(f"📸 POD Crítico {pod:.1%}")
        score -= 25
    elif pod < t['target_pod']:
        issues.append(f"📸 POD Bajo {pod:.1%}")
        score -= 10
    
    # === CC ===
    if cc < t['target_cc']:
        issues.append(f"📞 CC Bajo {cc:.1%}")
        score -= 10
    
    # === FDPS ===
    if fdps < t['target_fdps']:
        issues.append(f"🚚 FDPS Bajo {fdps:.1%}")
        score -= 10

    # === RTS ===
    if rts > t['target_rts'] * 2:
        issues.append(f"🔄 RTS Alto {rts:.1%}")
        score -= 15
    elif rts > t['target_rts']:
        issues.append(f"🔄 RTS {rts:.1%}")
        score -= 8
    
    # === CDF ===
    if cdf < t['target_cdf']:
        issues.append(f"⭐ CDF Bajo {cdf:.1%}")
        score -= 15
    
    # Asegurar score mínimo 0
    score = max(0, score)
    
    # Determinar calificación final
    if score >= 93:
        calificacion = "🌟 FANTASTIC+"
    elif score >= 90:
        calificacion = "💎 FANTASTIC"
    elif score >= 80:
        calificacion = "🥇 GREAT"
    elif score >= 60:
        calificacion = "⚠️ FAIR"
    else:
        calificacion = "🛑 POOR"
    
    return calificacion, ", ".join(issues) if issues else "Óptimo", score

# ═══════════════════════════════════════════════════════════════
# GENERACIÓN DE EXCEL
# ═══════════════════════════════════════════════════════════════

def extract_info_from_path(path: str) -> Tuple[str, str, Optional[int]]:
    """Extrae semana, centro y año del nombre del archivo con normalización inteligente"""
    if not path:
        return "N/A", "TDSL", None
    
    filename = os.path.basename(path)
    
    # 1. Normalizar Semana
    week = "N/A"
    
    # Patrón estándar: W05, Week 5, Semana 05, S5
    week_match = re.search(r'(?:W|Week|Semana|S)[_\s-]*(\d+)', filename, re.IGNORECASE)
    if week_match:
        num = int(week_match.group(1))
        if 1 <= num <= 53:
            week = f"W{num:02d}"
    else:
        # Intentar buscar por formato de fecha YYYY-MM-DD
        date_match = re.search(r'(\d{4})[_-](\d{1,2})[_-](\d{1,2})', filename)
        if date_match:
            try:
                from datetime import timedelta
                dt = datetime(int(date_match.group(1)), int(date_match.group(2)), int(date_match.group(3)))
                # Shifting 1 day forward to align Sunday (Amazon start) with Monday (ISO start)
                # This makes Sun-Sat belong to the same ISO week
                dt_shifted = dt + timedelta(days=1)
                isoyear, isoweek, isoday = dt_shifted.isocalendar()
                week = f"W{isoweek:02d}"
            except (ValueError, TypeError, AttributeError):
                pass  # Fecha inválida — continuar con otros patrones
        
        if week == "N/A":
            # Formato YYYY-MM: tratar como 1er día del mes y convertir a semana ISO
            # Evita interpretar "05" como W05 cuando es mes 5 (≈ W18-W22)
            date_month_match = re.search(r'(202\d)[_-](\d{1,2})(?!\d)', filename)
            if date_month_match:
                try:
                    _yr  = int(date_month_match.group(1))
                    _mo  = int(date_month_match.group(2))
                    if 1 <= _mo <= 12:
                        _dt = datetime(_yr, _mo, 1)
                        _, _isow, _ = _dt.isocalendar()
                        week = f"W{_isow:02d}"
                except (ValueError, TypeError):
                    pass
    
    # 2. Normalizar Centro (ej: DIC1, VLC1, MAD1, DMA3, ES-TDSL-DIC1)
    # Busca 3-4 letras seguidas de uno o más dígitos
    center_match = re.search(r'([A-Z]{3,4}\d+)', filename, re.IGNORECASE)
    if center_match:
        center = center_match.group(1).upper()
    else:
        # Buscar en el path si no está en el filename
        center_match_path = re.search(r'([A-Z]{3,4}\d+)', path, re.IGNORECASE)
        center = center_match_path.group(1).upper() if center_match_path else "TDSL"

    # 3. Extraer año si está explícito en el nombre (ej: 2025, 2026...)
    year_extracted: Optional[int] = None
    year_match = re.search(r'\b(20\d{2})\b', filename)
    if not year_match:
        year_match = re.search(r'\b(20\d{2})\b', path)
    if year_match:
        year_extracted = int(year_match.group(1))
    else:
        year_extracted = datetime.now().year

    return week, center, year_extracted

def create_professional_excel(df: pd.DataFrame, output_path: str, 
                              center_name: str = "TDSL", week: str = "N/A") -> bool:
    """
    Genera un Scorecard de Excel con diseño de Dashboard profesional optimizado.
    """
    if df is None or df.empty:
        logger.warning("No hay datos para generar el Excel.")
        return False
        
    try:
        logger.info(f"Generando Dashboard Excel para {len(df)} conductores...")
        
        wb = Workbook()
        wb.remove(wb.active)
        
        # --- PALETA DE COLORES AMAZON ---
        AMZ_DARK = '232F3E'
        AMZ_ORANGE = 'FF9900'
        AMZ_BLUE = '146EB4'
        WHITE = 'FFFFFF'
        
        # Estilos Base
        font_header = Font(name='Segoe UI', size=11, bold=True, color=WHITE)
        font_title = Font(name='Segoe UI', size=16, bold=True, color=WHITE)
        fill_header = PatternFill(start_color=AMZ_DARK, end_color=AMZ_DARK, fill_type='solid')
        fill_orange = PatternFill(start_color=AMZ_ORANGE, end_color=AMZ_ORANGE, fill_type='solid')
        align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
        border_thin = Border(
            left=Side(style='thin', color='DDDDDD'),
            right=Side(style='thin', color='DDDDDD'),
            top=Side(style='thin', color='DDDDDD'),
            bottom=Side(style='thin', color='DDDDDD')
        )

        # 1. 📊 HOJA RESUMEN (DASHBOARD)
        ws_summary = wb.create_sheet('📊 DASHBOARD RESUMEN', 0)
        ws_summary.sheet_view.showGridLines = False
        
        # Título Principal (Corregido merge a columna I para cubrir 4 tarjetas)
        ws_summary.merge_cells('B2:I3')
        title_cell = ws_summary['B2']
        title_cell.value = f'QUALITY SCORECARD - {center_name} - {week}'
        title_cell.font = font_title
        title_cell.fill = fill_header
        title_cell.alignment = align_center
        
        # --- CARDS DE MÉTRICAS CLAVE ---
        calif_counts = df['CALIFICACION'].value_counts()
        total_drv = len(df)
        
        def create_metric_card(ws, start_col, start_row, label, value, subtext="", color="146EB4"):
            fill_color = PatternFill(start_color=color, end_color=color, fill_type='solid')
            
            # Label box
            cell_label = ws.cell(row=start_row, column=start_col)
            cell_label.value = label
            cell_label.font = Font(name='Segoe UI', size=9, bold=True, color=WHITE)
            cell_label.fill = fill_color
            cell_label.alignment = align_center
            
            # Value box
            cell_val = ws.cell(row=start_row+1, column=start_col)
            cell_val.value = value
            cell_val.font = Font(name='Segoe UI', size=18, bold=True)
            cell_val.alignment = align_center
            
            # Subtext box
            cell_sub = ws.cell(row=start_row+2, column=start_col)
            cell_sub.value = subtext
            cell_sub.font = Font(name='Segoe UI', size=8, italic=True, color='666666')
            cell_sub.alignment = align_center

        row_cards = 5
        # ✅ CÁLCULO CORRECTO DE MÉTRICAS (en Python, no con fórmulas Excel)
        # Esto evita el bug de incluir la fila de totales en los promedios
        dnr_total = int(df['DNR'].sum()) if 'DNR' in df.columns else 0
        dnr_promedio = float(df['DNR'].mean()) if 'DNR' in df.columns else 0.0
        fs_total = int(df['FS_Count'].sum()) if 'FS_Count' in df.columns else 0
        fs_promedio = float(df['FS_Count'].mean()) if 'FS_Count' in df.columns else 0.0
        dcr_avg = float(df['DCR'].mean()) if 'DCR' in df.columns else 1.0
        pod_avg = float(df['POD'].mean()) if 'POD' in df.columns else 1.0
        
        metrics_to_show = [
            ("TOTAL CONDUCTORES", total_drv, f"Activos en {center_name}", AMZ_DARK),
            ("DNR PROMEDIO", f"{dnr_promedio:.2f}", f"Total: {dnr_total}", "E53935"),
            ("FALSE SCANS", f"{fs_promedio:.1f}", f"Total: {fs_total}", "FB8C00"),
            ("DCR PROMEDIO", f"{dcr_avg:.2%}", "Calidad Entrega", "43A047")
        ]
        
        for i, (l, v, s, c) in enumerate(metrics_to_show):
            col_pos = 2 + (i*2)
            create_metric_card(ws_summary, col_pos, row_cards, l, v, s, c)
            ws_summary.merge_cells(start_row=row_cards, start_column=col_pos, end_row=row_cards, end_column=col_pos+1)
            ws_summary.merge_cells(start_row=row_cards+1, start_column=col_pos, end_row=row_cards+2, end_column=col_pos+1)

        # --- DISTRIBUCIÓN DE CALIFICACIONES ---
        ws_summary.cell(row=9, column=2, value="ESTADO DE LA FLOTA").font = Font(bold=True, size=11)
        dist_row = 10
        dist_colors = {'🌟 FANTASTIC+': '7c3aed', '💎 FANTASTIC': '4CAF50', '🥇 GREAT': '8BC34A', '⚠️ FAIR': 'FFC107', '🛑 POOR': 'F44336'}
        
        for i, (cat, color) in enumerate(dist_colors.items()):
            count = int(calif_counts.get(cat, 0))
            pct = count / total_drv if total_drv > 0 else 0
            
            curr_r = dist_row + i
            ws_summary.cell(row=curr_r, column=2, value=cat).font = Font(bold=True)
            ws_summary.cell(row=curr_r, column=3, value=count).alignment = Alignment(horizontal='center')
            ws_summary.cell(row=curr_r, column=4, value=pct).number_format = '0%'
            # Barra visual
            ws_summary.cell(row=curr_r, column=5).fill = PatternFill(start_color=color, end_color=color, fill_type='solid')

        # 2. 📋 HOJA DETALLE
        ws_detail = wb.create_sheet('📋 DETALLE COMPLETO', 1)
        ws_detail.merge_cells('A1:O2')
        detail_header = ws_detail['A1']
        detail_header.value = f"REPORTE DETALLADO CONDUCTORES - {center_name} - SEMANA {week}"
        detail_header.font = font_title
        detail_header.fill = fill_header
        detail_header.alignment = align_center
        
        headers = ['CONDUCTOR', 'ID AGENTE', 'STATUS', 'SCORE', 'ENTREGADOS', 'DNR', 'FALSE SCAN',
                   'RIESGO DNR', 'DCR %', 'POD %', 'CC %', 'FDPS %', 'RTS %', 'CDF %', 'OBSERVACIONES']
        
        for col, h in enumerate(headers, 1):
            cell = ws_detail.cell(row=4, column=col, value=h)
            cell.font = font_header
            cell.fill = fill_orange
            cell.alignment = align_center
            cell.border = border_thin
        
        df_sorted = df.sort_values(['SCORE', 'DNR'], ascending=[False, True]).reset_index(drop=True)
        for idx, (_, row_data) in enumerate(df_sorted.iterrows(), start=5):
            ws_detail.cell(row=idx, column=1, value=str(row_data['Nombre']))
            ws_detail.cell(row=idx, column=2, value=str(row_data['ID']))
            
            status = str(row_data['CALIFICACION'])
            c_status = ws_detail.cell(row=idx, column=3, value=status)
            c_status.font = Font(bold=True)
            if '💎' in status: c_status.font = Font(bold=True, color='2E7D32')
            elif '🛑' in status: c_status.font = Font(bold=True, color='C62828')
            
            ws_detail.cell(row=idx, column=4, value=float(row_data['SCORE'])).number_format = '0'
            ws_detail.cell(row=idx, column=5, value=float(row_data['Entregados'])).number_format = '#,##0'
            ws_detail.cell(row=idx, column=6, value=float(row_data['DNR'])).number_format = '0'
            ws_detail.cell(row=idx, column=7, value=float(row_data['FS_Count'])).number_format = '#,##0'
            ws_detail.cell(row=idx, column=8, value=float(row_data['DNR_RISK_EVENTS'])).number_format = '0'
            
            for col_idx, col_name in enumerate(['DCR', 'POD', 'CC', 'FDPS', 'RTS', 'CDF'], start=9):
                cell = ws_detail.cell(row=idx, column=col_idx, value=float(row_data[col_name]))
                cell.number_format = '0.0%'
            
            cell_obs = ws_detail.cell(row=idx, column=15, value=str(row_data['DETALLES']))
            cell_obs.alignment = Alignment(horizontal='left', wrap_text=True, vertical='center')
            
            for col in range(1, 16): 
                ws_detail.cell(row=idx, column=col).border = border_thin

        # Formato Condicional Optimizado
        last_row = len(df_sorted) + 4
        if last_row >= 5:
            for col in ['D', 'I', 'J', 'K', 'L', 'N']:
                ws_detail.conditional_formatting.add(f'{col}5:{col}{last_row}',
                    ColorScaleRule(start_type='num', start_value=0.7, start_color='F8696B',
                                   mid_type='num', mid_value=0.9, mid_color='FFEB84',
                                   end_type='num', end_value=1.0, end_color='63BE7B'))
            for col in ['F', 'G', 'H', 'M']:
                ws_detail.conditional_formatting.add(f'{col}5:{col}{last_row}',
                    ColorScaleRule(start_type='percentile', start_value=0, start_color='63BE7B',
                                   mid_type='percentile', mid_value=85, mid_color='FFEB84',
                                   end_type='percentile', end_value=98, end_color='F8696B'))

        ws_detail.freeze_panes = 'A5'
        ws_detail.auto_filter.ref = f'A4:O{last_row}'
        
        # Ajuste de anchos
        col_widths = {'A': 28, 'B': 14, 'C': 15, 'D': 8, 'E': 12, 'O': 55}
        for col, width in col_widths.items(): ws_detail.column_dimensions[col].width = width
        for col in 'FGHIJKLMN': ws_detail.column_dimensions[col].width = 11

        # 3. 🏆 HOJA RANKING & ALERTAS
        ws_rank = wb.create_sheet('🏆 RANKING & ALERTAS', 2)
        ws_rank.sheet_view.showGridLines = False
        
        # Estilos específicos para rankings
        fill_green = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
        font_green = Font(color='006100', bold=True)
        fill_red = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
        font_red = Font(color='9C0006', bold=True)
        
        # TOP 10 FANTASTIC
        ws_rank.merge_cells('B2:D2')
        ws_rank['B2'] = "⭐ TOP 10 CONDUCTORES (FANTASTIC)"
        ws_rank['B2'].font = Font(name='Segoe UI', size=12, bold=True, color='2E7D32')
        
        headers_rank = ['Nombre', 'Score', 'DNR']
        for col, h in enumerate(headers_rank, 2):
            cell = ws_rank.cell(row=3, column=col, value=h)
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = align_center
            
        top_10 = df_sorted.sort_values(['SCORE', 'DNR'], ascending=[False, True]).head(10)
        for idx, (_, r) in enumerate(top_10.iterrows(), 4):
            ws_rank.cell(row=idx, column=2, value=r['Nombre']).fill = fill_green
            ws_rank.cell(row=idx, column=3, value=r['SCORE']).font = font_green
            ws_rank.cell(row=idx, column=4, value=r['DNR'])
            
        # BOTTOM 10 / RED FLAGS
        ws_rank.merge_cells('F2:H2')
        ws_rank['F2'] = "🚨 RED FLAGS (POOR/FAIR)"
        ws_rank['F2'].font = Font(name='Segoe UI', size=12, bold=True, color='C62828')
        
        for col, h in enumerate(headers_rank, 6):
            cell = ws_rank.cell(row=3, column=col, value=h)
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = align_center
            
        bottom_10 = df_sorted.sort_values(['SCORE', 'DNR'], ascending=[True, False]).head(15)
        for idx, (_, r) in enumerate(bottom_10.iterrows(), 4):
            ws_rank.cell(row=idx, column=6, value=r['Nombre']).fill = fill_red
            ws_rank.cell(row=idx, column=7, value=r['SCORE']).font = font_red
            ws_rank.cell(row=idx, column=8, value=r['DNR'])

        ws_rank.column_dimensions['B'].width = 35
        ws_rank.column_dimensions['F'].width = 35
        
        # 4. 💬 HOJA FEEDBACK (PULIDO)
        ws_feed = wb.create_sheet('💬 FEEDBACK', 3)
        ws_feed.append(['CONDUCTOR', 'CALIFICACIÓN', 'PUNTOS DE MEJORA'])
        for col, h in enumerate(['CONDUCTOR', 'CALIFICACIÓN', 'PUNTOS DE MEJORA'], 1):
            cell = ws_feed.cell(row=1, column=col)
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = align_center
            
        for _, r in df_sorted.iterrows(): 
            ws_feed.append([r['Nombre'], r['CALIFICACION'], r['DETALLES']])
            
        ws_feed.column_dimensions['A'].width = 30
        ws_feed.column_dimensions['B'].width = 15
        ws_feed.column_dimensions['C'].width = 85
        for row_idx in range(2, len(df_sorted) + 2):
            ws_feed.cell(row=row_idx, column=3).alignment = Alignment(wrap_text=True, vertical='center')

        wb.save(output_path)
        logger.info(f"✅ Dashboard Excel Optimizado guardado en: {output_path}")
        return True
    except Exception as e:
        logger.error(f"❌ Error Dashboard Excel: {str(e)}")
        return False

def delete_scorecard_batch(week: str, center: str, db_config: Optional[Dict] = None,
                           year: Optional[int] = None, preserve_pdf: bool = False) -> bool:
    """Elimina los datos de una semana y centro específicos (Para corregir errores de volcado).
    Si year es None, elimina TODOS los años para ese centro+semana.
    Si year es un int, elimina solo ese año concreto.
    Si preserve_pdf=True, conserva filas con pdf_loaded=1 (datos oficiales del PDF).
    """
    try:
        is_pg = db_config and db_config.get('type') == 'postgresql'
        ph = '%s' if is_pg else '?'
        pdf_filter = " AND (pdf_loaded IS NULL OR pdf_loaded = 0)" if preserve_pdf else ""
        with db_connection(db_config) as conn:
            cursor = conn.cursor()
            if year is not None:
                q = (f"DELETE FROM scorecards WHERE semana = {ph} AND centro = {ph} AND anio = {ph}{pdf_filter}")
                cursor.execute(q, (week, center, year))
            else:
                q = (f"DELETE FROM scorecards WHERE semana = {ph} AND centro = {ph}{pdf_filter}")
                cursor.execute(q, (week, center))
            rows = cursor.rowcount
            conn.commit()
        logger.info(f"🗑️ Se eliminaron {rows} registros de {center} semana {week}" +
                    (f" año {year}" if year else " (todos los años)") +
                    (" (preservando filas PDF)" if preserve_pdf else ""))
        return True
    except Exception as e:
        logger.error(f"Error eliminando lote: {e}")
        return False

def find_file_in_dir(pattern: str, directory: str) -> Optional[str]:
    """Busca un archivo en un directorio específico usando un patrón"""
    for file in os.listdir(directory):
        if re.match(pattern, file, re.IGNORECASE):
            return os.path.join(directory, file)
    return None

def process_single_batch(path_concessions, path_quality=None, path_false_scan=None,
                         path_dwc=None, path_fdps=None, path_daily=None,
                         path_dsc_concessions=None, targets=None) -> Optional[pd.DataFrame]:
    """Procesa un único lote de archivos y devuelve el DataFrame final.
    Soporta rutas individuales o listas de rutas/buffers.
    path_dsc_concessions: archivo DSC-Concessions (se excluye del cálculo DNR de conductores)."""
    try:
        # Función auxiliar para leer uno o varios archivos y concatenarlos
        def read_multiple(paths):
            if paths is None: return None
            if not isinstance(paths, list): paths = [paths]
            
            valid_dfs = []
            for p in paths:
                df = read_any_safe(p, str(p))
                if df is not None and not df.empty:
                    valid_dfs.append(df)
            
            if not valid_dfs: return None
            return pd.concat(valid_dfs, ignore_index=True) if len(valid_dfs) > 1 else valid_dfs[0]

        df_concessions = read_multiple(path_concessions)
        df_quality = read_multiple(path_quality)
        df_false_scan_html = read_multiple(path_false_scan)
        df_dwc = read_multiple(path_dwc)
        df_fdps = read_multiple(path_fdps)
        df_daily_raw = read_multiple(path_daily)

        # DSC-Concessions: se lee y se registra como info, NO entra en el cálculo de DNR
        df_dsc = read_multiple(path_dsc_concessions)
        if df_dsc is not None and not df_dsc.empty:
            logger.info(f"DSC-Concessions detectado ({len(df_dsc)} filas) — excluido del cálculo DNR de conductores")
        
        if df_concessions is None or df_concessions.empty:
            logger.error("No se pudo leer el archivo de Concessions (obligatorio)")
            return None
            
        df_conc_clean = process_concessions(df_concessions)
        
        df_qual_clean = process_quality(df_quality) if df_quality is not None and not df_quality.empty else None
        df_fs_clean = process_false_scan(df_false_scan_html) if df_false_scan_html is not None and not df_false_scan_html.empty else None
        df_dwc_clean = process_dwc(df_dwc) if df_dwc is not None and not df_dwc.empty else None
        df_fdps_clean = process_fdps(df_fdps) if df_fdps is not None and not df_fdps.empty else None
        df_daily_clean = process_daily_report(df_daily_raw) if df_daily_raw is not None and not df_daily_raw.empty else None
        
        df_merged = merge_data_smart(df_conc_clean, df_qual_clean, df_fs_clean, df_dwc_clean, df_fdps_clean, df_daily_clean)
        
        if 'DNR' in df_merged.columns:
            df_merged['DNR'] = pd.to_numeric(df_merged['DNR'], errors='coerce').fillna(0.0).clip(upper=Config.MAX_DNR)
        if 'FS_Count' in df_merged.columns:
            df_merged['FS_Count'] = pd.to_numeric(df_merged['FS_Count'], errors='coerce').fillna(0.0).clip(upper=Config.MAX_FALSE_SCAN)

        # Eliminar duplicados finales por si acaso
        if 'ID' in df_merged.columns:
            df_merged = df_merged.drop_duplicates(subset='ID', keep='first')
            
        results = df_merged.apply(lambda x: calculate_score_v3_robust(x, targets=targets), axis=1, result_type='expand')
        df_merged['CALIFICACION'] = results[0]
        df_merged['DETALLES'] = results[1]
        df_merged['SCORE'] = results[2]
        
        return df_merged
    except Exception as e:
        logger.error(f"Error en procesamiento: {str(e)}")
        import traceback
        traceback.print_exc()
        return None

# ═══════════════════════════════════════════════════════════════
# BASE DE DATOS (FASE 2/3 - SQLite & PostgreSQL)
# ═══════════════════════════════════════════════════════════════

def get_db_connection(db_config: Optional[Dict] = None):
    """Crea o conecta a la base de datos (PostgreSQL o SQLite)"""
    if db_config and db_config.get('type') == 'postgresql':
        if not HAS_POSTGRES:
            raise ImportError("Librería 'psycopg2' no encontrada. Instálala con: pip install psycopg2-binary")
        # Usar pool — reutiliza conexiones en vez de abrir una nueva cada vez
        _pg_kwargs = dict(
            host=db_config.get('host', 'localhost'),
            database=db_config.get('database', 'postgres'),
            user=db_config.get('user', 'postgres'),
            password=db_config.get('password', ''),
            port=db_config.get('port', 5432),
            connect_timeout=10,
            sslmode='require',
            keepalives=1,
            keepalives_idle=30,
            keepalives_interval=10,
            keepalives_count=5,
            options="-c statement_timeout=30000",
        )
        try:
            _pool = _get_pg_pool(db_config)
            conn = _pool.getconn()
            # Validar que la conexión sigue viva; si está rota, reconectar
            try:
                conn.cursor().execute("SELECT 1")
            except Exception:
                try:
                    _pool.putconn(conn, close=True)
                except Exception:
                    pass
                conn = psycopg2.connect(**_pg_kwargs)
            return conn
        except Exception:
            # Fallback a conexión directa si el pool falla
            return psycopg2.connect(**_pg_kwargs)
    else:
        if db_config and db_config.get('path'):
            db_path = db_config['path']
        else:
            db_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "amazon_quality.db")
        conn = sqlite3.connect(db_path)
        conn.row_factory = sqlite3.Row
        return conn


@contextmanager
def db_connection(db_config: Optional[Dict] = None):
    """
    Context manager para conexiones a BD.
    Para PostgreSQL: devuelve la conexión al pool al terminar (no la cierra).
    Para SQLite: cierra la conexión al terminar.
    """
    conn = get_db_connection(db_config)
    is_pg = db_config and db_config.get('type') == 'postgresql'
    try:
        yield conn
    except Exception:
        try:
            conn.rollback()
        except Exception as _e:
            logger.debug(f"db_connection rollback: {_e}")
        raise
    finally:
        try:
            if is_pg and _PG_POOL and not _PG_POOL.closed:
                _PG_POOL.putconn(conn)   # devolver al pool, no cerrar
            else:
                conn.close()
        except Exception as _e:
            logger.debug(f"db_connection finally: {_e}")


def hash_password(password: str) -> str:
    """
    Encripta la contraseña usando bcrypt (preferido) o SHA-256 (fallback).
    
    Args:
        password: Contraseña en texto plano
        
    Returns:
        Hash de la contraseña
        
    Notes:
        - bcrypt es más seguro (salt automático, resistente a rainbow tables)
        - SHA-256 se usa como fallback si bcrypt no está instalado
    """
    if HAS_BCRYPT:
        # Usar bcrypt con salt automático (más seguro)
        salt = bcrypt.gensalt()
        hashed = bcrypt.hashpw(password.encode('utf-8'), salt)
        # Agregar prefijo para identificar que es bcrypt
        return 'bcrypt:' + hashed.decode('utf-8')
    else:
        # Fallback a SHA-256 (menos seguro pero funcional)
        logger.warning("Usando SHA-256 para hash de contraseña. Considera instalar bcrypt para mejor seguridad.")
        return 'sha256:' + hashlib.sha256(password.encode()).hexdigest()

def verify_password(password: str, hashed: str) -> bool:
    """
    Verifica una contraseña contra su hash.
    
    Args:
        password: Contraseña en texto plano
        hashed: Hash almacenado en base de datos
        
    Returns:
        True si la contraseña coincide, False en caso contrario
        
    Notes:
        - Detecta automáticamente el método de hash usado (bcrypt o SHA-256)
        - Compatible con hashes antiguos (SHA-256) y nuevos (bcrypt)
    """
    try:
        if hashed.startswith('bcrypt:'):
            if HAS_BCRYPT:
                return bcrypt.checkpw(
                    password.encode('utf-8'),
                    hashed.removeprefix('bcrypt:').encode('utf-8')
                )
            logger.error("Hash es bcrypt pero biblioteca no está instalada")
            return False
        elif hashed.startswith('sha256:'):
            return hashlib.sha256(password.encode()).hexdigest() == hashed.removeprefix('sha256:')
        else:
            # Compatibilidad con hashes antiguos sin prefijo (asumimos SHA-256)
            test_hash = hashlib.sha256(password.encode()).hexdigest()
            return test_hash == hashed
    except Exception as e:
        logger.error(f"Error verificando contraseña: {e}")
        return False


def update_user_password(username: str, new_hash: str, db_config: dict) -> bool:
    """Actualiza el hash de contraseña y limpia must_change_password (primer login obligatorio)."""
    try:
        with db_connection(db_config) as conn:
            cursor = conn.cursor()
            q = ("UPDATE users SET password = %s, must_change_password = 0 WHERE username = %s"
                 if db_config and db_config.get('type') == 'postgresql' else
                 "UPDATE users SET password = ?, must_change_password = 0 WHERE username = ?")
            cursor.execute(q, (new_hash, username))
            conn.commit()
        return True
    except Exception as e:
        logger.warning(f"update_user_password error: {e}")
        return False



def init_database(db_config: Optional[Dict] = None):
    """Inicializa las tablas optimizadas para producción en PostgreSQL o SQLite"""
    conn = None
    try:
        conn = get_db_connection(db_config)
        cursor = conn.cursor()
        is_postgres = db_config and db_config.get('type') == 'postgresql'
        
        # 1. Tabla de Scorecards (Optimizada para Power BI)
        if is_postgres:
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS scorecards (
                    id SERIAL PRIMARY KEY,
                    semana VARCHAR(10),
                    fecha_semana DATE,
                    anio INTEGER,
                    centro VARCHAR(20),
                    driver_id VARCHAR(50),
                    driver_name VARCHAR(255),
                    calificacion VARCHAR(50),
                    score DOUBLE PRECISION,
                    entregados DOUBLE PRECISION,
                    dnr DOUBLE PRECISION,
                    fs_count DOUBLE PRECISION,
                    dnr_risk_events DOUBLE PRECISION,
                    dcr DOUBLE PRECISION,
                    pod DOUBLE PRECISION,
                    cc DOUBLE PRECISION,
                    fdps DOUBLE PRECISION,
                    rts DOUBLE PRECISION,
                    cdf DOUBLE PRECISION,
                    detalles TEXT,
                    uploaded_by VARCHAR(100),
                    timestamp TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    UNIQUE(semana, centro, anio, driver_id)
                )
            ''')
            # Índices PostgreSQL — algunos nombres coinciden con SQLite (if/else separados, no hay duplicado real)
            # PG: soporta INCLUDE (covering), WHERE parcial, LOWER() funcional. SQLite: sintaxis básica.
            pg_indexes = [
                # Filtro principal app: centro + semana
                "CREATE INDEX IF NOT EXISTS idx_centro_semana       ON scorecards (centro, semana)",
                # Power BI: covering index con métricas
                "CREATE INDEX IF NOT EXISTS idx_bi_query            ON scorecards (centro, semana, fecha_semana) INCLUDE (score, dnr, dcr, calificacion)",
                # Histórico ordenado por fecha
                "CREATE INDEX IF NOT EXISTS idx_fecha_desc          ON scorecards (fecha_semana DESC)",
                # Búsqueda por conductor
                "CREATE INDEX IF NOT EXISTS idx_driver_id           ON scorecards (driver_id)",
                "CREATE INDEX IF NOT EXISTS idx_driver_name         ON scorecards (LOWER(driver_name))",
                # Filtro por calificación (dashboard alertas)
                "CREATE INDEX IF NOT EXISTS idx_calificacion        ON scorecards (calificacion)",
                # Análisis temporal por centro
                "CREATE INDEX IF NOT EXISTS idx_centro_fecha        ON scorecards (centro, fecha_semana DESC)",
                # Ranking conductores por centro/semana
                "CREATE INDEX IF NOT EXISTS idx_ranking             ON scorecards (centro, semana, score DESC)",
                # Índice parcial: DNR crítico > 5
                "CREATE INDEX IF NOT EXISTS idx_dnr_alto            ON scorecards (centro, semana, dnr) WHERE dnr > 5",
                # Índice parcial: bajo rendimiento
                "CREATE INDEX IF NOT EXISTS idx_poor_fair           ON scorecards (centro, semana, score) WHERE score < 70",
                "CREATE INDEX IF NOT EXISTS idx_timestamp_desc      ON scorecards (timestamp DESC)",
                "CREATE INDEX IF NOT EXISTS idx_semana_timestamp    ON scorecards (semana, timestamp DESC)",
                "CREATE INDEX IF NOT EXISTS idx_anio                ON scorecards (anio)",
            ]
            for idx_sql in pg_indexes:
                try:
                    cursor.execute(idx_sql)
                except Exception as ie:
                    logger.warning(f"Índice ya existe o error menor: {ie}")
        else:
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS scorecards (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    semana TEXT, fecha_semana DATE, anio INTEGER, centro TEXT, driver_id TEXT, driver_name TEXT,
                    calificacion TEXT, score FLOAT, entregados FLOAT, dnr FLOAT,
                    fs_count FLOAT, dnr_risk_events FLOAT, dcr FLOAT, pod FLOAT,
                    cc FLOAT, fdps FLOAT, rts FLOAT, cdf FLOAT, detalles TEXT,
                    uploaded_by TEXT,
                    timestamp TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    UNIQUE(semana, centro, anio, driver_id)
                )
            ''')
            # Índices SQLite — sintaxis simplificada (no soporta INCLUDE ni WHERE en todos los casos)
            # Los nombres comunes con PG son intencionales (mismas tablas, mismos accesos)
            sqlite_indexes = [
                "CREATE INDEX IF NOT EXISTS idx_centro_semana  ON scorecards (centro, semana)",
                "CREATE INDEX IF NOT EXISTS idx_fecha_desc     ON scorecards (fecha_semana DESC)",
                "CREATE INDEX IF NOT EXISTS idx_driver_id      ON scorecards (driver_id)",
                "CREATE INDEX IF NOT EXISTS idx_driver_name    ON scorecards (driver_name COLLATE NOCASE)",
                "CREATE INDEX IF NOT EXISTS idx_calificacion   ON scorecards (calificacion)",
                "CREATE INDEX IF NOT EXISTS idx_centro_fecha   ON scorecards (centro, fecha_semana)",
                "CREATE INDEX IF NOT EXISTS idx_ranking        ON scorecards (centro, semana, score DESC)",
                "CREATE INDEX IF NOT EXISTS idx_centro_timestamp ON scorecards (centro, timestamp DESC)",
                "CREATE INDEX IF NOT EXISTS idx_anio           ON scorecards (anio)",
            ]
            for idx_sql in sqlite_indexes:
                try:
                    cursor.execute(idx_sql)
                except Exception as ie:
                    logger.warning(f"Índice SQLite: {ie}")

        # 2. Tabla de Usuarios
        user_id_type = "SERIAL PRIMARY KEY" if is_postgres else "INTEGER PRIMARY KEY AUTOINCREMENT"
        cursor.execute(f'''
            CREATE TABLE IF NOT EXISTS users (
                id {user_id_type},
                username VARCHAR(100) UNIQUE,
                password TEXT,
                role VARCHAR(20),
                active INTEGER DEFAULT 1,
                must_change_password INTEGER DEFAULT 0,
                centro_asignado VARCHAR(100) DEFAULT NULL
            )
        ''')
        # Índices de usuarios (se crean AQUÍ, después de que la tabla existe)
        users_indexes_pg = [
            "CREATE INDEX IF NOT EXISTS idx_users_username ON users (LOWER(username))",
            "CREATE INDEX IF NOT EXISTS idx_users_role     ON users (role) WHERE active = 1",
        ]
        users_indexes_sqlite = [
            "CREATE INDEX IF NOT EXISTS idx_users_username ON users (username)",
            "CREATE INDEX IF NOT EXISTS idx_users_role     ON users (role)",
        ]
        for idx_sql in (users_indexes_pg if is_postgres else users_indexes_sqlite):
            try:
                cursor.execute(idx_sql)
            except Exception as ie:
                logger.warning(f"Índice users: {ie}")

        # 3. Tabla de Targets (Optimizada)
        cursor.execute(f'''
            CREATE TABLE IF NOT EXISTS center_targets (
                centro VARCHAR(50) PRIMARY KEY,
                target_dnr DOUBLE PRECISION DEFAULT 0.5,
                target_dcr DOUBLE PRECISION DEFAULT 0.995,
                target_pod DOUBLE PRECISION DEFAULT 0.99,
                target_cc DOUBLE PRECISION DEFAULT 0.99,
                target_fdps DOUBLE PRECISION DEFAULT 0.98,
                target_rts DOUBLE PRECISION DEFAULT 0.01,
                target_cdf DOUBLE PRECISION DEFAULT 0.95,
                timestamp TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        ''')

        # Auto-Migración: Añadir columnas si no existen (Para bases de datos ya creadas)
        cols_to_add = [
            ("fecha_semana", "DATE" if is_postgres else "DATE"),
            ("uploaded_by", "VARCHAR(100)" if is_postgres else "TEXT")
        ]
        
        for col_name, col_type in cols_to_add:
            try:
                if is_postgres:
                    cursor.execute(f"ALTER TABLE scorecards ADD COLUMN IF NOT EXISTS {col_name} {col_type}")
                else:
                    # SQLite no soporta ADD COLUMN IF NOT EXISTS de forma directa
                    cursor.execute(f"PRAGMA table_info(scorecards)")
                    existing_cols = [c[1] for c in cursor.fetchall()]
                    if col_name not in existing_cols:
                        cursor.execute(f"ALTER TABLE scorecards ADD COLUMN {col_name} {col_type}")
            except Exception as e:
                logger.warning(f"Aviso migración columna {col_name}: {e}")
        
        # Auto-Migración tabla users: Añadir columnas nuevas si no existen
        user_cols_to_add = [
            ("must_change_password", "INTEGER DEFAULT 0"),
            ("centro_asignado",      "VARCHAR(100) DEFAULT NULL" if is_postgres else "TEXT DEFAULT NULL"),
        ]
        for _ucol, _utype in user_cols_to_add:
            try:
                if is_postgres:
                    cursor.execute(f"ALTER TABLE users ADD COLUMN IF NOT EXISTS {_ucol} {_utype}")
                else:
                    cursor.execute("PRAGMA table_info(users)")
                    existing_user_cols = [c[1] for c in cursor.fetchall()]
                    if _ucol not in existing_user_cols:
                        cursor.execute(f"ALTER TABLE users ADD COLUMN {_ucol} {_utype}")
            except Exception as _ue:
                logger.warning(f"Migración users col {_ucol}: {_ue}")



        # ── TABLAS NUEVAS v3.2: station_scorecards y wh_exceptions ────────────
        # 4. station_scorecards
        if is_postgres:
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS station_scorecards (
                    id SERIAL PRIMARY KEY,
                    semana VARCHAR(10) NOT NULL, fecha_semana DATE, anio INTEGER, centro VARCHAR(20) NOT NULL,
                    overall_score DOUBLE PRECISION, overall_standing VARCHAR(20),
                    rank_station INTEGER, rank_wow INTEGER,
                    safety_tier VARCHAR(20), fico DOUBLE PRECISION, fico_tier VARCHAR(20),
                    speeding_rate DOUBLE PRECISION, speeding_tier VARCHAR(20),
                    mentor_adoption DOUBLE PRECISION, mentor_tier VARCHAR(20),
                    vsa_compliance DOUBLE PRECISION, vsa_tier VARCHAR(20), boc VARCHAR(100),
                    whc_pct DOUBLE PRECISION, whc_tier VARCHAR(20), cas VARCHAR(50),
                    quality_tier VARCHAR(20), dcr_pct DOUBLE PRECISION, dcr_tier VARCHAR(20),
                    dnr_dpmo DOUBLE PRECISION, dnr_tier VARCHAR(20),
                    lor_dpmo DOUBLE PRECISION, lor_tier VARCHAR(20),
                    dsc_dpmo DOUBLE PRECISION, dsc_tier VARCHAR(20),
                    pod_pct DOUBLE PRECISION, pod_tier VARCHAR(20),
                    cc_pct DOUBLE PRECISION, cc_tier VARCHAR(20),
                    ce_dpmo DOUBLE PRECISION, ce_tier VARCHAR(20),
                    cdf_dpmo DOUBLE PRECISION, cdf_tier VARCHAR(20),
                    capacity_tier VARCHAR(20),
                    capacity_next_day DOUBLE PRECISION, capacity_next_day_tier VARCHAR(20),
                    capacity_same_day DOUBLE PRECISION, capacity_same_day_tier VARCHAR(20),
                    focus_area_1 VARCHAR(200), focus_area_2 VARCHAR(200), focus_area_3 VARCHAR(200),
                    uploaded_by VARCHAR(100), timestamp TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    UNIQUE(semana, centro, anio)
                )
            ''')
            for idx in ["CREATE INDEX IF NOT EXISTS idx_ss_centro_semana ON station_scorecards (centro, semana)",
                        "CREATE INDEX IF NOT EXISTS idx_ss_fecha ON station_scorecards (fecha_semana DESC)",
                        "CREATE INDEX IF NOT EXISTS idx_ss_standing ON station_scorecards (overall_standing)"]:
                try: cursor.execute(idx)
                except Exception: pass
        else:
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS station_scorecards (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    semana TEXT NOT NULL, fecha_semana DATE, anio INTEGER, centro TEXT NOT NULL,
                    overall_score FLOAT, overall_standing TEXT, rank_station INTEGER, rank_wow INTEGER,
                    safety_tier TEXT, fico FLOAT, fico_tier TEXT,
                    speeding_rate FLOAT, speeding_tier TEXT,
                    mentor_adoption FLOAT, mentor_tier TEXT,
                    vsa_compliance FLOAT, vsa_tier TEXT, boc TEXT,
                    whc_pct FLOAT, whc_tier TEXT, cas TEXT,
                    quality_tier TEXT, dcr_pct FLOAT, dcr_tier TEXT,
                    dnr_dpmo FLOAT, dnr_tier TEXT, lor_dpmo FLOAT, lor_tier TEXT,
                    dsc_dpmo FLOAT, dsc_tier TEXT, pod_pct FLOAT, pod_tier TEXT,
                    cc_pct FLOAT, cc_tier TEXT, ce_dpmo FLOAT, ce_tier TEXT,
                    cdf_dpmo FLOAT, cdf_tier TEXT,
                    capacity_tier TEXT, capacity_next_day FLOAT, capacity_next_day_tier TEXT,
                    capacity_same_day FLOAT, capacity_same_day_tier TEXT,
                    focus_area_1 TEXT, focus_area_2 TEXT, focus_area_3 TEXT,
                    uploaded_by TEXT, timestamp TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    UNIQUE(semana, centro, anio)
                )
            ''')
            for idx in ["CREATE INDEX IF NOT EXISTS idx_ss_centro_semana ON station_scorecards (centro, semana)",
                        "CREATE INDEX IF NOT EXISTS idx_ss_fecha ON station_scorecards (fecha_semana DESC)"]:
                try: cursor.execute(idx)
                except Exception: pass

        # 5. wh_exceptions
        wh_pk = "SERIAL PRIMARY KEY" if is_postgres else "INTEGER PRIMARY KEY AUTOINCREMENT"
        wh_str = "VARCHAR(10)" if is_postgres else "TEXT"
        wh_cen = "VARCHAR(20)" if is_postgres else "TEXT"
        wh_did = "VARCHAR(50)" if is_postgres else "TEXT"
        wh_uby = "VARCHAR(100)" if is_postgres else "TEXT"
        cursor.execute(f'''
            CREATE TABLE IF NOT EXISTS wh_exceptions (
                id {wh_pk},
                semana {wh_str} NOT NULL, fecha_semana DATE, anio INTEGER, centro {wh_cen} NOT NULL,
                driver_id {wh_did} NOT NULL,
                driver_name {wh_uby},
                daily_limit_exceeded INTEGER DEFAULT 0,
                weekly_limit_exceeded INTEGER DEFAULT 0,
                under_offwork_limit INTEGER DEFAULT 0,
                workday_limit_exceeded INTEGER DEFAULT 0,
                uploaded_by {wh_uby},
                timestamp TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                UNIQUE(semana, centro, anio, driver_id)
            )
        ''')
        for idx in ["CREATE INDEX IF NOT EXISTS idx_wh_centro_semana ON wh_exceptions (centro, semana)",
                    "CREATE INDEX IF NOT EXISTS idx_wh_driver ON wh_exceptions (driver_id)"]:
            try: cursor.execute(idx)
            except Exception: pass

        # ── MIGRACIÓN v3.2: columnas _oficial en scorecards (solo UPDATE desde PDF) ──
        pdf_cols_to_add = [
            ("entregados_oficial", "DOUBLE PRECISION" if is_postgres else "FLOAT"),
            ("dcr_oficial",        "DOUBLE PRECISION" if is_postgres else "FLOAT"),
            ("pod_oficial",        "DOUBLE PRECISION" if is_postgres else "FLOAT"),
            ("cc_oficial",         "DOUBLE PRECISION" if is_postgres else "FLOAT"),
            ("dsc_dpmo",           "DOUBLE PRECISION" if is_postgres else "FLOAT"),
            ("lor_dpmo",           "DOUBLE PRECISION" if is_postgres else "FLOAT"),
            ("ce_dpmo",            "DOUBLE PRECISION" if is_postgres else "FLOAT"),
            ("cdf_dpmo_oficial",   "DOUBLE PRECISION" if is_postgres else "FLOAT"),
            ("pdf_loaded",         "INTEGER DEFAULT 0"),
        ]
        for col_name, col_type in pdf_cols_to_add:
            try:
                if is_postgres:
                    cursor.execute(f"ALTER TABLE scorecards ADD COLUMN IF NOT EXISTS {col_name} {col_type}")
                else:
                    cursor.execute("PRAGMA table_info(scorecards)")
                    existing = [c[1] for c in cursor.fetchall()]
                    if col_name not in existing:
                        cursor.execute(f"ALTER TABLE scorecards ADD COLUMN {col_name} {col_type}")
            except Exception as e:
                logger.warning(f"Migración v3.2 col {col_name}: {e}")

        # ── MIGRACIÓN v3.9: driver_name en wh_exceptions + año en scorecards ─
        bi_cols_wh = [
            ("driver_name", "VARCHAR(255)" if is_postgres else "TEXT"),
        ]
        for col_name, col_type in bi_cols_wh:
            try:
                if is_postgres:
                    cursor.execute(f"ALTER TABLE wh_exceptions ADD COLUMN IF NOT EXISTS {col_name} {col_type}")
                else:
                    cursor.execute("PRAGMA table_info(wh_exceptions)")
                    existing = [c[1] for c in cursor.fetchall()]
                    if col_name not in existing:
                        cursor.execute(f"ALTER TABLE wh_exceptions ADD COLUMN {col_name} {col_type}")
            except Exception as e:
                logger.warning(f"Migración v3.9 wh_exceptions col {col_name}: {e}")

        # ── MIGRACIÓN v3.9b: columna 'anio' en scorecards y wh_exceptions ───
        # INTEGER con el año ISO extraído de fecha_semana.
        # Permite filtrar por año en Power BI sin parsear fechas.
        for _tbl in ("scorecards", "wh_exceptions"):
            try:
                if is_postgres:
                    cursor.execute(f"ALTER TABLE {_tbl} ADD COLUMN IF NOT EXISTS anio INTEGER")
                else:
                    cursor.execute(f"PRAGMA table_info({_tbl})")
                    if 'anio' not in [c[1] for c in cursor.fetchall()]:
                        cursor.execute(f"ALTER TABLE {_tbl} ADD COLUMN anio INTEGER")
            except Exception as _e:
                logger.warning(f"Migración v3.9b anio en {_tbl}: {_e}")

        # station_scorecards: añadir columna anio Y actualizar UNIQUE constraint
        # (el constraint antiguo era UNIQUE(semana,centro); el INSERT usa ON CONFLICT(semana,centro,anio))
        try:
            if is_postgres:
                cursor.execute("""
                    DO $$
                    BEGIN
                        -- 1. Añadir columna si no existe
                        IF NOT EXISTS (
                            SELECT 1 FROM information_schema.columns
                            WHERE table_name = 'station_scorecards' AND column_name = 'anio'
                        ) THEN
                            ALTER TABLE station_scorecards ADD COLUMN anio INTEGER;
                        END IF;

                        -- 2. Eliminar constraint antiguo sin anio (si existe con ese nombre)
                        IF EXISTS (
                            SELECT 1 FROM information_schema.table_constraints
                            WHERE table_name = 'station_scorecards'
                              AND constraint_type = 'UNIQUE'
                              AND constraint_name = 'station_scorecards_semana_centro_key'
                        ) THEN
                            ALTER TABLE station_scorecards
                                DROP CONSTRAINT station_scorecards_semana_centro_key;
                        END IF;

                        -- 3. Añadir constraint con anio si no existe aún
                        IF NOT EXISTS (
                            SELECT 1 FROM information_schema.table_constraints
                            WHERE table_name = 'station_scorecards'
                              AND constraint_type = 'UNIQUE'
                              AND constraint_name = 'station_scorecards_semana_centro_anio_key'
                        ) THEN
                            ALTER TABLE station_scorecards
                                ADD CONSTRAINT station_scorecards_semana_centro_anio_key
                                UNIQUE (semana, centro, anio);
                        END IF;
                    END $$;
                """)
            else:
                cursor.execute("PRAGMA table_info(station_scorecards)")
                if 'anio' not in [c[1] for c in cursor.fetchall()]:
                    cursor.execute("ALTER TABLE station_scorecards ADD COLUMN anio INTEGER")
        except Exception as _e:
            logger.warning(f"Migración v3.9b station_scorecards anio+constraint: {_e}")

        # Rellenar anio para filas existentes que tengan fecha_semana
        for _tbl in ("scorecards", "wh_exceptions", "station_scorecards"):
            try:
                if is_postgres:
                    cursor.execute(
                        f"UPDATE {_tbl} SET anio = EXTRACT(YEAR FROM fecha_semana::date) "
                        f"WHERE anio IS NULL AND fecha_semana IS NOT NULL"
                    )
                else:
                    cursor.execute(
                        f"UPDATE {_tbl} SET anio = CAST(STRFTIME('%Y', fecha_semana) AS INTEGER) "
                        f"WHERE anio IS NULL AND fecha_semana IS NOT NULL"
                    )
            except Exception as _e:
                logger.warning(f"Migración v3.9b relleno anio {_tbl}: {_e}")

        # Commit intermedio: los cambios DDL quedan persistidos aunque falle
        # algún paso posterior (bootstrap de usuario, etc.)
        try:
            conn.commit()
        except Exception as _e:
            logger.warning(f"Commit intermedio migración v3.9b: {_e}")

        # ── 6. Tabla de Rate Limiting persistente (v3.5) ────────────────────
        # Necesaria para que el bloqueo de login sobreviva reinicios del servidor
        # y funcione correctamente en entornos multi-worker (Streamlit Cloud).
        la_pk  = "SERIAL PRIMARY KEY" if is_postgres else "INTEGER PRIMARY KEY AUTOINCREMENT"
        la_str = "VARCHAR(100)" if is_postgres else "TEXT"
        cursor.execute(f"""
            CREATE TABLE IF NOT EXISTS login_attempts (
                id {la_pk},
                username {la_str} UNIQUE NOT NULL,
                attempt_count INTEGER DEFAULT 0,
                locked_until TIMESTAMP,
                last_attempt TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        """)
        try:
            if is_postgres:
                cursor.execute("CREATE INDEX IF NOT EXISTS idx_la_username ON login_attempts (LOWER(username))")
            else:
                cursor.execute("CREATE INDEX IF NOT EXISTS idx_la_username ON login_attempts (username)")
        except Exception as _e:
            logger.debug(f"init_database non-critical: {_e}")
            pass

        # Asegurar que el usuario superadmin inicial existe — configurable por env vars
        import os as _os
        _default_user = _os.environ.get("QS_ADMIN_USER")
        _default_pw   = _os.environ.get("QS_ADMIN_PASS")
        if not _default_user or not _default_pw:
            q_any_admin = ("SELECT COUNT(*) FROM users WHERE role IN ('admin','superadmin') AND active = 1"
                           if is_postgres else
                           "SELECT COUNT(*) FROM users WHERE role IN ('admin','superadmin') AND active = 1")
            cursor.execute(q_any_admin)
            _n_admins = cursor.fetchone()[0]
            if _n_admins > 0:
                logger.warning("QS_ADMIN_USER/PASS no configuradas — se omite bootstrap (ya existen admins en BD).")
            else:
                raise RuntimeError(
                    "QS_ADMIN_USER y QS_ADMIN_PASS son obligatorias. "
                    "Define estas variables de entorno antes de arrancar la app."
                )
        else:
            q_check = "SELECT id FROM users WHERE LOWER(username) = %s" if is_postgres else "SELECT id FROM users WHERE LOWER(username) = ?"
            cursor.execute(q_check, (_default_user.lower(),))
            if not cursor.fetchone():
                admin_pass = hash_password(_default_pw)
                q_ins = "INSERT INTO users (username, password, role, must_change_password) VALUES (%s, %s, %s, %s)" if is_postgres else "INSERT INTO users (username, password, role, must_change_password) VALUES (?, ?, ?, ?)"
                cursor.execute(q_ins, (_default_user, admin_pass, "superadmin", 1))
                logger.info(f"Usuario '{_default_user}' creado como SUPERADMIN (requiere cambio de contraseña).")
        
        conn.commit()
        return True
    except Exception as e:
        logger.error(f"Error DB: {e}")
        return False
    finally:
        try:
            if conn:
                _is_pg = db_config and db_config.get('type') == 'postgresql'
                if _is_pg and _PG_POOL and not _PG_POOL.closed:
                    _PG_POOL.putconn(conn)  # devolver al pool, no cerrar permanentemente
                else:
                    conn.close()
        except Exception:
            pass


# ═══════════════════════════════════════════════════════════════
# RATE LIMITING PERSISTENTE (Supabase/SQLite-compatible)
# ═══════════════════════════════════════════════════════════════
# Persiste entre workers de Streamlit Cloud y entre reinicios del servidor.
# Usa la tabla `login_attempts` (creada en init_database).

def record_login_attempt(username: str, success: bool, db_config: Optional[Dict] = None,
                          max_attempts: int = 5, lockout_minutes: int = 15) -> None:
    """
    Registra un intento de login (exitoso o fallido).
    - Si es exitoso, limpia los contadores del usuario.
    - Si es fallido, incrementa el contador y bloquea si se supera max_attempts.
    """
    try:
        with db_connection(db_config) as conn:
            cursor = conn.cursor()
            is_pg  = db_config and db_config.get('type') == 'postgresql'
            ph     = '%s' if is_pg else '?'
            now    = datetime.now()

            if success:
                cursor.execute(
                    f"DELETE FROM login_attempts WHERE LOWER(username) = {ph}",
                    (username.lower(),)
                )
            else:
                cursor.execute(
                    f"SELECT attempt_count FROM login_attempts WHERE LOWER(username) = {ph}",
                    (username.lower(),)
                )
                row = cursor.fetchone()
                new_count = (row[0] + 1) if row else 1
                locked_until = None
                if new_count >= max_attempts:
                    locked_until = (now + timedelta(minutes=lockout_minutes)).strftime("%Y-%m-%d %H:%M:%S")

                if is_pg:
                    cursor.execute(
                        """INSERT INTO login_attempts (username, attempt_count, locked_until, last_attempt)
                           VALUES (%s, %s, %s, %s)
                           ON CONFLICT (username) DO UPDATE SET
                               attempt_count = EXCLUDED.attempt_count,
                               locked_until  = EXCLUDED.locked_until,
                               last_attempt  = EXCLUDED.last_attempt""",
                        (username.lower(), new_count, locked_until, now.strftime("%Y-%m-%d %H:%M:%S"))
                    )
                else:
                    cursor.execute(
                        """INSERT OR REPLACE INTO login_attempts
                           (username, attempt_count, locked_until, last_attempt)
                           VALUES (?, ?, ?, ?)""",
                        (username.lower(), new_count, locked_until, now.strftime("%Y-%m-%d %H:%M:%S"))
                    )
            conn.commit()
    except Exception as e:
        logger.error(f"record_login_attempt error: {e}")


def check_login_locked(username: str, db_config: Optional[Dict] = None) -> Tuple[bool, int]:
    """
    Comprueba si un usuario está bloqueado.
    Returns: (bloqueado: bool, segundos_restantes: int)
    """
    try:
        with db_connection(db_config) as conn:
            cursor = conn.cursor()
            is_pg  = db_config and db_config.get('type') == 'postgresql'
            ph     = '%s' if is_pg else '?'
            cursor.execute(
                f"SELECT locked_until FROM login_attempts WHERE LOWER(username) = {ph}",
                (username.lower(),)
            )
            row = cursor.fetchone()

        if not row or not row[0]:
            return False, 0
        locked_until = datetime.strptime(str(row[0])[:19], "%Y-%m-%d %H:%M:%S")
        if datetime.now() < locked_until:
            remaining = int((locked_until - datetime.now()).total_seconds())
            return True, remaining
        return False, 0
    except Exception as e:
        logger.error(f"check_login_locked error: {e}")
        return False, 0

def reset_production_database(db_config: Optional[Dict] = None):
    """Limpia todos los datos de scorecards para empezar de cero (Mantiene usuarios y targets)"""
    try:
        is_pg = db_config and db_config.get('type') == 'postgresql'
        with db_connection(db_config) as conn:
            cursor = conn.cursor()
            if is_pg:
                cursor.execute("TRUNCATE TABLE scorecards RESTART IDENTITY CASCADE")
                cursor.execute("TRUNCATE TABLE station_scorecards RESTART IDENTITY CASCADE")
                cursor.execute("TRUNCATE TABLE wh_exceptions RESTART IDENTITY CASCADE")
            else:
                cursor.execute("DELETE FROM scorecards")
                cursor.execute("DELETE FROM station_scorecards")
                cursor.execute("DELETE FROM wh_exceptions")
            conn.commit()
        return True
    except Exception as e:
        logger.error(f"Error reset: {e}")
        return False


def week_to_date(week_str: str, year: int = None) -> str:
    """Convierte un string de semana 'W05' a la fecha del lunes de esa semana"""
    _fallback_year = year or datetime.now().year
    _fallback = f"{_fallback_year}-01-06"
    try:
        if not week_str or week_str == "N/A":
            return _fallback

        # Extraer número de semana
        match = re.search(r'(\d+)', week_str)
        if not match:
            return _fallback
            
        week_num = int(match.group(1))
        if not (1 <= week_num <= 53):
            return _fallback

        if year is None:
            year = datetime.now().year
            if week_num > 45:
                year -= 1
        # Cálculo ISO: 4 de enero es siempre semana 1
        d = datetime(year, 1, 4)
        # Retroceder al lunes de esa semana y saltar X semanas
        start_date = d - timedelta(days=d.weekday()) + timedelta(weeks=week_num-1)
        return start_date.strftime("%Y-%m-%d")
    except (ValueError, TypeError, AttributeError):
        return _fallback

def _safe_float(v, default=0.0):
    try: return float(v) if v is not None and str(v) not in ('nan','None','') else default
    except (ValueError, TypeError): return default


def save_to_database(df: pd.DataFrame, week: str, center: str, db_config: Optional[Dict] = None,
                     uploaded_by: str = "System", clean_first: bool = True,
                     year: Optional[int] = None) -> Tuple[bool, str]:
    """Guarda o actualiza los datos en la base de datos (SQLite o PostgreSQL)"""
    try:
        # Normalizar semana: W5 → W05, W9 → W09 (evita inconsistencias en ORDER BY y filtros)
        if week and week.startswith('W') and len(week) < 4:
            try:
                week = f"W{int(week[1:]):02d}"
            except ValueError:
                pass

        if clean_first:
            delete_scorecard_batch(week, center, db_config, year=year, preserve_pdf=True)

        is_postgres = db_config and db_config.get('type') == 'postgresql'
        ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        date_week = week_to_date(week, year=year)

        cols = [
            "semana", "fecha_semana", "anio", "centro", "driver_id", "driver_name", "calificacion", "score",
            "entregados", "dnr", "fs_count", "dnr_risk_events", "dcr", "pod", "cc",
            "fdps", "rts", "cdf", "detalles", "uploaded_by", "timestamp"
        ]

        placeholder = "%s" if is_postgres else "?"
        placeholders = ", ".join([placeholder] * len(cols))

        year_int = int(date_week[:4]) if date_week else None

        all_vals = [
            (
                week, date_week, year_int, center,
                str(row['ID']), str(row['Nombre']),
                str(row['CALIFICACION']), _safe_float(row['SCORE']),
                _safe_float(row['Entregados']), _safe_float(row['DNR']),
                _safe_float(row['FS_Count']), _safe_float(row['DNR_RISK_EVENTS']),
                _safe_float(row['DCR']), _safe_float(row['POD']),
                _safe_float(row['CC']), _safe_float(row['FDPS']),
                _safe_float(row['RTS']), _safe_float(row['CDF']),
                str(row['DETALLES']), uploaded_by, ts
            )
            for _, row in df.iterrows()
        ]

        with db_connection(db_config) as conn:
            cursor = conn.cursor()
            if is_postgres:
                query = f"""
                    INSERT INTO scorecards ({', '.join(cols)})
                    VALUES ({placeholders})
                    ON CONFLICT (semana, centro, anio, driver_id)
                    DO UPDATE SET
                        fecha_semana    = EXCLUDED.fecha_semana,
                        anio            = EXCLUDED.anio,
                        driver_name     = CASE
                                            WHEN EXCLUDED.driver_name != ''
                                             AND EXCLUDED.driver_name != EXCLUDED.driver_id
                                            THEN EXCLUDED.driver_name
                                            ELSE scorecards.driver_name
                                          END,
                        calificacion    = EXCLUDED.calificacion,
                        score           = EXCLUDED.score,
                        entregados      = CASE WHEN EXCLUDED.entregados > 0
                                               THEN EXCLUDED.entregados
                                               ELSE scorecards.entregados END,
                        dnr             = EXCLUDED.dnr,
                        fs_count        = EXCLUDED.fs_count,
                        dnr_risk_events = EXCLUDED.dnr_risk_events,
                        dcr             = CASE WHEN scorecards.pdf_loaded = 1
                                               THEN scorecards.dcr
                                               ELSE EXCLUDED.dcr END,
                        pod             = CASE WHEN scorecards.pdf_loaded = 1
                                               THEN scorecards.pod
                                               ELSE EXCLUDED.pod END,
                        cc              = CASE WHEN scorecards.pdf_loaded = 1
                                               THEN scorecards.cc
                                               ELSE EXCLUDED.cc  END,
                        fdps            = EXCLUDED.fdps,
                        rts             = EXCLUDED.rts,
                        cdf             = CASE WHEN scorecards.pdf_loaded = 1
                                               THEN scorecards.cdf
                                               ELSE EXCLUDED.cdf END,
                        detalles        = EXCLUDED.detalles,
                        uploaded_by     = EXCLUDED.uploaded_by,
                        timestamp       = EXCLUDED.timestamp
                """
                cursor.executemany(query, all_vals)
            else:
                # SQLite 3.24+: ON CONFLICT DO UPDATE (misma lógica que PostgreSQL)
                # En SQLite, la fila existente se referencia sin prefijo de tabla.
                query = f"""
                    INSERT INTO scorecards ({', '.join(cols)}) VALUES ({placeholders})
                    ON CONFLICT (semana, centro, anio, driver_id) DO UPDATE SET
                        fecha_semana    = excluded.fecha_semana,
                        anio            = excluded.anio,
                        driver_name     = CASE
                                            WHEN excluded.driver_name != ''
                                             AND excluded.driver_name != excluded.driver_id
                                            THEN excluded.driver_name
                                            ELSE driver_name
                                          END,
                        calificacion    = excluded.calificacion,
                        score           = excluded.score,
                        entregados      = CASE WHEN excluded.entregados > 0
                                               THEN excluded.entregados
                                               ELSE entregados END,
                        dnr             = excluded.dnr,
                        fs_count        = excluded.fs_count,
                        dnr_risk_events = excluded.dnr_risk_events,
                        dcr             = CASE WHEN pdf_loaded = 1 THEN dcr ELSE excluded.dcr END,
                        pod             = CASE WHEN pdf_loaded = 1 THEN pod ELSE excluded.pod END,
                        cc              = CASE WHEN pdf_loaded = 1 THEN cc  ELSE excluded.cc  END,
                        fdps            = excluded.fdps,
                        rts             = excluded.rts,
                        cdf             = CASE WHEN pdf_loaded = 1 THEN cdf ELSE excluded.cdf END,
                        detalles        = excluded.detalles,
                        uploaded_by     = excluded.uploaded_by,
                        timestamp       = excluded.timestamp
                """
                cursor.executemany(query, all_vals)
            # Recalcular scores para filas con pdf_loaded=1 (escenario PDF→CSV):
            # el UPSERT preservó dcr/pod/cc del PDF pero sobreescribió el score con
            # el valor provisional del CSV (calculado con dcr=1.0). Ahora recalculamos
            # usando los valores reales de la DB (dcr real + dnr/rts ya mergeados).
            if year_int:
                cursor.execute(
                    f"SELECT driver_id FROM scorecards "
                    f"WHERE semana={placeholder} AND centro={placeholder} "
                    f"AND anio={placeholder} AND pdf_loaded=1",
                    (week, center, year_int)
                )
                pdf_ids = [r[0] for r in cursor.fetchall()]
                if pdf_ids:
                    n_recalc = _recalculate_scores_for_ids(
                        cursor, placeholder, week, center, year_int, pdf_ids
                    )
                    if n_recalc:
                        logger.info(f"✓ save_to_database: {n_recalc} scores recalculados para filas PDF+CSV")
            conn.commit()

        logger.info(f"✅ {len(df)} registros sincronizados con DB ({'PostgreSQL' if is_postgres else 'SQLite'})")
        return True, ''
    except Exception as e:
        logger.error(f"Error guardando en DB: {str(e)}")
        return False, str(e)

def refresh_center_views(db_config=None) -> int:
    """
    Crea/actualiza vistas PostgreSQL por centro.
    Llamar solo desde Panel Admin — NO desde save_to_database.
    Devuelve el número de vistas creadas/actualizadas.
    """
    if not (db_config and db_config.get('type') == 'postgresql'):
        return 0
    try:
        with db_connection(db_config) as conn:
            cursor = conn.cursor()

            # Eliminar vistas antiguas con prefijo v_scorecard_
            cursor.execute("""
                SELECT table_name FROM information_schema.views
                WHERE table_schema = 'public'
                  AND table_name LIKE 'v_scorecard_%'
            """)
            old_views = [r[0] for r in cursor.fetchall()]
            for old_view in old_views:
                cursor.execute(sql.SQL('DROP VIEW IF EXISTS {}').format(sql.Identifier(old_view)))

            # Crear vistas nuevas con nombre DAS_{CENTRO}
            cursor.execute("SELECT DISTINCT centro FROM scorecards")
            centros = [r[0] for r in cursor.fetchall()]
            for c in centros:
                clean_name = "".join(ch if ch.isalnum() else "_" for ch in c.upper())[:50]
                view_name  = f"DAS_{clean_name}"
                cursor.execute(
                    sql.SQL('CREATE OR REPLACE VIEW {} AS SELECT * FROM scorecards WHERE centro = %s')
                    .format(sql.Identifier(view_name)), (c,)
                )
            conn.commit()
        logger.info(f"refresh_center_views: {len(centros)} vistas actualizadas (DAS_{{CENTRO}})")
        return len(centros)
    except Exception as e:
        logger.warning(f"refresh_center_views error: {e}")
        return 0


def run_maintenance(db_config: Optional[Dict] = None) -> Tuple[bool, int]:
    """
    Tarea de mantenimiento periódico — llamar desde panel Admin, no en cada save.
    1. Normaliza semanas (W5 → W05)
    2. Elimina duplicados físicos
    No se llama automáticamente en save_to_database() para evitar full-scan en cada guardado.
    """
    logger.info("⚙️ Iniciando mantenimiento de BD...")
    ok, removed = clean_database_duplicates(db_config)
    if ok:
        logger.info(f"⚙️ Mantenimiento completado: {removed} duplicados eliminados.")
    return ok, removed


def get_center_targets(center: str, db_config: Optional[Dict] = None) -> Dict:
    """Obtiene los targets guardados para un centro o los defaults"""
    try:
        with db_connection(db_config) as conn:
            cursor = conn.cursor()
            q = ("SELECT * FROM center_targets WHERE centro = %s"
                 if db_config and db_config.get('type') == 'postgresql'
                 else "SELECT * FROM center_targets WHERE centro = ?")
            cursor.execute(q, (center,))
            row = cursor.fetchone()
        if row:
            if db_config and db_config.get('type') == 'postgresql':
                return {
                    'centro': row[0], 'target_dnr': row[1], 'target_dcr': row[2],
                    'target_pod': row[3], 'target_cc': row[4], 'target_fdps': row[5],
                    'target_rts': row[6], 'target_cdf': row[7]
                }
            else:
                return dict(row)
    except Exception as e:
        logger.error(f"Error obteniendo targets: {e}")
    
    # Defaults si no existe o hay error — I-07: fuente única en Config.DEFAULT_TARGETS
    return {'centro': center, **Config.DEFAULT_TARGETS}

def save_center_targets(targets: Dict, db_config: Optional[Dict] = None):
    """Guarda o actualiza los targets para un centro"""
    try:
        cols = ['centro', 'target_dnr', 'target_dcr', 'target_pod', 'target_cc',
                'target_fdps', 'target_rts', 'target_cdf']
        vals = [targets[c] for c in cols]
        with db_connection(db_config) as conn:
            cursor = conn.cursor()
            is_postgres = db_config and db_config.get('type') == 'postgresql'
            if is_postgres:
                q = f"""
                    INSERT INTO center_targets ({', '.join(cols)}) VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
                    ON CONFLICT (centro) DO UPDATE SET
                    target_dnr=EXCLUDED.target_dnr, target_dcr=EXCLUDED.target_dcr,
                    target_pod=EXCLUDED.target_pod, target_cc=EXCLUDED.target_cc,
                    target_fdps=EXCLUDED.target_fdps, target_rts=EXCLUDED.target_rts,
                    target_cdf=EXCLUDED.target_cdf, timestamp=CURRENT_TIMESTAMP
                """
            else:
                q = f"INSERT OR REPLACE INTO center_targets ({', '.join(cols)}) VALUES (?, ?, ?, ?, ?, ?, ?, ?)"
            cursor.execute(q, vals)
            conn.commit()
        return True
    except Exception as e:
        logger.error(f"Error guardando targets: {e}")
        return False

def clean_database_duplicates(db_config: Optional[Dict] = None) -> Tuple[bool, int]:
    """
    1. Normaliza formatos de semana (ej: W5 -> W05)
    2. Elimina duplicados físicos que puedan haber quedado por versiones antiguas sin restricciones
    Llamar desde run_maintenance() o panel Admin, NO en save_to_database().
    """
    conn = None
    try:
        conn = get_db_connection(db_config)  # conexión explícita: función larga con commits intermedios
        cursor = conn.cursor()
        
        is_postgres = db_config and db_config.get('type') == 'postgresql'
        
        # 1. Normalizar Semanas (W5 -> W05)
        cursor.execute("SELECT id, semana FROM scorecards WHERE semana NOT LIKE 'W%' OR LENGTH(semana) < 3")
        rows = cursor.fetchall()
        
        updated = 0
        for r_id, sem in rows:
            match = re.search(r'(\d+)', str(sem))
            if match:
                new_sem = f"W{int(match.group(1)):02d}"
                try:
                    # Intentar actualizar
                    q_upd = "UPDATE scorecards SET semana = %s WHERE id = %s" if is_postgres else "UPDATE scorecards SET semana = ? WHERE id = ?"
                    cursor.execute(q_upd, (new_sem, r_id))
                    updated += 1
                except Exception as _dup_e:
                    # Si falla por UNIQUE constraint, borrar el antiguo mal nombrado
                    logger.debug(f"Semana normalización conflict id={r_id}: {_dup_e}")
                    q_del = "DELETE FROM scorecards WHERE id = %s" if is_postgres else "DELETE FROM scorecards WHERE id = ?"
                    cursor.execute(q_del, (r_id,))
                    updated += 1
        
        # 2. Eliminar duplicados físicos (por si acaso falló el UNIQUE en algún momento)
        # Mantenemos siempre el registro más reciente (el de ID más alto o timestamp mayor)
        if is_postgres:
            cursor.execute("""
                DELETE FROM scorecards 
                WHERE id IN (
                    SELECT id FROM (
                        SELECT id, ROW_NUMBER() OVER (PARTITION BY semana, centro, anio, driver_id ORDER BY timestamp DESC, id DESC) as row_num
                        FROM scorecards
                    ) t WHERE t.row_num > 1
                )
            """)
        else:
            cursor.execute("""
                DELETE FROM scorecards 
                WHERE id NOT IN (
                    SELECT MAX(id) FROM scorecards GROUP BY semana, centro, anio, driver_id
                )
            """)
        
        conn.commit()
        return True, updated
    except Exception as e:
        logger.error(f"clean_database_duplicates error: {e}")
        return False, 0
    finally:
        if conn:
            try:
                if db_config and db_config.get('type') == 'postgresql' and _PG_POOL and not _PG_POOL.closed:
                    _PG_POOL.putconn(conn)
                else:
                    conn.close()
            except Exception:
                pass

# NOTE: Procesamiento por lotes disponible en:
#   - App Streamlit: Tab Procesamiento → Importación masiva ZIP
#   - Directo:       process_single_batch() + save_to_database()
# La función main() CLI fue eliminada en v3.7 (dead code).


# ═══════════════════════════════════════════════════════════════════════════════
# DSP WEEKLY SCORECARD PDF — PARSING Y GUARDADO (v3.2)
# ═══════════════════════════════════════════════════════════════════════════════

def _parse_station_kpis(text: str, errors: list) -> dict:
    """
    Extrae los KPIs de estación de la página 2 del PDF.
    Probado contra el texto real extraído por pdfplumber del PDF oficial de Amazon.
    
    Formato real del PDF (verificado):
      'Safe Driving Metric (FICO) 831|Fantastic Vehicle Audit (VSA) Compliance 100%|Fantastic'
      'Working Hours Compliance (WHC) 86.36%|Poor'
      'Overall Score: 80.8 | Great'
      'Rank at DIC1: 2 ( 0 WoW)'
    """
    kpis = {}

    def find_val_tier(pattern):
        """Busca patrón, devuelve (float_value, tier_string) o (None, None)."""
        m = re.search(pattern, text, re.IGNORECASE)
        if not m:
            return None, None
        val_str = m.group(1).replace('%', '').strip()
        tier    = m.group(2).strip() if len(m.groups()) >= 2 else None
        try:
            return float(val_str), tier
        except (ValueError, TypeError):
            return None, tier

    # ── Overall Score y Standing ────────────────────────────────────────────
    m = re.search(r'Overall Score:\s*([\d.]+)\s*\|\s*(Fantastic|Great|Fair|Poor)', text)
    if m:
        kpis['overall_score']    = float(m.group(1))
        kpis['overall_standing'] = m.group(2)
    else:
        errors.append("overall_score")

    # ── Rank ────────────────────────────────────────────────────────────────
    # Formato real: "Rank at DIC1: 2 ( 0 WoW)"
    m = re.search(r'Rank at \w+:\s*(\d+)\s*\(\s*([+-]?\d+)\s*WoW\)', text)
    if m:
        kpis['rank_station'] = int(m.group(1))
        kpis['rank_wow']     = int(m.group(2))
    else:
        errors.append("rank_station")

    # ── Tiers de categorías ──────────────────────────────────────────────────
    m = re.search(r'Compliance and Safety\s+(Fantastic|Great|Fair|Poor)', text)
    kpis['safety_tier'] = m.group(1) if m else None

    m = re.search(r'Delivery Quality[^:]*:\s*(Fantastic|Great|Fair|Poor)', text)
    kpis['quality_tier'] = m.group(1) if m else None

    m = re.search(r'^Capacity:\s*(Fantastic|Great|Fair|Poor)', text, re.MULTILINE)
    kpis['capacity_tier'] = m.group(1) if m else None

    # ── Safety: FICO ─────────────────────────────────────────────────────────
    # Formato: "Safe Driving Metric (FICO) 831|Fantastic"
    kpis['fico'], kpis['fico_tier'] = find_val_tier(
        r'(?:Safe Driving Metric|FICO)[^0-9\n]+([\d.]+)\|(Fantastic|Great|Fair|Poor)')

    # ── Speeding ─────────────────────────────────────────────────────────────
    # Formato: "Speeding Event Rate (Per 100 Trips) 0|Fantastic"
    kpis['speeding_rate'], kpis['speeding_tier'] = find_val_tier(
        r'Speeding Event Rate[^\n]+?(\b[\d.]+)\|(Fantastic|Great|Fair|Poor)')

    # ── Mentor Adoption ──────────────────────────────────────────────────────
    # Formato: "Mentor Adoption Rate 100%|Fantastic"
    kpis['mentor_adoption'], kpis['mentor_tier'] = find_val_tier(
        r'Mentor Adoption Rate\s+([\d.]+)%?\|(Fantastic|Great|Fair|Poor)')

    # ── VSA ──────────────────────────────────────────────────────────────────
    # Formato: "Vehicle Audit (VSA) Compliance 100%|Fantastic"
    kpis['vsa_compliance'], kpis['vsa_tier'] = find_val_tier(
        r'Vehicle Audit.*?Compliance\s+([\d.]+)%?\|(Fantastic|Great|Fair|Poor)')

    # ── BOC ──────────────────────────────────────────────────────────────────
    # Formato: "Breach of Contract (BOC) None"
    m = re.search(r'Breach of Contract[^\n]*?(None|Yes)\b', text, re.IGNORECASE)
    kpis['boc'] = m.group(1) if m else None

    # ── WHC ──────────────────────────────────────────────────────────────────
    # Formato: "Working Hours Compliance (WHC) 86.36%|Poor"
    kpis['whc_pct'], kpis['whc_tier'] = find_val_tier(
        r'Working Hours Compliance[^\n]+?([\d.]+)%?\|(Fantastic|Great|Fair|Poor)')
    if kpis.get('whc_pct') is None:
        errors.append("whc_pct")

    # ── CAS ──────────────────────────────────────────────────────────────────
    # Formato: "Comprehensive Audit Score (CAS) In Compliance"
    m = re.search(r'Comprehensive Audit Score[^\n]*?(In Compliance|Non-Compliant)', text, re.IGNORECASE)
    kpis['cas'] = m.group(1) if m else None

    # ── DCR ──────────────────────────────────────────────────────────────────
    # Formato: "Delivery Completion Rate(DCR) 98.61%|Great"
    kpis['dcr_pct'], kpis['dcr_tier'] = find_val_tier(
        r'Delivery Completion Rate[^0-9\n]+([\d.]+)%?\|(Fantastic|Great|Fair|Poor)')
    if kpis.get('dcr_pct') is None:
        errors.append("dcr_pct")

    # ── DNR DPMO ─────────────────────────────────────────────────────────────
    # Formato: "Delivered Not Received(DNR DPMO) 1360|Great"
    kpis['dnr_dpmo'], kpis['dnr_tier'] = find_val_tier(
        r'Delivered Not Received[^0-9\n]+([\d.]+)\|(Fantastic|Great|Fair|Poor)')

    # ── LoR DPMO ─────────────────────────────────────────────────────────────
    # Formato: "Lost on Road (LoR) DPMO 111|Poor"
    kpis['lor_dpmo'], kpis['lor_tier'] = find_val_tier(
        r'Lost on Road[^0-9\n]+([\d.]+)\|(Fantastic|Great|Fair|Poor)')
    if kpis.get('lor_dpmo') is None:
        errors.append("lor_dpmo")

    # ── DSC DPMO ─────────────────────────────────────────────────────────────
    # Formato: "Delivery Success Conditions (DSC DPMO) 1134|Great"
    kpis['dsc_dpmo'], kpis['dsc_tier'] = find_val_tier(
        r'Delivery Success Conditions[^0-9\n]+([\d.]+)\|(Fantastic|Great|Fair|Poor)')

    # ── POD ──────────────────────────────────────────────────────────────────
    # Formato: "Photo-On-Delivery 97.51%|Fantastic"
    kpis['pod_pct'], kpis['pod_tier'] = find_val_tier(
        r'Photo.On.Delivery\s+([\d.]+)%?\|(Fantastic|Great|Fair|Poor)')

    # ── Contact Compliance ───────────────────────────────────────────────────
    # Formato: "Contact Compliance 97.72%|Great"
    kpis['cc_pct'], kpis['cc_tier'] = find_val_tier(
        r'Contact Compliance\s+([\d.]+)%?\|(Fantastic|Great|Fair|Poor)')

    # ── Customer Escalation DPMO ─────────────────────────────────────────────
    # Formato: "Customer escalation DPMO 0|Fantastic"
    kpis['ce_dpmo'], kpis['ce_tier'] = find_val_tier(
        r'Customer escalation DPMO\s+([\d.]+)\|(Fantastic|Great|Fair|Poor)')

    # ── CDF DPMO ─────────────────────────────────────────────────────────────
    # Formato: "Customer Delivery Feedback 1615|Fantastic"
    kpis['cdf_dpmo'], kpis['cdf_tier'] = find_val_tier(
        r'Customer Delivery Feedback\s+([\d.]+)\|(Fantastic|Great|Fair|Poor)')

    # ── Capacity ─────────────────────────────────────────────────────────────
    # Formato: "Next Day Capacity Reliability 145.91%|Fantastic"
    kpis['capacity_next_day'], kpis['capacity_next_day_tier'] = find_val_tier(
        r'Next Day Capacity Reliability\s+([\d.]+)%?\|(Fantastic|Great|Fair|Poor)')

    # Formato: "Same Day/Sub-Same Day Capacity Reliability 100%|Fantastic"
    kpis['capacity_same_day'], kpis['capacity_same_day_tier'] = find_val_tier(
        r'Same Day[^0-9\n]+([\d.]+)%?\|(Fantastic|Great|Fair|Poor)')

    # ── Focus Areas ──────────────────────────────────────────────────────────
    # Formato: "1. Lost on Road (LoR) DPMO\n2. Working Hours Compliance (WHC)\n3. ..."
    idx_focus = text.find('Recommended Focus Areas')
    if idx_focus >= 0:
        focus_block = text[idx_focus:]
        areas = re.findall(r'\d+\.\s+(.+?)(?=\n\d+\.|\nCurrent|$)', focus_block, re.DOTALL)
        kpis['focus_area_1'] = areas[0].strip() if len(areas) > 0 else None
        kpis['focus_area_2'] = areas[1].strip() if len(areas) > 1 else None
        kpis['focus_area_3'] = areas[2].strip() if len(areas) > 2 else None
    else:
        kpis['focus_area_1'] = kpis['focus_area_2'] = kpis['focus_area_3'] = None

    return kpis


def _build_drivers_df(all_rows: list, errors: list) -> pd.DataFrame:
    """
    Construye DataFrame de drivers desde filas extraídas de las páginas 3-5.
    
    Estructura real del PDF (verificada con pdfplumber):
      Página 3: fila 0 es cabecera ['Transporter ID','Delivered','DCR','DSC DPMO',
                'LoR DPMO','POD','CC','CE CDF DPMO', None] — CE y CDF en col[7] y col[8]
      Páginas 4-5: sin cabecera, empiezan directo en datos
    """
    def to_float(val):
        """Convierte valor a float; None si es '-', vacío o inválido."""
        if val in (None, '-', '', 'None'):
            return None
        try:
            return float(str(val).replace('%', '').strip())
        except (ValueError, TypeError):
            return None

    records = []
    header_skipped = False

    for row in all_rows:
        if not row or not row[0]:
            continue

        cell0 = str(row[0]).strip()

        # Saltar fila de cabecera (página 3)
        if not header_skipped and cell0 == 'Transporter ID':
            header_skipped = True
            continue

        # Saltar filas de título de sección (ej: "DSP WEEKLY SUMMARY")
        if not re.match(r'^[A-Z0-9]{10,20}$', cell0):
            continue

        # DCR, POD, CC vienen como "98.96%" → convertir a ratio 0-1
        # usando la función ya existente safe_percentage()
        raw_dcr = row[2] if len(row) > 2 else None
        raw_pod = row[5] if len(row) > 5 else None
        raw_cc  = row[6] if len(row) > 6 else None

        records.append({
            'driver_id':          cell0,
            'entregados_oficial': to_float(row[1] if len(row) > 1 else None),
            'dcr_oficial':        safe_percentage(raw_dcr),     # → 0-1
            'dsc_dpmo':           to_float(row[3] if len(row) > 3 else None),
            'lor_dpmo':           to_float(row[4] if len(row) > 4 else None),
            'pod_oficial':        safe_percentage(raw_pod),     # → 0-1
            'cc_oficial':         safe_percentage(raw_cc),      # → 0-1
            'ce_dpmo':            to_float(row[7] if len(row) > 7 else None),
            'cdf_dpmo_oficial':   to_float(row[8] if len(row) > 8 else None),
        })

    df = pd.DataFrame(records)
    if df.empty:
        errors.append("tabla de drivers vacía")
    else:
        logger.info(f"✓ Drivers extraídos del PDF: {len(df)}")
    return df


def _build_wh_df(rows: list, errors: list) -> pd.DataFrame:
    """
    Construye DataFrame de excepciones WHC desde página 6.
    
    Estructura real (verificada):
      row[0]: ['#','Transporter ID','Daily Limit Exceeded','Weekly Limit Exceeded',
               'Under Offwork Limit','Work Day Limit Exceeded','WH Exception']
      row[1..]: ['1','AOX3PX1MTVS0E','No','No','Yes','No','Yes']
    """
    def to_bool(val):
        return 1 if str(val or '').strip().lower() == 'yes' else 0

    records = []
    for row in rows:
        if not row or not row[0]:
            continue
        # Saltar cabecera
        if str(row[0]).strip() == '#':
            continue
        # Validar que sea fila de datos (primera col es número)
        try:
            int(str(row[0]).strip())
        except (ValueError, TypeError):
            continue

        if len(row) < 6:
            continue

        records.append({
            'driver_id':              str(row[1]).strip(),
            'daily_limit_exceeded':   to_bool(row[2]),
            'weekly_limit_exceeded':  to_bool(row[3]),
            'under_offwork_limit':    to_bool(row[4]),
            'workday_limit_exceeded': to_bool(row[5]),
        })

    df = pd.DataFrame(records)
    if df.empty:
        logger.info("PDF: sin excepciones WHC esta semana")
    else:
        logger.info(f"✓ WHC excepciones extraídas: {len(df)}")
    return df


def parse_dsp_scorecard_pdf(pdf_bytes: bytes) -> dict:
    """
    Función principal: extrae los 3 conjuntos de datos del DSP Weekly Scorecard PDF.

    Returns dict con claves:
      'ok'      → bool
      'meta'    → {'centro': str, 'semana': str, 'year': int}
      'station' → dict con todos los KPIs de estación
      'drivers' → pd.DataFrame con métricas oficiales por conductor
      'wh'      → pd.DataFrame con infracciones Working Hours
      'errors'  → list de campos no encontrados (no fatales)
    """
    if not HAS_PDFPLUMBER:
        return {'ok': False, 'errors': ['pdfplumber no instalado. Ejecuta: pip install pdfplumber'],
                'meta': {}, 'station': {}, 'drivers': pd.DataFrame(), 'wh': pd.DataFrame()}

    result = {
        'ok': False, 'meta': {}, 'station': {},
        'drivers': pd.DataFrame(), 'wh': pd.DataFrame(), 'errors': []
    }

    try:
        with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:

            if not pdf.pages:
                result['errors'].append("PDF sin páginas")
                return result

            # ── PÁGINA 1: Meta — centro, semana, año ────────────────────────
            p1 = pdf.pages[0].extract_text() or ""

            # "TDSL at DIC1" → DIC1
            m = re.search(r'TDSL\s+at\s+([A-Z]{2,5}\d+)', p1)
            centro = m.group(1).upper() if m else None

            # "Week 7" → W07
            m = re.search(r'Week\s+(\d+)', p1, re.IGNORECASE)
            if m and 1 <= int(m.group(1)) <= 53:
                semana = f"W{int(m.group(1)):02d}"
            else:
                semana = None

            # Año: Buscar en todo el PDF de forma agresiva
            year = None
            for p in pdf.pages:
                txt = p.extract_text() or ""
                # Buscar año explícito (20xx) con límites de palabra
                m = re.search(r'(?<!\d)(20[2-9]\d)(?!\d)', txt)
                if m:
                    year = int(m.group(1))
                    break

                # Búsqueda alternativa: fechas completas (ej: 02/15/2026 o 2026-02-15)
                m = re.search(r'\b(20[2-9]\d)[-/]\d{1,2}[-/]\d{1,2}\b', txt)
                if m:
                    year = int(m.group(1))
                    break
                m = re.search(r'\b\d{1,2}[-/]\d{1,2}[-/](20[2-9]\d)\b', txt)
                if m:
                    year = int(m.group(1))
                    break
            
            if not year:
                year = datetime.now().year
            
            if not centro:
                result['errors'].append("No se detectó el centro en el PDF")
                return result
            if not semana:
                result['errors'].append("No se detectó la semana en el PDF")
                return result

            result['meta'] = {'centro': centro, 'semana': semana, 'year': year}
            logger.info(f"PDF detectado: {centro} | {semana} | {year}")

            # ── PÁGINA 2: KPIs de estación ──────────────────────────────────
            if len(pdf.pages) > 1:
                p2_text = pdf.pages[1].extract_text() or ""
                result['station'] = _parse_station_kpis(p2_text, result['errors'])
            else:
                result['errors'].append("PDF sin página 2 (KPIs de estación)")

            # ── PÁGINAS 3-4: Tabla de drivers (índices 2 y 3) ────────────────
            # Página 3 (idx 2) tiene cabecera; página 4 (idx 3) solo datos
            all_driver_rows = []
            for page_idx in [2, 3]:
                if page_idx < len(pdf.pages):
                    tbl = pdf.pages[page_idx].extract_table()
                    if tbl:
                        all_driver_rows.extend(tbl)

            if all_driver_rows:
                result['drivers'] = _build_drivers_df(all_driver_rows, result['errors'])
            else:
                result['errors'].append("No se encontraron datos de drivers en páginas 3-4")

            # ── PÁGINA 5: Working Hours Exceptions (idx 4) ──────────────────
            if len(pdf.pages) > 4:
                tbl = pdf.pages[4].extract_table()
                if tbl:
                    result['wh'] = _build_wh_df(tbl, result['errors'])

            result['ok'] = True

    except Exception as e:
        result['errors'].append(f"Error al procesar PDF: {str(e)}")
        logger.error(f"parse_dsp_scorecard_pdf error: {e}")

    return result


def save_station_scorecard(station_data: dict, week: str, center: str,
                           db_config=None, uploaded_by: str = "System",
                           year: Optional[int] = None) -> Tuple[bool, str]:
    """
    Guarda o actualiza los KPIs de estación en station_scorecards.
    UPSERT por (semana, centro) — reemplaza si ya existe.
    """
    try:
        is_pg  = db_config and db_config.get('type') == 'postgresql'
        ph     = '%s' if is_pg else '?'
        fecha  = week_to_date(week, year=year)

        fields = [
            'semana', 'fecha_semana', 'anio', 'centro',
            'overall_score', 'overall_standing', 'rank_station', 'rank_wow',
            'safety_tier', 'fico', 'fico_tier', 'speeding_rate', 'speeding_tier',
            'mentor_adoption', 'mentor_tier', 'vsa_compliance', 'vsa_tier',
            'boc', 'whc_pct', 'whc_tier', 'cas',
            'quality_tier', 'dcr_pct', 'dcr_tier', 'dnr_dpmo', 'dnr_tier',
            'lor_dpmo', 'lor_tier', 'dsc_dpmo', 'dsc_tier',
            'pod_pct', 'pod_tier', 'cc_pct', 'cc_tier',
            'ce_dpmo', 'ce_tier', 'cdf_dpmo', 'cdf_tier',
            'capacity_tier', 'capacity_next_day', 'capacity_next_day_tier',
            'capacity_same_day', 'capacity_same_day_tier',
            'focus_area_1', 'focus_area_2', 'focus_area_3',
            'uploaded_by',
        ]

        anio_ss = int(fecha[:4]) if fecha else (year or datetime.now().year)

        vals = [
            week, fecha, anio_ss, center,
            station_data.get('overall_score'), station_data.get('overall_standing'),
            station_data.get('rank_station'), station_data.get('rank_wow'),
            station_data.get('safety_tier'),
            station_data.get('fico'), station_data.get('fico_tier'),
            station_data.get('speeding_rate'), station_data.get('speeding_tier'),
            station_data.get('mentor_adoption'), station_data.get('mentor_tier'),
            station_data.get('vsa_compliance'), station_data.get('vsa_tier'),
            station_data.get('boc'), station_data.get('whc_pct'), station_data.get('whc_tier'),
            station_data.get('cas'),
            station_data.get('quality_tier'),
            station_data.get('dcr_pct'), station_data.get('dcr_tier'),
            station_data.get('dnr_dpmo'), station_data.get('dnr_tier'),
            station_data.get('lor_dpmo'), station_data.get('lor_tier'),
            station_data.get('dsc_dpmo'), station_data.get('dsc_tier'),
            station_data.get('pod_pct'), station_data.get('pod_tier'),
            station_data.get('cc_pct'), station_data.get('cc_tier'),
            station_data.get('ce_dpmo'), station_data.get('ce_tier'),
            station_data.get('cdf_dpmo'), station_data.get('cdf_tier'),
            station_data.get('capacity_tier'),
            station_data.get('capacity_next_day'), station_data.get('capacity_next_day_tier'),
            station_data.get('capacity_same_day'), station_data.get('capacity_same_day_tier'),
            station_data.get('focus_area_1'), station_data.get('focus_area_2'),
            station_data.get('focus_area_3'),
            uploaded_by,
        ]

        col_list     = ', '.join(fields)
        placeholders = ', '.join([ph] * len(fields))

        with db_connection(db_config) as conn:
            cursor = conn.cursor()

            if is_pg:
                # Asegurar que la columna anio existe (puede faltar en BDs antiguas
                # si el usuario de la app no tiene permisos DDL — en ese caso hay
                # que ejecutar el ALTER TABLE manualmente en Supabase SQL Editor)
                try:
                    cursor.execute(
                        "SELECT 1 FROM information_schema.columns "
                        "WHERE table_name='station_scorecards' AND column_name='anio'"
                    )
                    anio_exists = cursor.fetchone() is not None
                except Exception:
                    anio_exists = False

                if anio_exists:
                    update_set = ', '.join(
                        f"{f} = EXCLUDED.{f}" for f in fields if f not in ('semana', 'centro', 'anio')
                    )
                    query = f"""
                        INSERT INTO station_scorecards ({col_list}) VALUES ({placeholders})
                        ON CONFLICT (semana, centro, anio) DO UPDATE SET {update_set}
                    """
                else:
                    # Fallback: insertar sin anio usando el constraint antiguo
                    fields_no_anio = [f for f in fields if f != 'anio']
                    vals_no_anio   = [v for f, v in zip(fields, vals) if f != 'anio']
                    col_list_fb    = ', '.join(fields_no_anio)
                    ph_fb          = ', '.join([ph] * len(fields_no_anio))
                    update_set_fb  = ', '.join(
                        f"{f} = EXCLUDED.{f}" for f in fields_no_anio if f not in ('semana', 'centro')
                    )
                    query = f"""
                        INSERT INTO station_scorecards ({col_list_fb}) VALUES ({ph_fb})
                        ON CONFLICT (semana, centro) DO UPDATE SET {update_set_fb}
                    """
                    vals = vals_no_anio
                    logger.warning(
                        "save_station_scorecard: columna 'anio' no existe en station_scorecards. "
                        "Ejecuta el SQL de migración en Supabase SQL Editor para añadirla."
                    )
            else:
                query = f"INSERT OR REPLACE INTO station_scorecards ({col_list}) VALUES ({placeholders})"

            cursor.execute(query, vals)
            conn.commit()

        logger.info(f"✓ station_scorecard guardado: {center} {week} | Score: {station_data.get('overall_score')} {station_data.get('overall_standing')}")
        return True, ''

    except Exception as e:
        logger.error(f"save_station_scorecard error: {e}")
        return False, str(e)


def _recalculate_scores_for_ids(cursor, ph: str, week: str, center: str,
                                year: int, driver_ids: list) -> int:
    """
    Re-calcula score/calificacion/detalles para un conjunto de driver_ids usando
    los valores actuales de la DB (dcr/pod/cc del PDF + dnr/rts del CSV).
    Se llama dentro de una transacción abierta (cursor ya activo).
    Devuelve el número de filas actualizadas.
    """
    if not driver_ids:
        return 0
    phs_ids = ", ".join([ph] * len(driver_ids))
    cursor.execute(
        f"SELECT driver_id, dnr, fs_count, dnr_risk_events, dcr, pod, cc, fdps, rts, cdf "
        f"FROM scorecards "
        f"WHERE semana={ph} AND centro={ph} AND anio={ph} AND driver_id IN ({phs_ids})",
        [week, center, year] + list(driver_ids)
    )
    rows = cursor.fetchall()
    score_updates = []
    for r in rows:
        did, dnr, fs_cnt, dnr_risk, dcr_r, pod_r, cc_r, fdps_r, rts_r, cdf_r = r
        fake = pd.Series({
            'DNR':             float(dnr      or 0),
            'FS_Count':        float(fs_cnt   or 0),
            'DNR_RISK_EVENTS': float(dnr_risk or 0),
            'DCR':             float(dcr_r    or 1),
            'POD':             float(pod_r    or 1),
            'CC':              float(cc_r     or 1),
            'FDPS':            float(fdps_r   or 1),
            'RTS':             float(rts_r    or 0),
            'CDF':             float(cdf_r    or 1),
        })
        calificacion, detalles, score_val = calculate_score_v3_robust(fake)
        score_updates.append((calificacion, float(score_val), detalles,
                               week, center, year, did))
    if score_updates:
        cursor.executemany(
            f"UPDATE scorecards SET calificacion={ph}, score={ph}, detalles={ph} "
            f"WHERE semana={ph} AND centro={ph} AND anio={ph} AND driver_id={ph}",
            score_updates
        )
    return len(score_updates)


def update_drivers_from_pdf(drivers_df: pd.DataFrame, week: str, center: str,
                             db_config=None, year: Optional[int] = None) -> Tuple[int, int]:
    """
    Actualiza las columnas _oficial y las columnas principales (dcr, pod, cc, entregados)
    en scorecards con los valores del PDF, y recalcula el score combinado (PDF+CSV).

    - NO elimina filas existentes
    - Para drivers ya existentes: actualiza _oficial + dcr/pod/cc + recalcula score
    - Para drivers nuevos (PDF subido antes del CSV): inserta fila seed con pdf_loaded=1

    Returns: (n_actualizados, n_no_encontrados)
    """
    if drivers_df is None or drivers_df.empty:
        return 0, 0

    try:
        with db_connection(db_config) as conn:
            cursor  = conn.cursor()
            is_pg   = db_config and db_config.get('type') == 'postgresql'
            ph      = '%s' if is_pg else '?'

            # ── Obtener driver_ids existentes de una sola query ───────────
            all_ids = drivers_df['driver_id'].astype(str).tolist()

            if year is None:
                year = datetime.now().year

            if not all_ids:
                return 0, 0

            phs = ", ".join([ph] * len(all_ids))
            cursor.execute(
                f"SELECT driver_id FROM scorecards "
                f"WHERE semana={ph} AND centro={ph} AND anio={ph} AND driver_id IN ({phs})",
                [week, center, year] + all_ids
            )
            existing_ids = {row[0] for row in cursor.fetchall()}

            not_found = [did for did in all_ids if did not in existing_ids]

            # ── Batch UPDATE para los que sí existen ─────────────────────
            update_vals = [
                (
                    row.get('entregados_oficial'), row.get('dcr_oficial'),
                    row.get('pod_oficial'),        row.get('cc_oficial'),
                    row.get('dsc_dpmo'),           row.get('lor_dpmo'),
                    row.get('ce_dpmo'),            row.get('cdf_dpmo_oficial'),
                    # columnas principales: los valores reales del PDF sobreescriben los defaults del CSV
                    row.get('dcr_oficial'), row.get('pod_oficial'), row.get('cc_oficial'),
                    row.get('entregados_oficial'), row.get('entregados_oficial'),  # CASE WHEN ? > 0 THEN ?
                    week, center, year, str(row['driver_id'])
                )
                for _, row in drivers_df.iterrows()
                if str(row['driver_id']) in existing_ids
            ]

            if update_vals:
                q_update = f"""
                    UPDATE scorecards SET
                        entregados_oficial = {ph},
                        dcr_oficial        = {ph},
                        pod_oficial        = {ph},
                        cc_oficial         = {ph},
                        dsc_dpmo           = {ph},
                        lor_dpmo           = {ph},
                        ce_dpmo            = {ph},
                        cdf_dpmo_oficial   = {ph},
                        dcr                = {ph},
                        pod                = {ph},
                        cc                 = {ph},
                        entregados         = CASE WHEN {ph} > 0 THEN {ph} ELSE entregados END,
                        pdf_loaded         = 1
                    WHERE semana={ph} AND centro={ph} AND anio={ph} AND driver_id={ph}
                """
                cursor.executemany(q_update, update_vals)
                updated = len(update_vals)
                # Recalcular score: ahora dcr/pod/cc son reales (PDF) y dnr/rts son reales (CSV)
                n_recalc = _recalculate_scores_for_ids(
                    cursor, ph, week, center, year, list(existing_ids)
                )
                if n_recalc:
                    logger.info(f"✓ update_drivers_from_pdf: {n_recalc} scores recalculados con datos combinados PDF+CSV")
            else:
                updated = 0

            # ── INSERT filas para drivers del PDF sin registro en scorecards ────────
            # Ocurre cuando el PDF se sube ANTES que el CSV de Concessions.
            # Se crea una fila con las métricas oficiales del PDF en las columnas
            # principales (dcr, pod, cc) y pdf_loaded=1.
            # El CSV de Concessions rellenará driver_name, dnr y rts después.
            if not_found:
                fecha_ins = week_to_date(week, year=year)
                anio_ins  = year if year else (int(fecha_ins[:4]) if fecha_ins else datetime.now().year)
                ts_ins    = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

                ins_cols = [
                    "semana", "fecha_semana", "anio", "centro", "driver_id", "driver_name",
                    "calificacion", "score",
                    "entregados", "dnr", "fs_count", "dnr_risk_events",
                    "dcr", "pod", "cc", "fdps", "rts", "cdf",
                    "detalles", "uploaded_by", "timestamp",
                    "entregados_oficial", "dcr_oficial", "pod_oficial", "cc_oficial",
                    "dsc_dpmo", "lor_dpmo", "ce_dpmo", "cdf_dpmo_oficial", "pdf_loaded",
                ]
                phs_ins = ', '.join([ph] * len(ins_cols))
                col_ins = ', '.join(ins_cols)

                not_found_set = set(not_found)
                insert_vals   = []
                for _, row in drivers_df.iterrows():
                    did = str(row['driver_id'])
                    if did not in not_found_set:
                        continue

                    dcr_v = float(row.get('dcr_oficial') or 1.0)
                    pod_v = float(row.get('pod_oficial') or 1.0)
                    cc_v  = float(row.get('cc_oficial')  or 1.0)
                    ent_v = float(row.get('entregados_oficial') or 0.0)

                    # Calcular score provisional con los datos del PDF (sin DNR/RTS aún)
                    fake = pd.Series({
                        'DNR': 0.0, 'FS_Count': 0.0, 'DNR_RISK_EVENTS': 0.0,
                        'DCR': dcr_v, 'POD': pod_v, 'CC': cc_v,
                        'FDPS': 1.0, 'RTS': 0.0, 'CDF': 1.0,
                    })
                    calificacion, detalles, score = calculate_score_v3_robust(fake)

                    insert_vals.append((
                        week, fecha_ins, anio_ins, center,
                        did, '',            # driver_name: vacío, el CSV lo rellenará
                        calificacion, float(score),
                        ent_v, 0, 0, 0,                 # dnr, fs_count, dnr_risk_events = 0
                        dcr_v, pod_v, cc_v, 1.0, 0.0, 1.0,  # dcr, pod, cc, fdps, rts, cdf
                        detalles, 'PDF', ts_ins,
                        # columnas _oficial
                        ent_v, dcr_v, pod_v, cc_v,
                        float(row.get('dsc_dpmo') or 0.0),
                        float(row.get('lor_dpmo') or 0.0),
                        float(row.get('ce_dpmo')  or 0.0),
                        float(row.get('cdf_dpmo_oficial') or 0.0),
                        1,  # pdf_loaded
                    ))

                if insert_vals:
                    if is_pg:
                        q_ins = f"""
                            INSERT INTO scorecards ({col_ins}) VALUES ({phs_ins})
                            ON CONFLICT (semana, centro, anio, driver_id) DO NOTHING
                        """
                    else:
                        q_ins = f"INSERT OR IGNORE INTO scorecards ({col_ins}) VALUES ({phs_ins})"
                    cursor.executemany(q_ins, insert_vals)
                    logger.info(f"✓ update_drivers_from_pdf: {len(insert_vals)} nuevas filas insertadas desde PDF")

            conn.commit()

        if not_found:
            logger.warning(f"Drivers del PDF sin match en scorecards ({len(not_found)}): "
                           f"{not_found[:5]}{'...' if len(not_found) > 5 else ''} — filas insertadas desde PDF")
        logger.info(f"✓ update_drivers_from_pdf: {updated} actualizados, {len(not_found)} sin match")
        return updated, len(not_found)

    except Exception as e:
        logger.error(f"update_drivers_from_pdf error: {e}")
        return 0, 0


def save_wh_exceptions(wh_df: pd.DataFrame, week: str, center: str,
                       db_config=None, uploaded_by: str = "System",
                       year: Optional[int] = None) -> bool:
    """
    Guarda las excepciones de Working Hours en wh_exceptions.
    Primero borra las del mismo centro+semana para evitar duplicados,
    luego inserta las nuevas.
    Hace lookup de driver_name en scorecards para que Power BI pueda
    filtrar por conductor sin necesitar un JOIN manual.
    """
    if wh_df is None or wh_df.empty:
        logger.info(f"Sin excepciones WHC para {center} {week}")
        return True

    try:
        with db_connection(db_config) as conn:
            cursor = conn.cursor()
            is_pg  = db_config and db_config.get('type') == 'postgresql'
            ph     = '%s' if is_pg else '?'
            fecha  = week_to_date(week, year=year)

            anio_wh = int(fecha[:4]) if fecha else (year or datetime.now().year)

            # ── Lookup driver_name desde scorecards (misma semana, centro y año) ──
            driver_ids = wh_df['driver_id'].astype(str).tolist()
            phs_list   = ', '.join([ph] * len(driver_ids))
            cursor.execute(
                f"SELECT driver_id, driver_name FROM scorecards "
                f"WHERE semana={ph} AND centro={ph} AND anio={ph} AND driver_id IN ({phs_list})",
                [week, center, anio_wh] + driver_ids
            )
            name_map = {str(r[0]): str(r[1]) for r in cursor.fetchall() if r[1]}

            cursor.execute(
                f"DELETE FROM wh_exceptions WHERE semana={ph} AND centro={ph} AND anio={ph}",
                (week, center, anio_wh)
            )

            wh_vals = [
                (
                    week, fecha, anio_wh, center,
                    str(row['driver_id']), name_map.get(str(row['driver_id'])),
                    int(row.get('daily_limit_exceeded', 0)),
                    int(row.get('weekly_limit_exceeded', 0)),
                    int(row.get('under_offwork_limit', 0)),
                    int(row.get('workday_limit_exceeded', 0)),
                    uploaded_by
                )
                for _, row in wh_df.iterrows()
            ]
            cursor.executemany(
                f"""INSERT INTO wh_exceptions
                    (semana, fecha_semana, anio, centro, driver_id, driver_name,
                     daily_limit_exceeded, weekly_limit_exceeded,
                     under_offwork_limit, workday_limit_exceeded, uploaded_by)
                    VALUES ({ph},{ph},{ph},{ph},{ph},{ph},{ph},{ph},{ph},{ph},{ph})
                """,
                wh_vals
            )
            conn.commit()
        n_named = sum(1 for r in wh_df['driver_id'] if str(r) in name_map)
        logger.info(f"✓ WHC exceptions guardadas: {len(wh_df)} para {center} {week} "
                    f"({n_named} con nombre, {len(wh_df)-n_named} solo ID)")
        return True

    except Exception as e:
        logger.error(f"save_wh_exceptions error: {e}")
        return False


def get_station_scorecards(db_config=None) -> pd.DataFrame:
    """
    Devuelve todos los station_scorecards ordenados por centro y semana desc.
    Incluye wh_count: número de drivers con excepción WHC esa semana/centro,
    obtenido via LEFT JOIN con wh_exceptions.
    """
    try:
        with db_connection(db_config) as conn:
            query = """
                SELECT ss.semana, ss.anio, ss.centro, ss.overall_score, ss.overall_standing,
                       ss.rank_station, ss.rank_wow,
                       ss.fico, ss.fico_tier, ss.whc_pct, ss.whc_tier,
                       ss.dcr_pct, ss.dcr_tier, ss.dnr_dpmo, ss.dnr_tier,
                       ss.lor_dpmo, ss.lor_tier, ss.dsc_dpmo, ss.dsc_tier,
                       ss.pod_pct, ss.pod_tier, ss.cc_pct, ss.cc_tier,
                       ss.ce_dpmo, ss.ce_tier, ss.cdf_dpmo, ss.cdf_tier,
                       ss.speeding_rate, ss.speeding_tier,
                       ss.mentor_adoption, ss.mentor_tier,
                       ss.vsa_compliance, ss.vsa_tier,
                       ss.boc, ss.cas,
                       ss.capacity_next_day, ss.capacity_next_day_tier,
                       ss.capacity_same_day, ss.capacity_same_day_tier,
                       ss.safety_tier, ss.quality_tier, ss.capacity_tier,
                       ss.focus_area_1, ss.focus_area_2, ss.focus_area_3,
                       ss.uploaded_by, ss.timestamp,
                       COALESCE(wh.wh_count, 0) AS wh_count
                FROM station_scorecards ss
                LEFT JOIN (
                    SELECT semana, centro, anio, COUNT(*) AS wh_count
                    FROM wh_exceptions
                    GROUP BY semana, centro, anio
                ) wh ON ss.semana = wh.semana AND ss.centro = wh.centro AND COALESCE(ss.anio, wh.anio) = wh.anio
                ORDER BY ss.centro ASC, ss.timestamp DESC
            """
            df = pd.read_sql_query(query, conn)
            df = df.loc[:, ~df.columns.duplicated()]  # fix: sqlite3.Row + LEFT JOIN puede duplicar cols
        return df
    except Exception as e:
        logger.error(f"get_station_scorecards error: {e}")
        return pd.DataFrame()


# ═══════════════════════════════════════════════════════════════
# HELPERS DE CENTRO ASIGNADO (restricción JT por centro v3.6)
# ═══════════════════════════════════════════════════════════════

def get_user_centro(username: str, db_config: Optional[Dict] = None) -> Optional[str]:
    """
    Devuelve el centro asignado a un JT, o None si no tiene restricción.
    None = ve todos los centros.
    """
    try:
        with db_connection(db_config) as conn:
            is_pg = db_config and db_config.get('type') == 'postgresql'
            ph = '%s' if is_pg else '?'
            cursor = conn.cursor()
            cursor.execute(
                f"SELECT centro_asignado FROM users WHERE LOWER(username) = {ph}",
                (username.lower(),)
            )
            row = cursor.fetchone()
            if row and row[0]:
                v = str(row[0]).strip().upper()
                return v if v else None
            return None
    except Exception as e:
        logger.error(f"get_user_centro error: {e}")
        return None


def set_user_centro(username: str, centro: Optional[str],
                    db_config: Optional[Dict] = None) -> bool:
    """
    Asigna (o quita) el centro a un usuario JT.
    centro=None → sin restricción (ve todos).
    centro='DIC1' → solo ve DIC1.
    """
    try:
        valor = centro.strip().upper() if centro and centro.strip() else None
        with db_connection(db_config) as conn:
            is_pg = db_config and db_config.get('type') == 'postgresql'
            ph = '%s' if is_pg else '?'
            cursor = conn.cursor()
            cursor.execute(
                f"UPDATE users SET centro_asignado = {ph} WHERE LOWER(username) = {ph}",
                (valor, username.lower())
            )
            conn.commit()
        logger.info(f"centro_asignado → '{username}': {valor!r}")
        return True
    except Exception as e:
        logger.error(f"set_user_centro error: {e}")
        return False


# ═══════════════════════════════════════════════════════════════
# ALERTAS POR EMAIL (v3.6)
# ═══════════════════════════════════════════════════════════════

import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart


def send_alert_email(smtp_cfg: Dict, to_email: str, subject: str, body_html: str) -> bool:
    """
    Envía un email de alerta vía SMTP.

    smtp_cfg esperado (desde st.secrets['smtp']):
        host, port, user, password, from_email

    Soporta TLS (port 587) y SSL (port 465).
    """
    try:
        msg = MIMEMultipart('alternative')
        msg['Subject'] = subject
        msg['From']    = smtp_cfg.get('from_email', smtp_cfg['user'])
        msg['To']      = to_email
        msg.attach(MIMEText(body_html, 'html', 'utf-8'))

        port = int(smtp_cfg.get('port', 587))
        host = smtp_cfg['host']

        if port == 465:
            with smtplib.SMTP_SSL(host, port) as srv:
                srv.login(smtp_cfg['user'], smtp_cfg['password'])
                srv.sendmail(msg['From'], [to_email], msg.as_string())
        else:
            with smtplib.SMTP(host, port) as srv:
                srv.ehlo()
                srv.starttls()
                srv.login(smtp_cfg['user'], smtp_cfg['password'])
                srv.sendmail(msg['From'], [to_email], msg.as_string())

        logger.info(f"✉️ Alerta enviada a {to_email}: {subject}")
        return True
    except Exception as e:
        logger.error(f"send_alert_email error: {e}")
        return False


def check_and_send_alerts(week: str, center: str,
                          smtp_cfg: Optional[Dict] = None,
                          alert_email: Optional[str] = None,
                          db_config: Optional[Dict] = None) -> int:
    """
    Comprueba conductores que llevan ≥2 semanas consecutivas en POOR
    y envía un email de alerta consolidado.

    Devuelve: número de conductores alertados.
    """
    if not smtp_cfg or not alert_email:
        logger.info("Alertas desactivadas (sin config SMTP o email destino).")
        return 0

    try:
        with db_connection(db_config) as conn:
            is_pg = db_config and db_config.get('type') == 'postgresql'
            ph = '%s' if is_pg else '?'

            # Conductores POOR en la semana actual
            cursor = conn.cursor()
            df_current = pd.read_sql_query(
                f"SELECT driver_id, driver_name, score, detalles "
                f"FROM scorecards WHERE centro = {ph} AND semana = {ph} "
                f"AND calificacion = '🛑 POOR'",
                conn, params=(center, week)
            )

            if df_current.empty:
                return 0

            # Semana inmediatamente anterior (la más reciente por fecha antes de esta)
            df_prev_meta = pd.read_sql_query(
                f"SELECT semana FROM scorecards "
                f"WHERE centro = {ph} "
                f"AND fecha_semana < (SELECT MIN(fecha_semana) FROM scorecards "
                f"                    WHERE centro = {ph} AND semana = {ph}) "
                f"GROUP BY semana ORDER BY fecha_semana DESC LIMIT 1",
                conn, params=(center, center, week)
            )

            if df_prev_meta.empty:
                return 0

            prev_week = df_prev_meta['semana'].iloc[0]

            df_prev_poor = pd.read_sql_query(
                f"SELECT driver_id FROM scorecards "
                f"WHERE centro = {ph} AND semana = {ph} AND calificacion = '🛑 POOR'",
                conn, params=(center, prev_week)
            )

        if df_prev_poor.empty:
            return 0

        # Intersección: POOR esta semana Y la anterior
        repeated_poor = df_current[
            df_current['driver_id'].isin(df_prev_poor['driver_id'])
        ]

        if repeated_poor.empty:
            return 0

        # Construir email HTML
        rows_html = ""
        for _, r in repeated_poor.iterrows():
            rows_html += (
                f"<tr>"
                f"<td style='padding:8px;border-bottom:1px solid #eee'>{_html_escape.escape(str(r['driver_name']))}</td>"
                f"<td style='padding:8px;border-bottom:1px solid #eee;color:#dc3545;font-weight:700'>{int(r['score'])}</td>"
                f"<td style='padding:8px;border-bottom:1px solid #eee;font-size:0.9em'>{_html_escape.escape(str(r['detalles']))}</td>"
                f"</tr>"
            )

        body = f"""
        <div style='font-family:Arial,sans-serif;max-width:700px;margin:0 auto'>
            <div style='background:#232f3e;color:white;padding:20px;border-radius:8px 8px 0 0'>
                <h2 style='margin:0'>🚨 Alerta de Calidad — {center} {week}</h2>
                <p style='margin:5px 0 0;opacity:.8'>
                    {len(repeated_poor)} conductor(es) en POOR durante 2 semanas consecutivas
                    ({prev_week} y {week})
                </p>
            </div>
            <div style='padding:20px;border:1px solid #dee2e6;border-top:none;border-radius:0 0 8px 8px'>
                <table style='width:100%;border-collapse:collapse'>
                    <thead>
                        <tr style='background:#f8f9fa'>
                            <th style='padding:10px;text-align:left'>Conductor</th>
                            <th style='padding:10px;text-align:left'>Score</th>
                            <th style='padding:10px;text-align:left'>Problemas</th>
                        </tr>
                    </thead>
                    <tbody>{rows_html}</tbody>
                </table>
                <p style='color:#6c757d;font-size:0.85em;margin-top:20px'>
                    Este mensaje ha sido generado automáticamente por Quality Scorecard.
                </p>
            </div>
        </div>
        """

        subject = f"🚨 [{center} {week}] {len(repeated_poor)} conductor(es) POOR 2 semanas consecutivas"
        send_alert_email(smtp_cfg, alert_email, subject, body)
        return len(repeated_poor)

    except Exception as e:
        logger.error(f"check_and_send_alerts error: {e}")
        return 0
